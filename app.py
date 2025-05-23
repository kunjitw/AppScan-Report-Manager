# -*- coding: utf-8 -*-

# --- 標準函式庫匯入 (Standard Library Imports) ---
import base64         # 用於圖片數據編碼 (Base64)
import glob           # 用於查找符合模式的檔案路徑名
import io             # 用於處理記憶體中的二進位數據流 (例如產生 Excel)
import json           # 用於處理 JSON 數據 (讀寫狀態檔、設定檔)
import logging        # 用於記錄程式運行訊息和錯誤
import os             # 用於與作業系統互動 (路徑操作、檔案系統)
import platform       # 用於獲取作業系統資訊 (例如判斷如何開啟檔案)
import re             # 用於正則表達式操作 (解析檔名、清理字串)
import signal         # 用於處理作業系統信號 (例如 Ctrl+C)
import subprocess     # 用於執行外部命令 (例如開啟 .scan 檔)
import sys            # 用於存取 Python 解譯器相關的變數和函式 (例如打包路徑)
import threading      # 用於實現多線程 (讓 Flask 伺服器和 GUI 同時運行)
import time           # 用於時間相關操作 (例如暫停)
import traceback      # 用於獲取和格式化錯誤追蹤訊息
import xml.etree.ElementTree as ET # 用於解析 XML 檔案 (AppScan 報告)
from datetime import datetime # 用於處理日期和時間 (例如產生時間戳)
from urllib.parse import ( # 用於 URL 解析和編碼
    unquote,          # URL 解碼
    urlparse,         # 解析 URL 結構
    quote             # URL 編碼
)
import urllib.error   # 處理 URL 請求錯誤
import urllib.request # 用於發送 HTTP 請求 (例如關閉伺服器)
import webbrowser     # 用於開啟網頁瀏覽器
import mimetypes      # 用於猜測檔案的 MIME 類型 (例如圖片)
import uuid           # 用於產生唯一的 ID (手動新增弱點)
import shutil         # 用於檔案操作 (例如移動檔案到垃圾桶)
from collections import defaultdict # 用於 JSON 匯出時方便建立巢狀字典

# --- 圖形介面函式庫匯入 (GUI Imports - Tkinter) ---
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# --- Excel 匯出函式庫匯入 (Excel Export Import - openpyxl) ---
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --- Web 框架函式庫匯入 (Web Framework Imports - Flask) ---
from flask import (
    Flask,            # Flask 應用程式主體
    render_template,  # 渲染 HTML 模板
    request,          # 處理傳入的請求數據
    jsonify,          # 將 Python dict 轉換為 JSON 回應
    abort,            # 中止請求並返回錯誤碼
    redirect,         # 重定向到其他 URL
    url_for,          # 產生 URL 路徑
    flash,            # 在網頁上顯示一次性訊息
    send_from_directory, # 從目錄安全地發送檔案
    send_file,        # 發送檔案作為回應
    Response          # 用於自訂 JSON 回應
)
from werkzeug.utils import secure_filename # 安全地處理用戶上傳的檔名

# --- Selenium 自動化瀏覽器匯入 (Selenium Imports) ---
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, NoSuchElementException, TimeoutException

# --- WebDriver 管理器 (可選，自動下載 ChromeDriver) ---
try:
    from webdriver_manager.chrome import ChromeDriverManager
    WEBDRIVER_MANAGER_AVAILABLE = True # 標記 webdriver-manager 是否可用
except ImportError:
    WEBDRIVER_MANAGER_AVAILABLE = False
    # 如果找不到 webdriver-manager，在控制台印出警告訊息
    print(
        "警告: 找不到 webdriver-manager。請安裝 (`pip install webdriver-manager`) "
        "或手動提供 chromedriver 的路徑。",
        file=sys.stderr,
    )

# --- Flask 應用程式初始化 ---
app = Flask(__name__) # 建立 Flask 應用程式實例
app.secret_key = os.urandom(24) # 設定 Session 使用的密鑰，用於 flash 訊息等功能

# --- 常數：路徑和檔名設定 (Constants: Paths and Filenames) ---
BASE_REPORT_FOLDER = "reports" # AppScan 報告 (.xml, .scan) 的根目錄
BASE_DATA_FOLDER = "data"      # 應用程式數據 (狀態、截圖、設定) 的根目錄
SCREENSHOTS_SUBFOLDER = "screenshots" # 儲存截圖的子目錄名稱
TRASH_SCREENSHOTS_SUBFOLDER = "_trash_screenshots" # 儲存已刪除截圖的子目錄名稱
STATUS_FILE_NAME = "vulnerability_status.json" # 儲存弱點狀態和筆記的檔名
RULES_FILE = "exclusion_rules.json" # 儲存通用排除規則的檔名
ADVISORY_FILE_NAME = "advisory.json" # 儲存弱點修補建議的檔名 (目前僅載入，未提供修改介面)
APP_CONFIG_FILE_NAME = "app_config.json" # 儲存專案特定設定 (例如顯示名稱) 的檔名
SERVER_CONFIG_FILE = "app_server_config.json" # 儲存伺服器設定 (例如 Port) 的檔名
TARGET_LIST_FILENAME = "target.xlsx" # 專案報告對應的目標清單檔名
WEAKNESS_LIST_FILE = os.path.join(BASE_DATA_FOLDER, "weakness_list.txt") # 預定義弱點名稱列表檔案路徑
COMMON_NOTES_FILE = "common_notes.json" # 新增：常用備註檔名

# --- 確保基礎數據資料夾存在 ---
if not os.path.exists(BASE_DATA_FOLDER):
    try:
        os.makedirs(BASE_DATA_FOLDER) # 建立基礎數據資料夾
        print(f"已建立基礎數據資料夾: {os.path.abspath(BASE_DATA_FOLDER)}")
    except OSError as e:
        # 如果建立失敗，印出嚴重錯誤並結束程式
        print(f"嚴重錯誤: 無法建立基礎數據資料夾 {BASE_DATA_FOLDER}: {e}", file=sys.stderr)
        sys.exit(1)

# --- 常數：顯示/邏輯相關設定 (Constants: Display/Logic related) ---
# 弱點狀態選項 (鍵值對，用於下拉選單和內部邏輯)
STATUS_OPTIONS = {
    "未審查": "未審查",
    "人工審查中": "人工審查中",
    "誤判": "誤判",
    "已確認弱點": "已確認弱點",
    "已自動排除": "已自動排除" # 此狀態由排除規則自動設定
}
AUTO_EXCLUDED_STATUS = "已自動排除" # 自動排除狀態的常數值
DEFAULT_STATUS = "未審查"           # 預設的弱點狀態
DEFAULT_NOTE = ""                   # 預設的筆記內容

# AppScan 嚴重性代碼到內部鍵值的映射 (XML 中的 severity ID)
SEVERITY_MAP = {
    "0": "informational",
    "1": "low",
    "2": "medium",
    "3": "high",
    "4": "critical"
}

# 內部嚴重性鍵值到中文顯示名稱的映射
SEVERITY_DISPLAY_MAP = {
    "informational": "參考資訊",
    "low": "低",
    "medium": "中",
    "high": "高",
    "critical": "重大"
}

# 嚴重性內部鍵值到數值等級的映射 (用於排序和比較)
SEVERITY_LEVELS = {
    "informational": 0,
    "low": 1,
    "medium": 2,
    "high": 3,
    "critical": 4
}

# 嚴重性摘要顯示的順序 (從高到低)
SEVERITY_SUMMARY_ORDER = ["critical", "high", "medium", "low", "informational"]

# AppScan 掃描狀態到中文顯示名稱的映射
SCAN_STATUS_MAP = {
    "Success": "成功",
    "Failed": "失敗",
    "Aborted": "已中斷",
    "Running": "執行中" # 理論上不會在靜態報告中看到
}
DEFAULT_SCAN_STATUS = "狀態未知"      # XML 中無法判斷時的預設狀態
PARSE_ERROR_STATUS = "解析錯誤"       # XML 解析失敗時的狀態
INCOMPLETE_STATUS = "格式不完整"     # XML 缺少必要標籤時的狀態
READ_ERROR_STATUS = "讀取錯誤"       # 讀取 XML 檔案失敗時的狀態
FILE_NOT_FOUND_STATUS = "檔案遺失"    # target.xlsx 中有但找不到對應 XML 檔的狀態
MISSING_FILE_STATUS = "Missing"      # 內部使用的缺失狀態標記 (似乎與 FILE_NOT_FOUND 重疊)
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "bmp"} # 允許上傳的截圖副檔名

# 截圖過濾器的選項 (用於前端下拉選單)
SCREENSHOT_FILTER_OPTIONS = {
    "completed": "已完成全部截圖",
    "not_completed": "未完成全部截圖",
    "has_files": "有截圖檔案",
    "no_files": "無截圖檔案"
}
REPORT_COMPLETED_KEY = "_report_review_completed" # 狀態檔中標記報告是否判讀完成的內部鍵名
EXTERNAL_LINK_REASONING = "AppScan 發現外部網站的鏈結，但無法解析它" # 特殊 AppScan 外部連結問題的原因文字
MANUAL_SOURCE_LABEL = "手動" # 手動新增弱點的來源標籤 (用於檔名和顯示)
APPSCAN_SOURCE_LABEL = "APPSCAN" # AppScan 來源弱點的標籤 (用於檔名和顯示)
MANUAL_ENTITY_TYPE_LABEL = "手動新增" # 手動新增弱點的固定實體類型標籤

# --- 全域變數與鎖 (Global variables & Locks) ---
project_file_locks = {}      # 用於儲存每個專案的檔案鎖 (例如 status.json, config.json)，防止多線程衝突
project_locks_lock = threading.Lock() # 用於保護 project_file_locks 字典本身的鎖
exclusion_rules = []         # 儲存全域排除規則的列表
rules_file_lock = threading.Lock() # 用於保護 exclusion_rules.json 讀寫的鎖
weakness_list_cache = None   # 快取弱點名稱列表，避免重複讀取檔案
weakness_list_lock = threading.Lock() # 保護弱點列表快取讀寫的鎖
common_notes_cache = None    # 新增：常用備註快取
common_notes_lock = threading.Lock() # 新增：常用備註鎖
server_thread = None         # 儲存 Flask 伺服器線程物件
server_port = 5001           # 預設的伺服器 Port
server_running = False       # 標記 Flask 伺服器是否正在運行
status_window_log_queue = [] # Tkinter GUI 顯示日誌的佇列
status_window_root = None    # Tkinter GUI 的根視窗物件
selenium_driver_lock = threading.Lock() # 保護 Selenium WebDriver 實例創建和使用的鎖
selenium_driver_instance = None # 儲存 Selenium WebDriver 的實例 (共用以加快速度)

# --- 日誌記錄設定 (Logging Setup) ---
log_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s") # 日誌格式

# 控制台日誌處理器
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(log_formatter)

# Tkinter GUI 日誌處理器 (將日誌訊息放入佇列)
class TkinterLogHandler(logging.Handler):
    def emit(self, record):
        # 將日誌級別和格式化後的訊息加入佇列
        status_window_log_queue.append((record.levelname, self.format(record) + "\n"))


tkinter_handler = TkinterLogHandler()
tkinter_handler.setFormatter(log_formatter)
tkinter_handler.setLevel(logging.INFO) # GUI 只顯示 INFO 及以上級別的日誌

# 設定 Flask 應用程式的 logger
for handler in app.logger.handlers[:]: # 移除 Flask 預設的 handler
    app.logger.removeHandler(handler)

app.logger.addHandler(console_handler) # 加入控制台 handler
app.logger.addHandler(tkinter_handler) # 加入 Tkinter handler
app.logger.setLevel(logging.DEBUG)     # 設定 Flask logger 的最低級別為 DEBUG
app.logger.propagate = False           # 防止日誌訊息向 root logger 傳播

# 設定 Werkzeug (Flask 底層 WSGI 伺服器) 的 logger
werkzeug_logger = logging.getLogger("werkzeug")
for handler in werkzeug_logger.handlers[:]: # 移除 Werkzeug 預設的 handler
    werkzeug_logger.removeHandler(handler)

werkzeug_logger.addHandler(console_handler)
werkzeug_logger.addHandler(tkinter_handler)
werkzeug_logger.setLevel(logging.INFO)     # Werkzeug 的日誌級別設為 INFO (過濾掉 DEBUG 請求)
werkzeug_logger.propagate = False          # 防止向 root logger 傳播
# --- END 日誌記錄設定 ---


# --- 輔助函式：處理打包後的資源路徑 ---
def resource_path(relative_path):
    """
    獲取資源的絕對路徑，適用於開發環境和 PyInstaller 打包後的環境。
    PyInstaller 會將數據文件放在一個臨時目錄 _MEIPASS 中。
    """
    try:
        # PyInstaller 建立一個暫存資料夾並將路徑儲存在 _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # 如果不是在 PyInstaller 環境中，則使用腳本所在的目錄
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# --- 輔助函式：路徑相關 (Path Helper Functions) ---
def is_safe_project_name(project_name):
    """檢查專案名稱是否安全 (防止路徑遍歷等問題)"""
    if not project_name:
        return False # 不能是空字串

    # 不能包含路徑分隔符，不能以 '.' 開頭 (隱藏檔案)
    if "/" in project_name or "\\" in project_name or project_name.startswith("."):
        return False

    return True


def get_project_report_folder(project_name):
    """獲取專案的報告資料夾絕對路徑"""
    if not is_safe_project_name(project_name):
        abort(400, "提供了無效的專案名稱。")

    # 拼接基礎報告目錄和專案名稱，並取得絕對路徑
    return os.path.abspath(os.path.join(BASE_REPORT_FOLDER, project_name))


def get_project_data_folder(project_name, ensure_exists=False):
    """
    獲取專案的數據資料夾絕對路徑。
    如果 ensure_exists 為 True，則會確保該資料夾及其必要的子資料夾 (截圖、垃圾桶) 存在。
    """
    if not is_safe_project_name(project_name):
        abort(400, "提供了無效的專案名稱。")

    path = os.path.abspath(os.path.join(BASE_DATA_FOLDER, project_name))

    if ensure_exists:
        try:
            # 確保主數據資料夾、截圖資料夾、截圖垃圾桶資料夾都存在
            os.makedirs(path, exist_ok=True)
            os.makedirs(os.path.join(path, SCREENSHOTS_SUBFOLDER), exist_ok=True)
            os.makedirs(os.path.join(path, TRASH_SCREENSHOTS_SUBFOLDER), exist_ok=True)

            # 記錄資料夾創建 (如果它原本不存在)
            if not os.path.isdir(path) or \
               not os.path.isdir(os.path.join(path, SCREENSHOTS_SUBFOLDER)) or \
               not os.path.isdir(os.path.join(path, TRASH_SCREENSHOTS_SUBFOLDER)):
                 app.logger.info(f"已確保專案數據資料夾及其子資料夾存在: {path}")

        except OSError as e:
            # 如果創建資料夾失敗，記錄錯誤並回傳 500 錯誤
            app.logger.error(f"無法在 '{path}' 中建立/確保數據子資料夾: {e}")
            abort(500, f"無法為專案 '{project_name}' 建立數據目錄/子資料夾。")

    return path


def get_project_screenshot_folder(project_name, ensure_exists=False):
    """獲取專案的截圖資料夾絕對路徑"""
    # 依賴 get_project_data_folder 來獲取基礎路徑並可能創建資料夾
    project_data_folder = get_project_data_folder(project_name, ensure_exists=ensure_exists)
    return os.path.join(project_data_folder, SCREENSHOTS_SUBFOLDER)


def get_project_status_file(project_name):
    """獲取專案的狀態檔絕對路徑"""
    return os.path.join(get_project_data_folder(project_name), STATUS_FILE_NAME)


def get_project_advisory_file(project_name):
    """獲取專案的修補建議檔絕對路徑"""
    return os.path.join(get_project_data_folder(project_name), ADVISORY_FILE_NAME)


def get_project_config_file(project_name):
    """獲取專案的設定檔絕對路徑"""
    return os.path.join(get_project_data_folder(project_name), APP_CONFIG_FILE_NAME)


def get_target_list_file(project_name):
    """獲取專案的目標清單 (target.xlsx) 絕對路徑"""
    return os.path.join(get_project_report_folder(project_name), TARGET_LIST_FILENAME)


def get_project_lock(project_name, lock_type):
    """
    獲取特定專案和特定類型檔案的線程鎖。
    這確保了同時只有一個線程能修改同一個專案的同一個檔案 (例如狀態檔)。
    """
    with project_locks_lock: # 鎖住全局的 project_file_locks 字典
        # 取得或建立該專案的鎖字典
        project_locks = project_file_locks.setdefault(project_name, {})
        # 取得或建立該檔案類型的鎖
        lock = project_locks.setdefault(lock_type, threading.Lock())
        return lock


# --- 輔助函式：數據讀取 (Data Reading Helpers) ---
def load_weakness_list():
    """從檔案載入弱點名稱列表，並進行快取"""
    global weakness_list_cache # 使用全域變數快取

    with weakness_list_lock: # 鎖住以進行讀取或寫入快取
        if weakness_list_cache is not None:
            return weakness_list_cache # 如果快取存在，直接返回

        weaknesses = []
        weakness_list_path = resource_path(WEAKNESS_LIST_FILE) # 使用 resource_path 以支援打包

        if os.path.exists(weakness_list_path):
            try:
                with open(weakness_list_path, "r", encoding="utf-8") as f:
                    # 讀取檔案，去除空白行和首尾空白
                    weaknesses = [line.strip() for line in f if line.strip()]

                app.logger.info(f"從 {weakness_list_path} 載入了 {len(weaknesses)} 個弱點名稱")
                # 去重並排序後存入快取
                weakness_list_cache = sorted(list(set(weaknesses)))

            except Exception as e:
                app.logger.error(f"讀取弱點列表檔案 {weakness_list_path} 時發生錯誤: {e}")
                weakness_list_cache = [] # 出錯時返回空列表
        else:
            app.logger.warning(f"找不到弱點列表檔案: {weakness_list_path}")
            weakness_list_cache = [] # 找不到檔案時返回空列表

        return weakness_list_cache


# --- 新增：載入常用備註 ---
def load_common_notes():
    """從 data/common_notes.json 載入常用備註。"""
    global common_notes_cache

    with common_notes_lock:
        if common_notes_cache is not None:
            return common_notes_cache

        notes_file_path = os.path.join(BASE_DATA_FOLDER, COMMON_NOTES_FILE)
        # 如果 BASE_DATA_FOLDER 可能在打包後的內部，請使用 resource_path
        # notes_file_path = resource_path(os.path.join(BASE_DATA_FOLDER, COMMON_NOTES_FILE))

        if os.path.exists(notes_file_path) and os.path.getsize(notes_file_path) > 0:
            try:
                with open(notes_file_path, "r", encoding="utf-8") as f:
                    loaded_notes = json.load(f)

                if isinstance(loaded_notes, dict):
                    common_notes_cache = loaded_notes
                    app.logger.info(f"載入了 {len(common_notes_cache)} 個弱點類型的常用備註。")
                else:
                    common_notes_cache = {}
                    app.logger.warning(f"常用備註檔案 {notes_file_path} 不是字典格式。")

            except json.JSONDecodeError as e:
                common_notes_cache = {}
                app.logger.error(f"解碼常用備註檔案 {notes_file_path} 的 JSON 時發生錯誤: {e}。")

            except Exception as e:
                common_notes_cache = {}
                app.logger.error(f"載入常用備註檔案 {notes_file_path} 時發生錯誤: {e}。")
        else:
            common_notes_cache = {}
            app.logger.warning(f"找不到常用備註檔案或檔案為空: {notes_file_path}")

        return common_notes_cache


# --- 載入/儲存 專案設定 (Load/Save Project Config) ---
def load_project_config(project_name):
    """載入指定專案的設定檔 (app_config.json)"""
    config_file = get_project_config_file(project_name)
    # 預設設定，如果檔案不存在或無效，則使用專案名稱作為顯示名稱
    defaults = {"project_display_name": project_name}
    loaded_config = defaults.copy()

    if os.path.exists(config_file) and os.path.getsize(config_file) > 0: # 檢查檔案是否存在且非空
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                config_data = json.load(f) # 讀取 JSON

            if isinstance(config_data, dict):
                # 讀取顯示名稱，若無效則使用預設值
                name = config_data.get("project_display_name", defaults["project_display_name"])
                if isinstance(name, str) and name.strip():
                    loaded_config["project_display_name"] = name.strip()
                else:
                    app.logger.warning(f"設定檔 {config_file} 中的 'project_display_name' 無效。")
            else:
                app.logger.warning(f"設定檔 {config_file} 格式無效 (非字典)。")

        except json.JSONDecodeError as e:
            app.logger.error(f"解碼設定檔 {config_file} 的 JSON 時發生錯誤: {e}。")

        except Exception as e:
            app.logger.error(f"載入設定檔 {config_file} 時發生錯誤: {e}。")
    else:
        # 如果檔案不存在或為空，記錄訊息並創建預設設定檔
        app.logger.info(f"找不到專案 '{project_name}' 的設定檔或檔案為空。正在建立...")
        save_project_config(project_name, defaults) # 儲存預設設定

    app.logger.debug(f"載入專案 '{project_name}' 的設定: {loaded_config}")
    return loaded_config


def save_project_config(project_name, config_data):
    """儲存指定專案的設定檔 (app_config.json)"""
    config_file = get_project_config_file(project_name)
    get_project_data_folder(project_name, ensure_exists=True) # 確保資料夾存在
    lock = get_project_lock(project_name, "config") # 獲取設定檔的鎖

    # 只儲存必要的設定欄位
    data_to_save = {
        "project_display_name": config_data.get("project_display_name", project_name)
    }

    with lock: # 鎖定檔案操作
        try:
            with open(config_file, "w", encoding="utf-8") as f:
                # 將設定數據寫入 JSON 檔案，美化格式
                json.dump(data_to_save, f, ensure_ascii=False, indent=4)

            app.logger.info(f"已儲存專案 '{project_name}' 的設定。")

        except Exception as e:
            app.logger.error(f"儲存專案 '{project_name}' 的設定時發生錯誤: {e}")


# --- 載入/儲存 伺服器設定 (Load/Save Server Config) ---
def load_server_config():
    """載入伺服器設定檔 (app_server_config.json)，主要是 Port 號"""
    global server_port # 使用全域變數 server_port

    default_port = 5001 # 預設 Port
    loaded_port = default_port
    config_path = resource_path(SERVER_CONFIG_FILE) # 使用 resource_path 支援打包

    if os.path.exists(config_path) and os.path.getsize(config_path) > 0:
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)

            # 從設定檔讀取 Port，若無效則使用預設值
            port_from_file = config.get("server_port", default_port) if isinstance(config, dict) else default_port

            # 驗證 Port 是否在有效範圍內
            if isinstance(port_from_file, int) and 1024 <= port_from_file <= 65535:
                loaded_port = port_from_file
            else:
                app.logger.warning(f"無效的 Port 號 '{port_from_file}'。將使用預設 Port {default_port}。")
                loaded_port = default_port

        except json.JSONDecodeError as e:
            app.logger.error(f"解碼伺服器設定檔時發生錯誤: {e}。將使用預設 Port。")
            loaded_port = default_port

        except Exception as e:
            app.logger.error(f"載入伺服器設定檔時發生錯誤: {e}。將使用預設 Port。")
            loaded_port = default_port
    else:
        # 如果檔案不存在或為空，使用預設值並創建設定檔
        app.logger.info(f"找不到伺服器設定檔或檔案為空。將使用預設 Port 並建立檔案。")
        save_server_config() # 儲存預設設定 (使用目前的 server_port，此時應為 default_port)

    server_port = loaded_port # 更新全域變數


def save_server_config():
    """儲存伺服器設定檔 (app_server_config.json)"""
    global server_port # 使用全域變數

    config_path = resource_path(SERVER_CONFIG_FILE) # 使用 resource_path 支援打包

    try:
        with open(config_path, "w", encoding="utf-8") as f:
            # 將伺服器 Port 寫入 JSON 檔案
            json.dump({"server_port": server_port}, f, indent=4)

        app.logger.info(f"已儲存伺服器設定: Port={server_port}")

    except Exception as e:
        app.logger.error(f"儲存伺服器設定時發生錯誤: {e}")


# --- 載入/儲存 專案狀態 (Load/Save Project Statuses) ---
def load_statuses(project_name):
    """載入指定專案的弱點狀態檔 (vulnerability_status.json)，確保處理 screenshots_meta"""
    status_file = get_project_status_file(project_name)
    lock = get_project_lock(project_name, "status") # 獲取狀態檔的鎖
    statuses = {} # 初始化為空字典

    with lock: # 鎖定檔案操作
        if os.path.exists(status_file) and os.path.getsize(status_file) > 0: # 檢查檔案是否存在且非空
            try:
                with open(status_file, "r", encoding="utf-8") as f:
                    loaded_data = json.load(f) # 讀取 JSON

                if not isinstance(loaded_data, dict):
                    app.logger.error(f"狀態檔 {status_file} 格式無效 (非字典)。")
                    return {} # 格式錯誤返回空字典

                converted_statuses = {} # 儲存轉換後的狀態
                report_count = 0
                issue_count = 0

                # 遍歷讀取的數據 (以報告檔名為鍵)
                for report_filename, report_data in loaded_data.items():
                    report_count += 1

                    if not isinstance(report_data, dict):
                        # 如果報告數據不是字典，記錄警告並跳過
                        app.logger.warning(f"跳過無效的報告條目 '{report_filename}'。")
                        converted_statuses[report_filename] = {}
                        continue

                    # 初始化當前報告的狀態，包含判讀完成標記
                    current_issues = {REPORT_COMPLETED_KEY: report_data.get(REPORT_COMPLETED_KEY, False)}

                    # 遍歷報告中的每個弱點 (以 issue_id 為鍵)
                    for issue_id, status_data in report_data.items():
                        if issue_id == REPORT_COMPLETED_KEY:
                            continue # 跳過判讀完成標記

                        issue_count += 1

                        # 初始化預設值
                        status_value = DEFAULT_STATUS
                        ss_taken = False
                        note_value = DEFAULT_NOTE
                        manual_details = None
                        source = "appscan" # 預設來源為 AppScan
                        screenshots_meta = {} # 初始化截圖元數據

                        # 處理不同格式的舊數據
                        if isinstance(status_data, dict): # 新格式 (字典)
                            status_value = status_data.get("status", DEFAULT_STATUS)
                            ss_taken = status_data.get("screenshot_taken", False)
                            note_value = status_data.get("note", DEFAULT_NOTE)
                            manual_details = status_data.get("manual_details") # 取得手動新增的詳細資訊
                            # 根據是否有手動資訊決定來源，或直接讀取來源欄位
                            source = status_data.get("source", "manual" if manual_details else "appscan")
                            # 讀取截圖元數據，確保是字典
                            screenshots_meta = status_data.get("screenshots_meta", {})
                            if not isinstance(screenshots_meta, dict):
                                app.logger.warning(f"狀態檔中問題 {issue_id} 的 screenshots_meta 格式無效，已重設。")
                                screenshots_meta = {}

                        elif isinstance(status_data, str): # 舊格式 (僅狀態字串)
                            status_value = status_data
                            app.logger.debug(f"讀取到舊版狀態格式 (僅字串) for {issue_id}。")

                        else:
                            app.logger.warning(f"弱點 {issue_id} 的狀態格式非預期。")

                        # 建立標準化的弱點條目
                        issue_entry = {
                            "status": status_value,
                            "screenshot_taken": ss_taken,
                            "note": note_value,
                            "source": source,
                            "screenshots_meta": screenshots_meta # 包含截圖元數據
                        }

                        # 如果是手動新增的弱點，加入詳細資訊
                        if manual_details:
                            issue_entry["manual_details"] = manual_details

                        current_issues[issue_id] = issue_entry

                    converted_statuses[report_filename] = current_issues # 儲存處理完的報告狀態

                statuses = converted_statuses # 更新最終的狀態字典
                app.logger.info(f"從 {status_file} 載入了 {report_count} 個報告, {issue_count} 個弱點的狀態")

            except json.JSONDecodeError as e:
                app.logger.error(f"解碼狀態檔 {status_file} 的 JSON 時發生錯誤: {e}。")
                return {} # 解碼錯誤返回空字典

            except Exception as e:
                app.logger.error(f"載入狀態時發生錯誤: {e}, {traceback.format_exc()}。")
                return {} # 其他錯誤返回空字典
        else:
            app.logger.info(f"狀態檔 {status_file} 不存在或為空。")

    return statuses # 返回載入的狀態字典


def save_statuses(project_name, statuses):
    """儲存指定專案的弱點狀態檔 (vulnerability_status.json)，確保包含 screenshots_meta"""
    status_file = get_project_status_file(project_name)
    get_project_data_folder(project_name, ensure_exists=True) # 確保資料夾存在
    lock = get_project_lock(project_name, "status") # 獲取狀態檔的鎖

    with lock: # 鎖定檔案操作
        try:
            data_to_save = {} # 初始化要儲存的數據

            # 遍歷傳入的狀態字典
            for report_filename, report_data in statuses.items():
                if not isinstance(report_data, dict):
                    app.logger.warning(f"跳過非字典格式的報告數據 '{report_filename}'。")
                    continue

                # 初始化報告條目，包含判讀完成標記
                report_entry = {REPORT_COMPLETED_KEY: report_data.get(REPORT_COMPLETED_KEY, False)}

                # 遍歷報告中的每個弱點
                for issue_id, status_info in report_data.items():
                    if issue_id == REPORT_COMPLETED_KEY:
                        continue # 跳過判讀完成標記

                    if isinstance(status_info, dict):
                        # 如果狀態資訊是字典，確保有 'source' 欄位
                        if "source" not in status_info:
                            # 如果缺少來源，根據 issue_id 判斷是手動還是 AppScan
                            status_info["source"] = "manual" if issue_id.startswith("_manual_") else "appscan"

                        # 確保 'screenshots_meta' 存在且為字典
                        if "screenshots_meta" not in status_info or not isinstance(status_info["screenshots_meta"], dict):
                            status_info["screenshots_meta"] = {}

                        report_entry[issue_id] = status_info # 直接儲存字典
                    else:
                        # 如果狀態資訊不是字典 (可能是舊格式)，轉換為標準格式
                        app.logger.warning(f"轉換報告 '{report_filename}' 中弱點 '{issue_id}' 的非字典狀態。")
                        report_entry[issue_id] = {
                            "status": str(status_info or DEFAULT_STATUS), # 轉換為字串，若為空則用預設值
                            "screenshot_taken": False, # 預設為 False
                            "note": DEFAULT_NOTE,      # 預設為空筆記
                            # 根據 issue_id 判斷來源
                            "source": "manual" if issue_id.startswith("_manual_") else "appscan",
                            "screenshots_meta": {} # 初始化為空字典
                        }

                data_to_save[report_filename] = report_entry # 儲存處理完的報告條目

            # 將數據寫入 JSON 檔案
            with open(status_file, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=4)

            app.logger.debug(f"已儲存專案 '{project_name}' 的狀態。")

        except Exception as e:
            app.logger.error(f"儲存專案 '{project_name}' 的狀態時發生錯誤: {e}")


# --- 載入/儲存 全域規則 (Load/Save Global Rules) ---
def load_rules():
    """載入全域排除規則檔 (exclusion_rules.json)"""
    global exclusion_rules # 使用全域變數

    rules_path = resource_path(RULES_FILE) # 使用 resource_path 支援打包

    with rules_file_lock: # 鎖定檔案操作
        if os.path.exists(rules_path) and os.path.getsize(rules_path) > 0: # 檢查檔案是否存在且非空
            try:
                with open(rules_path, "r", encoding="utf-8") as f:
                    loaded_rules = json.load(f) # 讀取 JSON

                if isinstance(loaded_rules, list): # 確保讀取的是列表
                    exclusion_rules = loaded_rules # 更新全域規則列表
                    app.logger.info(f"載入了 {len(exclusion_rules)} 條全域規則。")
                else:
                    exclusion_rules = [] # 格式不對則清空
                    app.logger.warning(f"規則檔 {rules_path} 不是列表格式。")

            except Exception as e:
                exclusion_rules = [] # 出錯時清空
                app.logger.error(f"載入規則時發生錯誤: {e}。")
        else:
            exclusion_rules = [] # 檔案不存在或為空則清空
            app.logger.info(f"找不到規則檔 {rules_path} 或檔案為空。")


def save_rules():
    """儲存全域排除規則檔 (exclusion_rules.json)"""
    global exclusion_rules # 使用全域變數

    rules_path = resource_path(RULES_FILE) # 使用 resource_path 支援打包

    with rules_file_lock: # 鎖定檔案操作
        try:
            with open(rules_path, "w", encoding="utf-8") as f:
                # 將規則列表寫入 JSON 檔案
                json.dump(exclusion_rules, f, ensure_ascii=False, indent=4)

            app.logger.debug(f"已儲存 {len(exclusion_rules)} 條規則。")

        except Exception as e:
            app.logger.error(f"儲存規則時發生錯誤: {e}")


# --- 載入 專案修補建議 (Load Project Advisory) ---
def load_advisory_data(project_name):
    """載入指定專案的修補建議檔 (advisory.json)"""
    advisory_file = get_project_advisory_file(project_name)
    data = {} # 初始化為空字典

    if os.path.exists(advisory_file):
        try:
            with open(advisory_file, "r", encoding="utf-8") as f:
                loaded_data = json.load(f) # 讀取 JSON

            if isinstance(loaded_data, dict): # 確保是字典格式
                data = loaded_data # 更新數據
                app.logger.info(f"為專案 '{project_name}' 載入了 {len(data)} 條修補建議。")
            else:
                app.logger.error(f"修補建議檔 {advisory_file} 格式無效 (非字典)。")

        except Exception as e:
            app.logger.error(f"載入專案 '{project_name}' 的修補建議時發生錯誤: {e}。")
    else:
        app.logger.warning(f"找不到修補建議檔 '{advisory_file}'。")

    return data


# --- 輔助函式：報告摘要/狀態 (Report Summary/Status Helpers) ---
def get_report_summary_severities(filepath):
    """從 XML 檔案解析掃描摘要中的各嚴重性數量"""
    filename = os.path.basename(filepath) # 獲取檔名用於日誌

    try:
        tree = ET.parse(filepath) # 解析 XML
        root = tree.getroot()
        summary = root.find("scan-summary") # 找到掃描摘要標籤
    except Exception as e:
        # XML 解析或讀取錯誤
        app.logger.error(f"解析/讀取報告 '{filename}' 的摘要時發生錯誤: {e}")
        return None # 返回 None 表示失敗

    if summary is None:
        # 找不到摘要標籤
        app.logger.warning(f"報告 '{filename}' 中找不到 <scan-summary> 標籤。")
        return None # 返回 None

    severity_counts = {} # 初始化數量字典

    # 遍歷定義的嚴重性等級
    for level_key in SEVERITY_LEVELS.keys():
        xml_tag = f"total-issues-severity-{level_key}" # 構造 XML 標籤名
        count_text = summary.findtext(xml_tag, "0") # 查找對應標籤的文本，預設為 "0"

        try:
            # 轉換為整數，若為空字串或無效則視為 0
            severity_counts[level_key] = int(count_text or "0")
        except (ValueError, TypeError):
            # 轉換失敗記錄警告，設為 0
            app.logger.warning(f"報告 '{filename}' 中標籤 '{xml_tag}' 的值 '{count_text}' 無效。")
            severity_counts[level_key] = 0

    return severity_counts # 返回包含各嚴重性數量的字典


def get_scan_status(filepath):
    """從 XML 檔案解析掃描狀態"""
    filename = os.path.basename(filepath) # 獲取檔名用於日誌

    try:
        tree = ET.parse(filepath) # 解析 XML
        root = tree.getroot()
        summary = root.find("scan-summary") # 找到掃描摘要標籤
    except ET.ParseError as e:
        # XML 解析錯誤
        app.logger.error(f"XML 解析錯誤 (狀態) '{filename}': {e}")
        return PARSE_ERROR_STATUS # 返回解析錯誤狀態
    except Exception as e:
        # 其他讀取錯誤
        app.logger.error(f"讀取檔案錯誤 (狀態) '{filename}': {e}")
        return READ_ERROR_STATUS # 返回讀取錯誤狀態

    if summary is None:
        # 找不到摘要標籤
        app.logger.warning(f"報告 '{filename}' 中找不到 <scan-summary> 標籤。")
        return INCOMPLETE_STATUS # 返回格式不完整狀態

    status_elem = summary.find("scan-run-status") # 查找狀態標籤

    if status_elem is not None and status_elem.text:
        # 如果找到狀態標籤且有文本
        raw_status = status_elem.text.strip() # 獲取原始狀態文字
        # 從映射中查找對應的中文狀態，若找不到則返回原始狀態
        return SCAN_STATUS_MAP.get(raw_status, raw_status)

    # --- 如果沒有明確的狀態標籤，嘗試根據掃描頁數推斷 ---
    pages_scanned = -1 # 初始化掃描頁數
    pages_scanned_text = summary.findtext("num-pages-scanned") # 查找掃描頁數標籤

    if pages_scanned_text:
        try:
            pages_scanned = int(pages_scanned_text) # 轉換為整數
        except (ValueError, TypeError):
            pass # 轉換失敗則忽略

    if pages_scanned == 0:
        # 如果掃描頁數為 0，通常表示失敗
        return SCAN_STATUS_MAP.get("Failed", "失敗")

    if pages_scanned > 0:
        # 如果掃描頁數大於 0，通常表示成功
        return SCAN_STATUS_MAP.get("Success", "成功")

    # 如果以上都無法判斷，返回預設未知狀態
    return DEFAULT_SCAN_STATUS


# --- 輔助函式：檔名處理 (Filename Helpers) ---
def sanitize_filename_part(part, max_len=50):
    """清理字串，使其適合作為檔名的一部分"""
    if not part or str(part).strip().lower() == 'n/a':
        return "na" # 特殊處理 'N/A'

    sanitized = str(part).strip(" .") # 去除首尾空格和點
    sanitized = re.sub(r'^[a-zA-Z]+://', '', sanitized) # 移除 URL scheme (http://)
    sanitized = re.sub(r'[?#].*$', '', sanitized) # 移除 URL 查詢參數和片段 (#)
    # 移除所有不安全字元 (非字母數字、-、.、_、~、/ 及 CJK/日韓文)，替換為 '_'
    sanitized = re.sub(r'[^\w\-\._~\/\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+', '_', sanitized, flags=re.UNICODE)
    sanitized = re.sub(r'\/+', '_', sanitized) # 將路徑分隔符 / 替換為 _
    sanitized = re.sub(r"_+", "_", sanitized) # 將多個連續的 _ 替換為單個 _
    sanitized = sanitized.strip("_") # 去除首尾的 _
    sanitized = sanitized[:max_len] # 截斷到最大長度

    if sanitized:
        return sanitized
    else:
        return "sanitized_empty" # 如果清理後變為空字串，返回特定標記


def allowed_file(filename):
    """檢查檔名是否具有允許的圖片副檔名"""
    return "." in filename and \
           filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# --- 輔助函式：截圖處理 (Screenshot Helpers) ---
def construct_screenshot_filename_prefix(report_num, source_label, sanitized_name, sanitized_url, sanitized_entity):
    """構造截圖檔名的前綴部分"""
    report_num_str = str(report_num) # 報告編號
    # 確保來源標籤是預定義的 "APPSCAN" 或 "手動"
    source_part = APPSCAN_SOURCE_LABEL if source_label.upper() == APPSCAN_SOURCE_LABEL else MANUAL_SOURCE_LABEL
    # 格式: 編號-來源-弱點名-URL-實體名-
    return f"{report_num_str}-{source_part}-{sanitized_name}-{sanitized_url}-{sanitized_entity}-"


def get_next_screenshot_sequence(project_name, filename_prefix):
    """根據指定的前綴，在截圖資料夾中查找下一個可用的序列號"""
    screenshot_dir = get_project_screenshot_folder(project_name)
    max_num = 0 # 初始化最大序列號

    if not os.path.isdir(screenshot_dir):
        return 1 # 如果資料夾不存在，從 1 開始

    try:
        if not filename_prefix.endswith('-'):
            filename_prefix += '-' # 確保前綴以 '-' 結尾

        prefix_len = len(filename_prefix)

        # 遍歷截圖資料夾中的檔案
        for filename in os.listdir(screenshot_dir):
            if filename.startswith(filename_prefix): # 檢查是否符合前綴
                name_part, dot, extension = filename.rpartition(".") # 分離檔名和副檔名
                if dot and extension.lower() in ALLOWED_EXTENSIONS: # 確保是有效的圖片檔案
                    sequence_part = name_part[prefix_len:] # 提取序列號部分
                    if sequence_part.isdigit(): # 檢查是否為數字
                        try:
                            max_num = max(max_num, int(sequence_part)) # 更新最大序列號
                        except ValueError:
                            pass # 轉換失敗則忽略

    except Exception as e:
        app.logger.error(f"查找前綴 '{filename_prefix}' 的序列號時發生錯誤: {e}")
        return 1 # 出錯時從 1 開始

    return max_num + 1 # 返回下一個可用的序列號

def get_existing_screenshots(project_name, filename_prefix):
    """
    根據指定的前綴，獲取所有已存在的截圖檔名列表 (大小寫不敏感)。
    返回的檔名保持其原始大小寫。
    """
    screenshot_list = []
    screenshot_dir = get_project_screenshot_folder(project_name)

    # --- 調試日誌：記錄搜索目錄和前綴 ---
    app.logger.debug(f"get_existing_screenshots: Searching in '{screenshot_dir}' for prefix '{filename_prefix}'")

    if not os.path.isdir(screenshot_dir):
        # --- 調試日誌：目錄不存在 ---
        app.logger.warning(f"  Screenshot directory does not exist or is not a directory: {screenshot_dir}")
        return [] # 資料夾不存在返回空列表

    try:
        # 確保前綴以 '-' 結尾，這對於後續提取序號很重要
        if not filename_prefix.endswith('-'):
            filename_prefix += '-'

        prefix_lower = filename_prefix.lower() # 轉換為小寫以進行不區分大小寫的比較
        prefix_len = len(filename_prefix)      # 保留原始前綴的長度，用於後續從原始檔名切割序號

        # --- 調試日誌：列出目錄內容 ---
        all_files_in_dir = os.listdir(screenshot_dir)
        app.logger.debug(f"  Directory contents ({len(all_files_in_dir)} files): {all_files_in_dir[:20]}...") # 只記錄前 20 個，避免過長

        # 遍歷截圖資料夾中的檔案
        for filename in all_files_in_dir:
            filename_lower = filename.lower() # 取得當前檔案的小寫版本

            # --- 調試日誌：記錄每個檔案的比較過程 ---
            match = filename_lower.startswith(prefix_lower)
            app.logger.debug(f"    Checking: '{filename_lower}'.startswith('{prefix_lower}') -> {match}")

            # **核心修改：使用大小寫不敏感的方式比較前綴**
            if match:
                # 分割 *原始* 檔名以獲取副檔名 (保留原始大小寫)
                name_part, dot, extension = filename.rpartition(".")

                if dot and extension.lower() in ALLOWED_EXTENSIONS: # 檢查副檔名是否允許
                    # 從 *原始* 檔名部分提取序列號部分 (使用原始前綴的長度)
                    sequence_part = name_part[prefix_len:]
                    if sequence_part.isdigit(): # 確保序列號部分是數字
                        # 將 *原始* 檔名加入列表
                        screenshot_list.append(filename)
                        # --- 調試日誌：記錄找到的匹配項 ---
                        app.logger.debug(f"      --> Match found and valid image: '{filename}'")
                    else:
                         # --- 調試日誌：記錄為何未加入 (序號問題) ---
                         app.logger.debug(f"      --> Prefix matched, but sequence part '{sequence_part}' is not digits for '{filename}'")
                else:
                    # --- 調試日誌：記錄為何未加入 (副檔名問題) ---
                    app.logger.debug(f"      --> Prefix matched, but invalid extension or format for '{filename}'")

        # 定義排序函式，根據檔名中的序列號排序 (保持不變)
        def get_sequence_num(fname):
            match = re.search(r"-(\d+)\.\w+$", fname) # 查找結尾的 '-數字.副檔名'
            if match:
                try:
                    return int(match.group(1)) # 返回數字部分
                except ValueError:
                     app.logger.warning(f"Could not parse sequence number from filename during sort: {fname}")
                     return float('inf') # 解析失敗的排後面
            else:
                # --- 警告日誌：無法從檔名提取序號 ---
                app.logger.warning(f"Could not extract sequence number from filename during sort: {fname}")
                return float('inf') # 找不到序號的排後面

        screenshot_list.sort(key=get_sequence_num) # 排序列表

        # --- 調試日誌：記錄最終結果 ---
        app.logger.debug(f"  Found and sorted {len(screenshot_list)} matching screenshots for prefix '{filename_prefix}'.")

    except Exception as e:
        app.logger.error(f"Error listing screenshots for prefix '{filename_prefix}' in '{screenshot_dir}': {e}", exc_info=True) # 加入 exc_info 記錄 traceback
        return [] # 出錯時返回空列表

    return screenshot_list


# --- XML 解析器 (XML Parser) ---
def parse_appscan_xml(project_name, filepath, report_filename):
    """解析 AppScan XML 報告檔案，提取所需資訊 (不再查找截圖)。"""
    app.logger.debug(f"Parsing XML: Project='{project_name}', File='{filepath}'")

    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
    except ET.ParseError as e:
        app.logger.error(f"XML Parse Error for {report_filename}: {e}")
        return None
    except Exception as e:
        app.logger.error(f"XML Read Error for {report_filename}: {e}")
        return None

    # --- 建立查找表 (用於 ID 到名稱的轉換) ---
    lookups = { "dict": {}, "issue_type": {}, "url": {}, "entity": {} }

    try:
        for item in root.findall(".//dictionary/item"):
            lookups["dict"][item.get("id")] = item.text or ""

        for item in root.findall(".//issue-type-group/item"):
            lookups["issue_type"][item.get("id")] = item.findtext("name", "Unknown Type")

        for item in root.findall(".//url-group/item"):
            lookups["url"][item.get("id")] = item.findtext("name", "N/A")

        for item in root.findall(".//entity-group/item"):
            item_id = item.get("id")
            type_key = item.findtext("entity-type", "N/A") # 實體類型的 ID (需查 dict)
            name = item.findtext("name", "N/A")           # 實體名稱
            # 儲存實體名稱和轉換後的類型名稱
            lookups["entity"][item_id] = {"name": name, "type": lookups["dict"].get(type_key, type_key)}

    except Exception as e:
        app.logger.warning(f"建立報告 {report_filename} 的查找表時發生錯誤: {e}")

    # --- 提取掃描資訊 ---
    scan_info = {
        "scan_name": "N/A", # 掃描名稱
        "scan_date": "N/A", # 掃描日期
        "base_filename": os.path.splitext(report_filename)[0] # 報告檔名 (不含副檔名)
    }
    scan_info_elem = root.find("scan-information")

    if scan_info_elem is not None:
        scan_info["scan_name"] = scan_info_elem.findtext("scan-name", "N/A")
        scan_info["scan_date"] = scan_info_elem.findtext("scan-date-and-time", "N/A")

    # --- 提取摘要資訊 ---
    severity_summary_raw = get_report_summary_severities(filepath) # 獲取各嚴重性數量
    display_summary = {} # 用於前端顯示的摘要

    # 初始化掃描統計數據
    scan_stats = { k: "N/A" for k in ["pages_scanned", "total_pages", "entities_tested", "total_entities", "issues_found"] }
    summary_elem = root.find("scan-summary")

    if summary_elem is not None:
        if severity_summary_raw:
            # 將原始數量轉換為顯示格式 (分開 informational)
            display_summary = { k: str(severity_summary_raw.get(k, 0)) for k in SEVERITY_SUMMARY_ORDER if k != "informational" }
            display_summary["info"] = str(severity_summary_raw.get("informational", 0))
            total_issues_from_sum = sum(severity_summary_raw.values()) # 從摘要計算總數
            issues_found_tag = summary_elem.findtext("num-issues-found") # 嘗試讀取 <num-issues-found>

            try:
                scan_stats["issues_found"] = str(int(issues_found_tag)) if issues_found_tag else str(total_issues_from_sum) # 若無標籤則用計算值
            except (ValueError, TypeError):
                scan_stats["issues_found"] = str(total_issues_from_sum) # 轉換失敗也用計算值

            display_summary["total_issues"] = scan_stats["issues_found"] # 加入總數到顯示摘要
        else:
            # 如果無法獲取嚴重性摘要，則全部設為 0
            display_summary = {k: "0" for k in ["total_issues", "critical", "high", "medium", "low", "info"]}
            scan_stats["issues_found"] = "0"

        # 映射統計數據鍵名到 XML 標籤名
        scan_stats_map = {
            "pages_scanned": "num-pages-scanned",
            "total_pages": "total-num-pages",
            "entities_tested": "num-security-entities-tested",
            "total_entities": "total-num-security-entities"
        }

        # 讀取其他統計數據
        for stat_key, xml_tag in scan_stats_map.items():
            scan_stats[stat_key] = summary_elem.findtext(xml_tag, "N/A")

        # 再次檢查 issues_found 是否為 N/A，若是且有嚴重性摘要，則用計算值
        if scan_stats["issues_found"] == "N/A" and severity_summary_raw:
            scan_stats["issues_found"] = str(sum(severity_summary_raw.values()))
    else:
        # 如果找不到 <scan-summary>
        app.logger.warning(f"報告 '{report_filename}' 中找不到 <scan-summary> 標籤。")
        display_summary = {k: "0" for k in ["total_issues", "critical", "high", "medium", "low", "info"]}
        scan_stats["issues_found"] = "0"

    # --- 提取問題列表 ---
    issues = [] # 初始化問題列表

    # 從檔名解析報告編號 (例如 '01-...')
    report_num_match = re.match(r"(\d+)-.*", report_filename)
    report_num = report_num_match.group(1) if report_num_match else "unknown" # 若無則為 'unknown'

    issue_group_elem = root.find("issue-group") # 查找問題組標籤

    if issue_group_elem is not None:
        # 遍歷每個問題項目
        for i, item in enumerate(issue_group_elem.findall("item")):
            issue_data = {} # 初始化單個問題的數據

            try:
                issue_id = item.get("id") # 獲取問題 ID
                if not issue_id:
                    raise ValueError(f"問題 {i+1} 缺少 ID")

                issue_data["id"] = issue_id

                # --- 解析嚴重性 ---
                severity_text = item.findtext("severity", "").lower().strip() # 直接讀取 <severity>
                severity_id_ref = item.findtext("severity-id", "-1")         # 讀取 <severity-id> (參考 dict)

                # 優先使用 <severity> 的文字，若不在預期範圍內，則嘗試用 <severity-id> 查找
                severity_key = severity_text if severity_text in SEVERITY_LEVELS else lookups["dict"].get(severity_id_ref, "unknown").lower()

                # 最後確認是否為有效鍵值，否則設為 'unknown'
                final_severity_key = severity_key if severity_key in SEVERITY_LEVELS else "unknown"
                issue_data["severity_key"] = final_severity_key # 儲存內部鍵值

                # 查找對應的中文顯示名稱
                issue_data["severity_display"] = SEVERITY_DISPLAY_MAP.get(final_severity_key, final_severity_key.capitalize())

                # --- 解析其他資訊 ---
                issue_data["cvss_score"] = item.findtext("cvss-score", "N/A") # CVSS 分數
                issue_data["cve_name"] = item.findtext("cve/name", "").strip() or None # CVE 名稱
                issue_data["cve_url"] = item.findtext("cve/url", "").strip() or None   # CVE 連結

                # 查找問題類型 (透過 ref 查找 lookups["issue_type"])
                issue_type_ref = item.findtext(".//issue-type/ref")
                issue_type_name = lookups["issue_type"].get(issue_type_ref, "Unknown Type")
                issue_data["issue_type"] = issue_type_name

                # 查找 URL (透過 ref 查找 lookups["url"])
                url_ref = item.findtext(".//url/ref")
                issue_data["url"] = lookups["url"].get(url_ref, "N/A")

                # 查找實體 (透過 ref 查找 lookups["entity"])
                entity_ref = item.findtext(".//entity/ref")
                entity_data = lookups["entity"].get(entity_ref, {"name": "N/A", "type": "N/A"})
                issue_data["entity_name"] = entity_data["name"]
                issue_data["entity_type"] = entity_data["type"]

                # --- 解析 Variant (原因和 HTTP 流量) ---
                reasoning_text = "N/A" # 預設原因
                http_traffic = "沒有可用的 HTTP 流量數據。" # 預設流量
                variant_item = item.find(".//variant-group/item") # 查找第一個 variant

                if variant_item is not None:
                    reasoning_text = variant_item.findtext("reasoning", "N/A").strip() # 讀取原因
                    raw_traffic = variant_item.findtext("test-http-traffic", "") # 讀取流量
                    if raw_traffic:
                        http_traffic = raw_traffic.strip() # 清理流量文本

                issue_data["reasoning"] = reasoning_text
                issue_data["http_traffic"] = http_traffic

                # --- 初始化截圖列表為空 (由 get_report_data 填充詳細資訊) ---
                issue_data["screenshots"] = [] # 保留舊鍵，以防萬一，但主要用 detailed
                issue_data["screenshots_detailed"] = [] # 初始化新的詳細列表

                # 加入掃描資訊和來源標記
                issue_data["scan_info"] = scan_info
                issue_data["source"] = "appscan"
                issues.append(issue_data) # 將處理完的問題加入列表

            except Exception as e:
                # 如果處理單個問題時出錯
                app.logger.error(f"解析報告 {report_filename} 中的問題 {i} 時發生錯誤: {e}", exc_info=True)
                # 加入一個錯誤標記的問題條目
                issues.append({
                    "id": f"error_item_{i+1}",
                    "status": "處理錯誤",
                    "screenshot_taken": False,
                    "note": f"解析錯誤 ({e})",
                    "severity_key": "error",
                    "severity_display": "錯誤",
                    "issue_type": f"處理錯誤",
                    "url": "N/A", "entity_name": "N/A", "entity_type": "N/A",
                    "reasoning": "N/A", "http_traffic": "N/A",
                    "screenshots": [], "screenshots_detailed": [], # 初始化截圖
                    "scan_info": scan_info, "source": "error"
                })
    else:
        # 如果找不到 <issue-group>
        app.logger.warning(f"報告 '{report_filename}' 中找不到 <issue-group> 標籤。")

    app.logger.info(f"已解析 '{filepath}'。找到 {len(issues)} 個潛在的 AppScan 問題。")

    # 返回包含所有解析結果的字典
    return {
        "scan_info": scan_info,         # 掃描資訊
        "summary": display_summary,     # 嚴重性摘要 (顯示用)
        "issues": issues,               # 問題列表
        "status_summary": {},           # 狀態摘要 (稍後填充)
        "scan_stats": scan_stats        # 掃描統計
    }


# --- 讀取目標 Excel (Read Target Excel) ---
def read_target_details_from_excel(project_name):
    """從專案的 target.xlsx 讀取報告編號對應的 URL 和名稱"""
    target_file = get_target_list_file(project_name) # 獲取 Excel 檔案路徑
    target_details = {} # 初始化結果字典

    if not os.path.exists(target_file):
        app.logger.warning(f"找不到目標列表檔案: {target_file}")
        return {} # 找不到檔案返回空字典

    try:
        # 載入 Excel 工作簿 (唯讀模式，只讀取數據)
        workbook = load_workbook(filename=target_file, read_only=True, data_only=True)
        sheet = workbook.active # 獲取活動工作表
        header_row = [cell.value for cell in sheet[1]] # 讀取第一行 (標頭)
        # 將標頭轉換為小寫並去除空白，以便查找
        header_lower = [str(h).lower().strip() if h is not None else "" for h in header_row]

        try:
            id_col_idx = header_lower.index("編號") # 查找 '編號' 欄位的索引
        except ValueError:
            # 如果找不到 '編號' 欄位，記錄錯誤並返回空字典
            app.logger.error(f"在 {target_file} 中找不到 '編號' 欄位。")
            return {}

        # 查找 URL 欄位的索引 (接受 'url' 或 '網址')
        url_col_idx = -1
        for header in ["url", "網址"]:
            try:
                url_col_idx = header_lower.index(header)
                break
            except ValueError:
                pass

        if url_col_idx == -1:
            app.logger.warning(f"找不到 'URL'/'網址' 欄位。")

        # 查找名稱欄位的索引 (接受 '標的名稱' 或 '名稱')
        name_col_idx = -1
        for header in ["標的名稱", "名稱"]:
            try:
                name_col_idx = header_lower.index(header)
                break
            except ValueError:
                pass

        if name_col_idx == -1:
            app.logger.warning(f"找不到 '標的名稱'/'名稱' 欄位。")

        # 從第二行開始遍歷數據行
        for row_index in range(2, sheet.max_row + 1):
            report_num = None
            url = "N/A"
            name = "N/A"
            id_cell_value = sheet.cell(row=row_index, column=id_col_idx + 1).value # 讀取編號儲存格

            if id_cell_value is not None:
                try:
                    # 嘗試將編號轉換為整數 (先轉 float 是為了處理可能的浮點數表示)
                    report_num = int(float(str(id_cell_value)))
                    if report_num < 0: # 忽略負數編號
                        app.logger.warning(f"跳過第 {row_index} 行，無效的負數 ID。")
                        report_num = None
                except (ValueError, TypeError):
                    # 轉換失敗則忽略此行
                    app.logger.warning(f"跳過第 {row_index} 行，ID 非數值。")
                    report_num = None

            if report_num is None:
                continue # 如果沒有有效的編號，跳到下一行

            # 如果找到了 URL 欄位，讀取 URL
            if url_col_idx != -1:
                url_cell_value = sheet.cell(row=row_index, column=url_col_idx + 1).value
                url = str(url_cell_value).strip() if url_cell_value is not None else "N/A"

            # 如果找到了名稱欄位，讀取名稱
            if name_col_idx != -1:
                name_cell_value = sheet.cell(row=row_index, column=name_col_idx + 1).value
                name = str(name_cell_value).strip() if name_cell_value is not None else "N/A"

            # 將讀取的 URL 和名稱存入字典，以報告編號為鍵
            target_details[report_num] = {"url": url, "name": name}

        app.logger.info(f"從 {target_file} 讀取了 {len(target_details)} 個目標。")

    except Exception as e:
        # 處理讀取 Excel 時的任何錯誤
        app.logger.error(f"讀取目標 Excel '{target_file}' 時發生錯誤: {e}", exc_info=True)
        return {} # 出錯時返回空字典

    return target_details


# --- Flask 路由 (Flask Routes) ---

# 路由：首頁，顯示專案列表
@app.route("/")
def list_projects():
    """顯示所有可用的專案列表"""
    projects = []

    if not os.path.isdir(BASE_REPORT_FOLDER):
        flash(f"報告資料夾 '{BASE_REPORT_FOLDER}' 不存在。", "danger")
        app.logger.error("基礎報告資料夾遺失。")
    else:
        try:
            for item in os.listdir(BASE_REPORT_FOLDER):
                project_path = os.path.join(BASE_REPORT_FOLDER, item)

                if os.path.isdir(project_path) and is_safe_project_name(item):
                    try:
                        config = load_project_config(item)
                        display_name = config.get("project_display_name", item)
                        project_report_folder = get_project_report_folder(item)
                        project_status_file = get_project_status_file(item)
                        project_severities = {level: 0 for level in SEVERITY_LEVELS.keys()}

                        if os.path.isdir(project_report_folder):
                            xml_files = glob.glob(os.path.join(project_report_folder, "*.xml"))
                            for xml_file in xml_files:
                                try:
                                    report_summary = get_report_summary_severities(xml_file)
                                    if report_summary:
                                        for level, count in report_summary.items():
                                            if level in project_severities:
                                                project_severities[level] += count
                                except Exception as xml_e:
                                    app.logger.warning(f"解析專案 '{item}' 的報告 '{os.path.basename(xml_file)}' 摘要時發生錯誤: {xml_e}")

                        if os.path.exists(project_status_file):
                            try:
                                project_statuses = load_statuses(item)
                                for report_data in project_statuses.values():
                                    if not isinstance(report_data, dict):
                                        continue

                                    for issue_id, status_data in report_data.items():
                                        if issue_id.startswith("_manual_") and isinstance(status_data, dict) and "manual_details" in status_data:
                                            severity_key = status_data["manual_details"].get("severity_key", "medium")
                                            if severity_key in project_severities:
                                                project_severities[severity_key] += 1
                                            else:
                                                app.logger.warning(f"專案 '{item}' 中的手動弱點 '{issue_id}' 具有未知的嚴重性 '{severity_key}'")
                            except Exception as status_e:
                                app.logger.error(f"處理專案 '{item}' 的狀態檔以統計手動弱點時發生錯誤: {status_e}")

                        projects.append({
                            "name": item,
                            "display_name": display_name,
                            "stats": project_severities
                        })

                    except Exception as load_err:
                        app.logger.error(f"載入專案 '{item}' 時發生錯誤: {load_err}")
                        projects.append({
                            "name": item,
                            "display_name": f"{item} (載入錯誤)",
                            "stats": {level: 0 for level in SEVERITY_LEVELS.keys()}
                        })

            projects.sort(key=lambda p: p["display_name"])

        except Exception as e:
            flash(f"讀取專案列表錯誤: {e}", "danger")
            app.logger.error(f"列出專案時發生錯誤: {e}")

    return render_template(
        "project_select.html",
        projects=projects,
        BASE_REPORT_FOLDER=os.path.abspath(BASE_REPORT_FOLDER),
        severity_levels=SEVERITY_LEVELS,
        severity_display_map=SEVERITY_DISPLAY_MAP
    )


# 路由：單個專案的主頁面
@app.route("/project/<project_name>/")
def project_index(project_name):
    """顯示特定專案的主介面"""
    if not is_safe_project_name(project_name):
        abort(400, "無效的專案名稱。")

    report_folder = get_project_report_folder(project_name)
    if not os.path.isdir(report_folder):
        app.logger.error(f"找不到報告資料夾: {report_folder}")
        flash(f"專案 '{project_name}' 報告資料夾不存在。", "warning")
        return redirect(url_for('list_projects'))

    target_details = read_target_details_from_excel(project_name)
    if not target_details:
        target_file_path = get_target_list_file(project_name)
        error_message = f"無法載入專案 '{project_name}'：必要的 '{os.path.basename(target_file_path)}' 檔案遺失或無效。"
        app.logger.error(f"阻止進入專案 '{project_name}': 缺少目標清單。")
        flash(error_message, "danger")
        return redirect(url_for("list_projects"))

    weakness_names = load_weakness_list()
    common_notes_data = load_common_notes() # <<< NEW: Load common notes
    config = load_project_config(project_name)
    display_name = config.get("project_display_name", project_name)

    return render_template(
        "index.html",
        project_name=project_name,
        project_display_name=display_name,
        severities=SEVERITY_DISPLAY_MAP,
        status_options=STATUS_OPTIONS,
        severity_levels_map=SEVERITY_LEVELS,
        screenshot_filter_options=SCREENSHOT_FILTER_OPTIONS,
        auto_excluded_status_value=AUTO_EXCLUDED_STATUS,
        weakness_name_list=weakness_names,
        external_link_reasoning_text=EXTERNAL_LINK_REASONING,
        common_notes=common_notes_data              # <<< NEW: Pass common notes data
    )


# API 路由：獲取特定問題類型的修補建議
@app.route("/project/<project_name>/api/advisory/<path:issue_type>")
def get_advisory(project_name, issue_type):
    """根據問題類型名稱，從 advisory.json 查找修補建議"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    try:
        decoded_type = unquote(issue_type)
        advisory_data = load_advisory_data(project_name)
        suggestion = advisory_data.get(decoded_type)

        if suggestion:
            return jsonify({"suggestion": suggestion})
        else:
            # 如果直接匹配不到，嘗試不區分大小寫匹配
            for key, value in advisory_data.items():
                if key.lower() == decoded_type.lower():
                    return jsonify({"suggestion": value})

            # 如果都找不到
            app.logger.info(f"找不到問題類型 '{decoded_type}' 的修補建議。")
            return jsonify({"suggestion": None, "message": "找不到修補建議。"}), 404

    except Exception as e:
        app.logger.error(f"獲取修補建議時發生錯誤: {e}")
        return jsonify({"error": "伺服器錯誤。"}), 500


# API 路由：獲取報告列表
@app.route("/project/<project_name>/api/reports")
def get_report_list(project_name):
    """獲取專案的報告列表，包含狀態、摘要等資訊，並根據前端篩選條件過濾"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    report_folder = get_project_report_folder(project_name)
    if not os.path.isdir(report_folder):
        app.logger.error(f"找不到報告資料夾: {report_folder}")
        return jsonify({"error": "找不到專案報告資料夾。"}), 404

    report_filter = request.args.get("filter", "low").lower()
    app.logger.info(f"API get_report_list: 專案='{project_name}', 篩選='{report_filter}'")

    try:
        statuses = load_statuses(project_name)
        target_details = read_target_details_from_excel(project_name)
        expected_numbers = sorted(list(target_details.keys()))
        expected_set = set(expected_numbers)
    except Exception as e:
        app.logger.error(f"載入狀態/目標數據時發生錯誤: {e}")
        return jsonify({"error": "伺服器載入數據時發生錯誤。"}), 500

    report_list = []
    processed_numbers = set()
    problematic_statuses_list = [
        SCAN_STATUS_MAP.get("Failed", "失敗"), PARSE_ERROR_STATUS, READ_ERROR_STATUS,
        INCOMPLETE_STATUS, FILE_NOT_FOUND_STATUS, MISSING_FILE_STATUS,
        DEFAULT_SCAN_STATUS, SCAN_STATUS_MAP.get("Aborted", "已中斷")
    ]

    try:
        xml_files = [f for f in os.listdir(report_folder) if f.lower().endswith(".xml") and os.path.isfile(os.path.join(report_folder, f))]

        for filename in xml_files:
            filepath = os.path.join(report_folder, filename)
            file_number = None
            match = re.match(r"(\d+)-.*\.xml", filename, re.IGNORECASE)

            if match:
                try:
                    file_number = int(match.group(1))
                    if file_number in expected_set:
                        processed_numbers.add(file_number)
                    else:
                        app.logger.warning(f"報告 '{filename}' 的編號 {file_number} 不在目標清單中。")
                except ValueError:
                    file_number = None

            scan_status = get_scan_status(filepath)
            severity_summary = None
            scan_stats = { k: "N/A" for k in ["pages_scanned", "total_pages", "entities_tested", "total_entities", "issues_found"] }
            is_error_or_abnormal_scan = scan_status in problematic_statuses_list

            if scan_status not in [PARSE_ERROR_STATUS, READ_ERROR_STATUS, FILE_NOT_FOUND_STATUS]:
                try:
                    severity_summary = get_report_summary_severities(filepath)

                    if severity_summary is not None:
                        try:
                            # 重新解析以獲取統計數據 (這裡可以優化，避免重複解析)
                            tree = ET.parse(filepath)
                            summary_elem = tree.find("scan-summary")

                            # --- MODIFIED: Use 'is not None' ---
                            if summary_elem is not None:
                                scan_stats_map = { "pages_scanned": "num-pages-scanned", "total_pages": "total-num-pages", "entities_tested": "num-security-entities-tested", "total_entities": "total-num-security-entities", "issues_found": "num-issues-found" }
                                for k, t in scan_stats_map.items():
                                    scan_stats[k] = summary_elem.findtext(t, "N/A")

                                # 如果 issues_found 為 N/A，用摘要計算值
                                if scan_stats["issues_found"] == "N/A":
                                    # 確保 severity_summary 不是 None
                                    if severity_summary:
                                       scan_stats["issues_found"] = str(sum(severity_summary.values()))
                                    else:
                                       scan_stats["issues_found"] = "0" # Fallback if summary exists but counts failed
                            # 如果沒有 summary 標籤，但有嚴重性摘要，也用計算值
                            elif scan_stats["issues_found"] == "N/A":
                                # 確保 severity_summary 不是 None
                                if severity_summary:
                                    scan_stats["issues_found"] = str(sum(severity_summary.values()))
                                else:
                                    scan_stats["issues_found"] = "0" # Fallback

                        except Exception as stat_e:
                            app.logger.error(f"獲取報告 '{filename}' 的掃描統計時發生錯誤: {stat_e}")
                except Exception as e:
                    app.logger.error(f"獲取報告 '{filename}' 的摘要時發生錯誤: {e}")
                    scan_status = READ_ERROR_STATUS
                    severity_summary = None
                    is_error_or_abnormal_scan = True

            # --- 判斷報告是否滿足前端的篩選條件 ---
            meets_filter_criteria = False
            if report_filter == "errors_only":
                meets_filter_criteria = is_error_or_abnormal_scan
            elif report_filter == "informational":
                meets_filter_criteria = True
            else:
                min_level = SEVERITY_LEVELS.get(report_filter, 1)
                has_qualifying = False
                if severity_summary:
                    has_qualifying = any(
                        (lv := SEVERITY_LEVELS.get(lk)) is not None and lv >= min_level and c > 0
                        for lk, c in severity_summary.items()
                    )
                meets_filter_criteria = has_qualifying or is_error_or_abnormal_scan

            review_completed = statuses.get(filename, {}).get(REPORT_COMPLETED_KEY, False)

            report_list.append({
                "filename": filename,
                "status": scan_status,
                "severity_summary": severity_summary or {k: 0 for k in SEVERITY_LEVELS},
                "scan_stats": scan_stats,
                "file_number": file_number,
                "is_missing": False,
                "review_completed": review_completed,
                "meets_threshold": meets_filter_criteria
            })

        # --- 處理遺失的報告 ---
        if expected_numbers:
            missing_numbers = sorted(list(expected_set - processed_numbers))
            for m_num in missing_numbers:
                missing_filename = f"{m_num}-找不到掃描檔"
                report_list.append({
                    "filename": missing_filename,
                    "status": FILE_NOT_FOUND_STATUS,
                    "severity_summary": {k: 0 for k in SEVERITY_LEVELS},
                    "scan_stats": {k: "N/A" for k in scan_stats},
                    "file_number": m_num,
                    "is_missing": True,
                    "review_completed": False,
                    "meets_threshold": True # 遺失報告總是滿足篩選條件
                })
        elif not expected_numbers:
             app.logger.warning(f"專案 '{project_name}': 目標清單為空或遺失。")

        report_list.sort(key=lambda item: item.get("file_number") if isinstance(item.get("file_number"), int) else float("inf"))
        app.logger.info(f"API get_report_list: 返回 {len(report_list)} 個報告。")
        return jsonify(report_list)

    except Exception as e:
        app.logger.error(f"處理專案 '{project_name}' 的報告列表時發生錯誤: {traceback.format_exc()}")
        return jsonify({"error": f"伺服器錯誤: {e}"}), 500

# API 路由：獲取單個報告的詳細數據
@app.route("/project/<project_name>/report/<path:filename>")
def get_report_data(project_name, filename):
    """獲取指定報告檔案的詳細內容，包含問題、狀態、筆記、截圖備註、目標資訊等"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    app.logger.info(f"API get_report_data: 專案='{project_name}', 報告='{filename}'")
    report_folder = get_project_report_folder(project_name)
    original_basename = os.path.basename(filename)
    secure_fn = secure_filename(original_basename)
    if secure_fn != original_basename:
        app.logger.warning(f"檔名可能不安全: 原始='{original_basename}', 清理後='{secure_fn}'")

    is_placeholder = original_basename.endswith("-找不到掃描檔")
    filepath = os.path.abspath(os.path.join(report_folder, original_basename))

    # --- 解析報告編號 (用於查找目標資訊 和 截圖前綴) ---
    report_num_match = re.match(r"(\d+)-.*", original_basename)
    report_num_int = "unknown" # 用於查找 target_details 的整數版本
    report_num_str = "unknown" # 用於生成檔名前綴的字串版本

    if report_num_match:
        report_num_str = report_num_match.group(1) # 直接獲取原始字串 (例如 "08")
        try:
            report_num_int = int(report_num_str) # 嘗試轉換為整數，用於查找 Excel
        except ValueError:
            app.logger.warning(f"無法從檔名 '{original_basename}' 解析出有效的數字編號 (用於查找目標)。")
            report_num_int = "unknown" # 整數版本設為 unknown
            # 注意：report_num_str 仍然保留原始字串，即使無法轉為 int
    else:
        # 如果檔名不匹配模式，兩者都保持 "unknown"
        app.logger.warning(f"無法從檔名 '{original_basename}' 的開頭解析出編號。")

    app.logger.debug(f"Parsed report number: Int='{report_num_int}', Str='{report_num_str}'")
    # --- END 解析報告編號 ---

    # --- 初始化預設值 ---
    parsed_data = None
    # 使用 report_num_str (如果有效) 或 report_num_int 來創建 base_filename
    base_filename_for_scan_info = report_num_str if report_num_str != "unknown" else str(report_num_int) if report_num_int != "unknown" else "unknown"
    scan_info = {"scan_name": original_basename, "scan_date": "N/A", "base_filename": base_filename_for_scan_info}
    summary = {k: "0" for k in SEVERITY_SUMMARY_ORDER + ["info", "total_issues"]}
    stats = {k: "N/A" for k in ["pages_scanned", "total_pages", "entities_tested", "total_entities", "issues_found"]}
    target_url_for_report = "N/A"
    target_name_for_report = "N/A"
    # --- END 初始化預設值 ---

    # --- 處理不同情況下的報告數據 ---
    if is_placeholder:
        stats = {k: "遺失" for k in stats}
        parsed_data = {"scan_info": scan_info, "issues": [], "summary": summary, "scan_stats": stats}
        app.logger.info(f"處理佔位符報告: {original_basename}")
    elif not filepath.startswith(os.path.abspath(report_folder)):
        app.logger.error(f"拒絕存取嘗試: '{filepath}'")
        return jsonify({"error": "拒絕存取。"}), 403
    elif not os.path.isfile(filepath):
        app.logger.warning(f"找不到報告檔案: {filepath}")
        stats = {k: "遺失" for k in stats}
        parsed_data = {"scan_info": scan_info, "issues": [], "summary": summary, "scan_stats": stats}
    else:
        try:
            parsed_data = parse_appscan_xml(project_name, filepath, original_basename)
            if parsed_data is None:
                app.logger.error(f"XML 解析失敗: {original_basename}")
                stats = {k: "錯誤" for k in stats}
                parsed_data = {"scan_info": scan_info, "issues": [], "summary": summary, "scan_stats": stats}
            else:
                # 如果解析成功，更新 scan_info, summary, stats
                scan_info = parsed_data.get("scan_info", scan_info)
                summary = parsed_data.get("summary", summary)
                stats = parsed_data.get("scan_stats", stats)
        except Exception as e:
            app.logger.error(f"解析 XML '{original_basename}' 時發生錯誤: {traceback.format_exc()}")
            return jsonify({"error": f"伺服器解析 XML 時發生錯誤: {e}"}), 500
    # --- END 處理報告數據 ---

    try:
        # --- 載入目標詳情 (使用整數版本 report_num_int) ---
        try:
            target_details = read_target_details_from_excel(project_name)
        except Exception as excel_e:
             app.logger.error(f"讀取 target.xlsx 時發生錯誤: {excel_e}")
             target_details = {} # 出錯時設為空字典

        if report_num_int != "unknown" and isinstance(report_num_int, int): # 使用 report_num_int
            target_info = target_details.get(report_num_int)
            if target_info:
                target_url_for_report = target_info.get("url", "N/A")
                target_name_for_report = target_info.get("name", "N/A")
                app.logger.debug(f"找到編號 {report_num_int} 的目標資訊: URL='{target_url_for_report}', Name='{target_name_for_report}'")
            else:
                app.logger.warning(f"在 target.xlsx 中找不到報告編號 {report_num_int} 的目標詳情。")
        elif is_placeholder: # 嘗試為佔位符也查找資訊 (使用 report_num_int)
             if report_num_int != "unknown" and isinstance(report_num_int, int):
                 target_info = target_details.get(report_num_int)
                 if target_info:
                     target_url_for_report = target_info.get("url", "N/A")
                     target_name_for_report = target_info.get("name", "N/A")
                     app.logger.debug(f"找到佔位符報告編號 {report_num_int} 的目標資訊。")
                 else:
                     app.logger.warning(f"在 target.xlsx 中找不到佔位符報告編號 {report_num_int} 的目標詳情。")
        else:
             app.logger.warning(f"報告編號 '{report_num_int}' 無效或未知，無法查找目標詳情。")
        # --- END 載入目標詳情 ---

        project_statuses = load_statuses(project_name)
        processed_issues = [] # 儲存所有處理過的 issue (過濾前)
        statuses_modified = False # 標記狀態檔是否需要儲存
        report_statuses_entry = project_statuses.get(original_basename, {}) # 獲取此報告的狀態條目

        # 確保報告條目存在於 project_statuses 中，即使是空的
        if original_basename not in project_statuses:
            project_statuses[original_basename] = {}
            report_statuses_entry = project_statuses[original_basename]
            # report_statuses_entry[REPORT_COMPLETED_KEY] = False # 初始化完成狀態 (可選)
            statuses_modified = True # 因為新增了報告條目，所以標記為已修改

        manual_override_statuses = { STATUS_OPTIONS["誤判"], STATUS_OPTIONS["已確認弱點"], STATUS_OPTIONS["人工審查中"] }
        # 初始化此報告最終要儲存的狀態 (包含判讀完成標記)
        updated_report_statuses = {REPORT_COMPLETED_KEY: report_statuses_entry.get(REPORT_COMPLETED_KEY, False)}

        # --- Process AppScan issues ---
        appscan_issues_count = 0
        if parsed_data and parsed_data.get("issues"):
            appscan_issues_count = len(parsed_data["issues"])
            for issue in parsed_data["issues"]: # 遍歷從 XML 解析出的 issue
                issue_id = issue.get("id")
                if issue.get("source") == "error": # 跳過解析錯誤的標記
                    issue["status"] = issue.get("status", "處理錯誤")
                    issue["screenshot_taken"] = issue.get("screenshot_taken", False)
                    issue["note"] = issue.get("note", "XML解析錯誤")
                    processed_issues.append(issue) # 加入列表以便計算總數
                    continue

                if not issue_id:
                    app.logger.warning(f"報告 {original_basename} 中發現缺少 ID 的 AppScan 問題，跳過。")
                    continue

                # --- 獲取或初始化此問題的狀態資訊 ---
                status_info = report_statuses_entry.get(issue_id) # 從已載入的狀態中獲取
                issue_status_changed_in_code = False # 標記此 issue 狀態是否在本輪處理中改變

                if not isinstance(status_info, dict):
                    # 如果狀態不是字典 (舊格式)，轉換
                    saved_status_val = str(status_info or DEFAULT_STATUS)
                    status_info = {
                        "status": saved_status_val,
                        "screenshot_taken": False,
                        "note": DEFAULT_NOTE,
                        "screenshots_meta": {}
                    }
                    app.logger.debug(f"將報告 {original_basename} 問題 {issue_id} 的舊狀態 '{saved_status_val}' 轉換為新格式。")
                    issue_status_changed_in_code = True # 因為格式轉換了，需要儲存
                else:
                    # 如果是字典，確保必要欄位存在
                    if "status" not in status_info: status_info["status"] = DEFAULT_STATUS; issue_status_changed_in_code = True
                    if "screenshot_taken" not in status_info: status_info["screenshot_taken"] = False; issue_status_changed_in_code = True
                    if "note" not in status_info: status_info["note"] = DEFAULT_NOTE; issue_status_changed_in_code = True
                    if "screenshots_meta" not in status_info or not isinstance(status_info["screenshots_meta"], dict):
                        status_info["screenshots_meta"] = {}; issue_status_changed_in_code = True

                # 讀取當前儲存的值
                saved_status = status_info["status"]
                saved_screenshot_taken = status_info["screenshot_taken"]
                saved_note = status_info["note"]
                screenshots_meta = status_info["screenshots_meta"] # 確定 screenshots_meta 是字典

                # --- 計算顯示狀態 (應用排除規則) ---
                display_status = saved_status
                display_screenshot_taken = saved_screenshot_taken

                if saved_status not in manual_override_statuses:
                    issue_type = issue.get("issue_type")
                    entity_name = issue.get("entity_name", "")
                    matches_a_rule = False
                    for rule in exclusion_rules:
                         rule_match_type = rule.get("match_type", "entity_starts_with")
                         rule_issue_type = rule.get("issue_type")
                         if not rule_issue_type or issue_type != rule_issue_type: continue
                         rule_applied = False
                         if rule_match_type == "issue_type_only": rule_applied = True
                         elif rule_match_type == "entity_starts_with":
                             pattern = rule.get("entity_pattern")
                             if pattern and entity_name is not None and entity_name.startswith(pattern): rule_applied = True
                         elif rule_match_type == "entity_contains":
                             pattern = rule.get("entity_pattern")
                             if pattern and entity_name is not None and pattern in entity_name: rule_applied = True
                         if rule_applied: matches_a_rule = True; break

                    if matches_a_rule:
                        if display_status != AUTO_EXCLUDED_STATUS:
                            display_status = AUTO_EXCLUDED_STATUS
                            display_screenshot_taken = True
                            if saved_status != display_status or saved_screenshot_taken != display_screenshot_taken:
                                issue_status_changed_in_code = True
                            app.logger.debug(f"規則匹配 AppScan 問題 {issue_id}，狀態設為 {display_status}。")
                    elif saved_status == AUTO_EXCLUDED_STATUS:
                        display_status = DEFAULT_STATUS
                        display_screenshot_taken = False
                        if saved_status != display_status or saved_screenshot_taken != display_screenshot_taken:
                             issue_status_changed_in_code = True
                        app.logger.debug(f"規則不再匹配 AppScan 問題 {issue_id}，狀態恢復為 {display_status}。")

                # --- 更新 issue 物件以供前端使用 ---
                issue["status"] = display_status
                issue["screenshot_taken"] = display_screenshot_taken
                issue["note"] = saved_note
                issue["source"] = "appscan" # 確保來源

                # --- 查找並添加截圖詳細資訊 ---
                issue["screenshots_detailed"] = [] # 初始化為空列表
                sanitized_name = sanitize_filename_part(issue.get('issue_type', 'Unknown'), max_len=50)
                sanitized_url = sanitize_filename_part(issue.get('url', 'N_A'), max_len=30)
                sanitized_entity = sanitize_filename_part(issue.get('entity_name', 'N_A'), max_len=30)

                # <<< 修改點：使用 report_num_str 來生成前綴 >>>
                if report_num_str != "unknown" and report_num_str.isdigit() and sanitized_name not in ["na", "sanitized_empty"]:
                    filename_prefix = construct_screenshot_filename_prefix(
                        report_num_str, # <<< 使用字串版本
                        APPSCAN_SOURCE_LABEL,
                        sanitized_name,
                        sanitized_url,
                        sanitized_entity
                    )
                    app.logger.debug(f"查找 AppScan 截圖: ReportNumStr='{report_num_str}', Source={APPSCAN_SOURCE_LABEL}, Name='{sanitized_name}', URL='{sanitized_url}', Entity='{sanitized_entity}'")
                    app.logger.debug(f"  生成的 Prefix: '{filename_prefix}'")
                    try:
                        existing_filenames = get_existing_screenshots(project_name, filename_prefix)
                        app.logger.debug(f"  get_existing_screenshots 返回: {existing_filenames}")
                        for ss_filename in existing_filenames:
                            # 從 status_info 中讀取此檔案的備註
                            caption = screenshots_meta.get(ss_filename, {}).get("caption", "")
                            issue["screenshots_detailed"].append({"filename": ss_filename, "caption": caption})
                        if existing_filenames:
                             app.logger.debug(f"找到 {len(existing_filenames)} 個 AppScan 問題 {issue_id} 的截圖。")
                    except Exception as ss_err:
                         app.logger.error(f"查找 AppScan 問題 {issue_id} 的截圖時出錯: {ss_err}")
                else:
                    app.logger.debug(f"無法為 AppScan 問題 {issue_id} 查找截圖 (無效的報告編號字串 '{report_num_str}' 或名稱 '{sanitized_name}')。")

                # --- 更新儲存用的狀態字典 ---
                status_info["status"] = display_status
                status_info["screenshot_taken"] = display_screenshot_taken
                # note 和 screenshots_meta 在之前確保存在/轉換時已設定好
                updated_report_statuses[issue_id] = status_info # 將此 issue 的最終狀態加入儲存字典

                if issue_status_changed_in_code: statuses_modified = True
                processed_issues.append(issue) # 加入處理後的 issue 列表
        # --- End Process AppScan issues ---

        # --- Process manual issues ---
        app.logger.debug(f"正在檢查報告 {original_basename} 的手動弱點")
        manual_issue_count = 0
        for issue_id, status_data in report_statuses_entry.items():
            if not issue_id.startswith("_manual_") or issue_id == REPORT_COMPLETED_KEY:
                continue # 跳過非手動或特殊鍵

            if isinstance(status_data, dict) and isinstance(status_data.get("manual_details"), dict):
                manual_issue_count += 1
                manual_details = status_data["manual_details"]

                # 確保 screenshots_meta 存在且為字典
                screenshots_meta = status_data.get("screenshots_meta", {})
                if not isinstance(screenshots_meta, dict):
                    screenshots_meta = {}; status_data["screenshots_meta"] = screenshots_meta; statuses_modified = True # 需要儲存

                # 查找手動問題的截圖
                manual_screenshots_detailed = []
                # <<< 修改點：使用 report_num_str 來生成前綴 >>>
                if report_num_str != "unknown" and report_num_str.isdigit():
                    sanitized_name = sanitize_filename_part(manual_details.get('issue_type', 'Unknown'), max_len=50)
                    sanitized_url = sanitize_filename_part(manual_details.get('url', 'N_A'), max_len=30)
                    sanitized_entity = sanitize_filename_part(manual_details.get('entity_name', 'N/A'), max_len=30)

                    if sanitized_name not in ["na", "sanitized_empty"]:
                        filename_prefix = construct_screenshot_filename_prefix(
                            report_num_str, # <<< 使用字串版本
                            MANUAL_SOURCE_LABEL,
                            sanitized_name,
                            sanitized_url,
                            sanitized_entity
                        )
                        app.logger.debug(f"查找 Manual 截圖: ReportNumStr='{report_num_str}', Source={MANUAL_SOURCE_LABEL}, Name='{sanitized_name}', URL='{sanitized_url}', Entity='{sanitized_entity}'")
                        app.logger.debug(f"  生成的 Prefix: '{filename_prefix}'")
                        try:
                            existing_filenames = get_existing_screenshots(project_name, filename_prefix)
                            app.logger.debug(f"  get_existing_screenshots 返回: {existing_filenames}")
                            for ss_filename in existing_filenames:
                                caption = screenshots_meta.get(ss_filename, {}).get("caption", "")
                                manual_screenshots_detailed.append({"filename": ss_filename, "caption": caption})
                            if existing_filenames:
                                 app.logger.debug(f"找到 {len(existing_filenames)} 個手動問題 {issue_id} 的截圖。")
                        except Exception as ss_err:
                             app.logger.error(f"查找手動問題 {issue_id} 的截圖時出錯: {ss_err}")
                    else:
                        app.logger.warning(f"無法為手動問題 {issue_id} 查找截圖 (無效的名稱 '{sanitized_name}')。")
                else:
                     app.logger.warning(f"無法為手動問題 {issue_id} 查找截圖 (無效的報告編號字串 '{report_num_str}')。")

                # 構建手動問題物件以供前端使用
                manual_issue = {
                    "id": issue_id,
                    "status": status_data.get("status", DEFAULT_STATUS),
                    "screenshot_taken": status_data.get("screenshot_taken", False),
                    "note": status_data.get("note", DEFAULT_NOTE),
                    "source": "manual",
                    **manual_details, # 合併 manual_details 字典
                    "severity_display": manual_details.get("severity_display", SEVERITY_DISPLAY_MAP.get(manual_details.get("severity_key", "medium"), "未知")),
                    "screenshots_detailed": manual_screenshots_detailed, # 加入截圖列表
                    # 添加 AppScan issue 有但手動 issue 通常沒有的欄位，設為預設值
                    "cvss_score": "N/A", "cve_name": None, "cve_url": None, "http_traffic": "N/A",
                    "scan_info": scan_info # 使用從 XML 或預設獲取的 scan_info
                }
                processed_issues.append(manual_issue)
                updated_report_statuses[issue_id] = status_data # 將手動問題的狀態數據加入儲存字典
            else:
                app.logger.warning(f"手動弱點條目 {issue_id} (在報告 {original_basename} 中) 格式無效或缺少 'manual_details'。跳過。")
        # --- End Process manual issues ---

        app.logger.debug(f"報告 {original_basename}: 處理了 {appscan_issues_count} 個 AppScan 問題和 {manual_issue_count} 個手動問題。")

        # --- Save updated statuses if modified ---
        report_needs_saving = statuses_modified # 如果有任何轉換或規則應用，就儲存
        if report_needs_saving:
            app.logger.info(f"報告 '{original_basename}' 的狀態需要儲存 (Modified={statuses_modified})。")
            project_statuses[original_basename] = updated_report_statuses # 確保使用更新後的字典
            save_statuses(project_name, project_statuses)
        else:
             app.logger.debug(f"報告 '{original_basename}' 的狀態無需儲存。")

        # --- Filter issues based on request ---
        filtered_issues = list(processed_issues) # 複製一份用於過濾
        severity_filter = request.args.getlist("severity")
        status_filter = request.args.getlist("status_filter")
        screenshot_filter = request.args.getlist("screenshot_status_filter")
        source_filter = request.args.getlist("source_filter")

        # --- (過濾邏輯保持不變) ---
        if severity_filter:
            valid_severities = {s.lower() for s in severity_filter if s.lower() in SEVERITY_LEVELS}
            if valid_severities: filtered_issues = [i for i in filtered_issues if i.get("severity_key") in valid_severities]
        if status_filter:
            valid_statuses = set(status_filter)
            if valid_statuses: filtered_issues = [i for i in filtered_issues if i.get("status") in valid_statuses]
        if screenshot_filter:
            valid_ss_filters = set(screenshot_filter) & set(SCREENSHOT_FILTER_OPTIONS.keys())
            if valid_ss_filters and len(valid_ss_filters) < len(SCREENSHOT_FILTER_OPTIONS):
                temp_filtered = []
                for issue in filtered_issues:
                    include = False; ss_taken = issue.get("screenshot_taken", False); has_files = bool(issue.get("screenshots_detailed"))
                    if "completed" in valid_ss_filters and ss_taken: include = True
                    if not include and "not_completed" in valid_ss_filters and not ss_taken: include = True
                    if not include and "has_files" in valid_ss_filters and has_files: include = True
                    if not include and "no_files" in valid_ss_filters and not has_files: include = True
                    if include: temp_filtered.append(issue)
                filtered_issues = temp_filtered
        if source_filter:
             valid_sources = {s for s in source_filter if s in ["appscan", "manual"]}
             if valid_sources and len(valid_sources) < 2: # 僅當只選了一個來源時才過濾
                 filtered_issues = [i for i in filtered_issues if i.get("source", "appscan") in valid_sources]

        # 過濾掉錯誤源的問題
        final_filtered_issues = [i for i in filtered_issues if i.get("source") != "error"]

        # --- Calculate final status summary based on *all* processed issues ---
        final_status_summary = {status_display: 0 for status_display in STATUS_OPTIONS.values()}
        final_status_summary["處理錯誤"] = 0; final_status_summary["__unexpected__"] = 0
        for issue in processed_issues: # Use the list *before* filtering
            status_key = issue.get("status"); source = issue.get("source")
            if source == "error": final_status_summary["處理錯誤"] += 1
            elif status_key in STATUS_OPTIONS: final_status_summary[STATUS_OPTIONS[status_key]] += 1
            elif status_key: final_status_summary["__unexpected__"] += 1
        final_status_summary = {k: v for k, v in final_status_summary.items() if v > 0 or k == "__unexpected__"}
        if final_status_summary.get("__unexpected__", 0) == 0: final_status_summary.pop("__unexpected__", None)

        # --- Build final response data ---
        final_data = {
            "scan_info": scan_info,
            "summary": summary,
            "issues": final_filtered_issues, # Return the *filtered* list
            "status_summary": final_status_summary, # Return the *unfiltered* summary
            "scan_stats": stats,
            "target_url": target_url_for_report, # Add target info
            "target_name": target_name_for_report
        }
        app.logger.info(f"返回報告 '{original_basename}' 的 {len(final_filtered_issues)} 個篩選後問題 (總共 {len(processed_issues)} 個)。")
        return jsonify(final_data)

    except Exception as e:
        app.logger.error(f"處理報告數據 '{original_basename}' 時發生錯誤: {traceback.format_exc()}")
        return jsonify({"error": f"伺服器錯誤: {e}"}), 500


# API 路由：批次更新弱點狀態或截圖完成標記
@app.route("/project/<project_name>/api/batch_update_status", methods=["POST"])
def batch_update_vulnerability_status(project_name):
    """處理前端發送的批次更新請求 (用於群組操作)"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.get_json()
    updates_list = data.get("updates")
    new_status = data.get("status")
    new_screenshot_taken = data.get("screenshot_taken")
    app.logger.info(f"API batch_update_status: 專案='{project_name}', 項目數={len(updates_list) if updates_list else 0}, 狀態='{new_status}', 截圖完成='{new_screenshot_taken}'")

    if not updates_list or not isinstance(updates_list, list):
        return jsonify({"error": "缺少 'updates' 列表。"}), 400

    if new_status is None and new_screenshot_taken is None:
        return jsonify({"error": "請求必須包含 'status' 或 'screenshot_taken'。"}), 400

    if new_status is not None:
        if new_status not in STATUS_OPTIONS:
            return jsonify({"error": f"無效的狀態: '{new_status}'"}), 400
        if new_status == AUTO_EXCLUDED_STATUS:
            return jsonify({"error": f"無法手動設定狀態為 '{AUTO_EXCLUDED_STATUS}'。"}), 400

    if new_screenshot_taken is not None and not isinstance(new_screenshot_taken, bool):
        return jsonify({"error": "無效的 'screenshot_taken' 值。"}), 400

    updated_count, skipped_missing, modified = 0, 0, False
    reports_to_save = set()

    try:
        statuses = load_statuses(project_name)
        for item in updates_list:
            report_filename = item.get("reportFilename")
            issue_id = item.get("issueId")

            if not report_filename or not issue_id:
                app.logger.warning(f"跳過無效的批次項目: {item}")
                continue

            if report_filename.endswith("-找不到掃描檔"):
                app.logger.warning(f"跳過遺失報告的佔位符 '{report_filename}'。")
                skipped_missing += 1
                continue

            report_entry = statuses.setdefault(report_filename, {})
            issue_entry = report_entry.get(issue_id)

            if not isinstance(issue_entry, dict):
                app.logger.warning(f"報告 {report_filename} 中的問題 {issue_id} 不是字典格式。將創建預設值。")
                issue_entry = { "status": str(issue_entry or DEFAULT_STATUS), "screenshot_taken": False, "note": DEFAULT_NOTE, "screenshots_meta": {} } # Initialize meta
                if issue_id.startswith("_manual_") and isinstance(report_entry.get(issue_id), dict):
                    original_data = report_entry.get(issue_id)
                    if original_data and "manual_details" in original_data:
                        issue_entry.update(original_data)
                    if "screenshots_meta" in original_data and isinstance(original_data["screenshots_meta"], dict):
                        issue_entry["screenshots_meta"] = original_data["screenshots_meta"]

            item_modified = False
            if new_status is not None and issue_entry.get("status") != new_status:
                issue_entry["status"] = new_status
                item_modified = True
            if new_screenshot_taken is not None and issue_entry.get("screenshot_taken") != new_screenshot_taken:
                issue_entry["screenshot_taken"] = new_screenshot_taken
                item_modified = True

            if item_modified:
                report_entry[issue_id] = issue_entry
                updated_count += 1
                modified = True
                reports_to_save.add(report_filename)

        if modified:
             for fname in reports_to_save:
                 statuses[fname] = statuses.get(fname, {})
             save_statuses(project_name, statuses)
             app.logger.info(f"批次更新成功: 更新 {updated_count} 個項目, 跳過 {skipped_missing} 個遺失項目。")
             return jsonify({ "message": f"成功批次更新 {updated_count} 個項目。", "skipped_missing": skipped_missing, "updated_count": updated_count }), 200
        else:
             app.logger.info(f"批次更新: 無需變更。跳過 {skipped_missing} 個遺失項目。")
             return jsonify({ "message": "無需變更。", "skipped_missing": skipped_missing, "updated_count": 0 }), 200

    except Exception as e:
        app.logger.error(f"批次更新期間發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "伺服器批次更新期間發生錯誤。"}), 500


# API 路由：更新單個弱點的筆記
@app.route("/project/<project_name>/api/note", methods=["POST"])
def update_vulnerability_note(project_name):
    """處理前端發送的更新單個弱點筆記的請求"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.get_json()
    report_filename = data.get("reportFilename")
    issue_id = data.get("issueId")
    new_note = data.get("note", "")
    app.logger.info(f"API update_note: 專案='{project_name}', 報告='{report_filename}', 問題='{issue_id}'")

    if not report_filename or not issue_id:
        return jsonify({"error": "缺少必要數據: reportFilename, issueId"}), 400

    if not isinstance(new_note, str):
        new_note = str(new_note)
        app.logger.warning(f"將報告 {report_filename}/{issue_id} 的非字串筆記強制轉換。")

    if report_filename.endswith("-找不到掃描檔"):
        return jsonify({"error": "無法更新遺失報告檔案的筆記。"}), 400

    try:
        statuses = load_statuses(project_name)
        report_entry = statuses.setdefault(report_filename, {})
        issue_entry = report_entry.get(issue_id)

        if not isinstance(issue_entry, dict):
            app.logger.warning(f"報告 {report_filename} 中的問題 {issue_id} 不是字典格式。將創建預設值。")
            issue_entry = { "status": str(issue_entry or DEFAULT_STATUS), "screenshot_taken": False, "note": DEFAULT_NOTE, "screenshots_meta": {} } # Initialize meta
            if issue_id.startswith("_manual_") and isinstance(report_entry.get(issue_id), dict):
                original_data = report_entry.get(issue_id)
                if original_data and "manual_details" in original_data:
                    issue_entry.update(original_data)
                if "screenshots_meta" in original_data and isinstance(original_data["screenshots_meta"], dict):
                    issue_entry["screenshots_meta"] = original_data["screenshots_meta"]

        issue_entry["note"] = new_note
        report_entry[issue_id] = issue_entry
        save_statuses(project_name, statuses)
        app.logger.debug(f"報告 {report_filename}/{issue_id} 的筆記已成功更新。")
        return jsonify({"message": "筆記已成功更新。"}), 200

    except Exception as e:
        app.logger.error(f"更新報告 {report_filename} 中問題 {issue_id} 的筆記時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "伺服器更新筆記時發生錯誤。"}), 500


# API 路由：更新報告的判讀完成狀態
@app.route("/project/<project_name>/api/report_completion_status", methods=["POST"])
def update_report_completion_status(project_name):
    """處理前端發送的更新報告判讀完成狀態的請求"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.get_json()
    report_filename = data.get("reportFilename")
    is_completed = data.get("isCompleted")

    if not report_filename or not isinstance(is_completed, bool):
        return jsonify({"error": "缺少或無效的數據: reportFilename, isCompleted。"}), 400

    if report_filename.endswith("-找不到掃描檔"):
        app.logger.info(f"跳過遺失報告佔位符的完成狀態更新: {report_filename}")
        return jsonify({"message": "已跳過遺失檔案佔位符的更新。"}), 200

    try:
        statuses = load_statuses(project_name)
        report_entry = statuses.setdefault(report_filename, {})
        report_entry[REPORT_COMPLETED_KEY] = is_completed
        save_statuses(project_name, statuses)
        app.logger.info(f"報告 '{report_filename}' 的判讀完成狀態已更新為 {is_completed}。")
        return jsonify({"message": "報告判讀完成狀態已更新。"}), 200

    except Exception as e:
        app.logger.error(f"更新報告 {report_filename} 的完成狀態時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "伺服器錯誤。"}), 500


# API 路由：處理截圖上傳
@app.route("/project/<project_name>/api/upload_screenshot", methods=["POST"])
def upload_screenshot(project_name):
    """處理前端透過表單上傳的截圖檔案，包含備註"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    # 從表單數據中獲取資訊
    report_filename = request.form.get("reportFilename")
    issue_id = request.form.get("issueId")
    issue_name = request.form.get("issueName")
    image_file = request.files.get("imageFile")
    image_caption = request.form.get("imageCaption", "") # <<< NEW: Get caption
    source = request.form.get('source', APPSCAN_SOURCE_LABEL)
    issue_url_raw = request.form.get('issueUrl', 'N_A')
    entity_name_raw = request.form.get('entityName', 'N_A')

    app.logger.info(f"收到截圖上傳請求: Proj={project_name}, Report={report_filename}, IssueID={issue_id}, Name={issue_name}, Source={source}, URL={issue_url_raw}, Entity={entity_name_raw}, Caption='{image_caption[:20]}...'")

    # --- 基本驗證 ---
    if not all([report_filename, issue_id, issue_name]):
        return jsonify({"error": "缺少表單數據。"}), 400
    if not image_file:
        return jsonify({"error": "沒有圖片檔案。"}), 400
    if report_filename.endswith("-找不到掃描檔"):
        return jsonify({"error": "無法為遺失的報告上傳截圖。"}), 400

    # --- 檔名和副檔名檢查 ---
    match = re.match(r"(\d+)-.*", report_filename)
    if not match:
        return jsonify({"error": "無效的報告檔名格式。"}), 400
    report_num = match.group(1)

    sanitized_issue_name = sanitize_filename_part(issue_name, max_len=50)
    if not sanitized_issue_name or sanitized_issue_name in ["na", "sanitized_empty"]:
        app.logger.error(f"無效的清理後弱點名稱: '{issue_name}' -> '{sanitized_issue_name}'")
        return jsonify({"error": "無效的弱點名稱。"}), 400

    original_filename = secure_filename(image_file.filename)
    if not original_filename or not allowed_file(original_filename):
        return jsonify({"error": f"無效的檔案類型。允許: {', '.join(ALLOWED_EXTENSIONS)}"}), 400
    file_extension = original_filename.rsplit(".", 1)[1].lower()

    try:
        # --- 構造檔名前綴 ---
        sanitized_url = sanitize_filename_part(issue_url_raw, max_len=30)
        sanitized_entity = sanitize_filename_part(entity_name_raw, max_len=30)
        source_label = MANUAL_SOURCE_LABEL if source.lower() == "manual" else APPSCAN_SOURCE_LABEL
        filename_prefix = construct_screenshot_filename_prefix(
            report_num, source_label, sanitized_issue_name, sanitized_url, sanitized_entity
        )
        app.logger.debug(f"構造的截圖檔名前綴: {filename_prefix}")

        # --- 獲取唯一的檔名和儲存路徑 ---
        screenshot_dir = get_project_screenshot_folder(project_name, ensure_exists=True)
        save_path = None
        final_filename = None
        max_retries = 5
        for attempt in range(max_retries):
            sequence_num = get_next_screenshot_sequence(project_name, filename_prefix)
            final_filename = f"{filename_prefix}{sequence_num}.{file_extension}"
            save_path = os.path.join(screenshot_dir, final_filename)
            if not os.path.exists(save_path):
                break # 找到可用路徑，跳出迴圈
            elif attempt < max_retries - 1:
                app.logger.warning(f"檔名衝突嘗試 {attempt+1}: {final_filename}。正在重試...")
                time.sleep(0.1 * (attempt + 1))
                save_path = None # 重設以便重新產生
            else:
                app.logger.error(f"嘗試 {max_retries} 次後無法為前綴產生唯一檔名: {filename_prefix}")
                raise IOError("無法產生唯一檔名。")

        if save_path is None:
            raise IOError("檔名產生失敗。")

        # --- 儲存圖片檔案 ---
        image_file.save(save_path)
        app.logger.info(f"截圖已儲存: {save_path}")

        # --- NEW: 更新狀態檔以包含備註 ---
        try:
            statuses = load_statuses(project_name)
            report_entry = statuses.setdefault(report_filename, {})
            # 確保 issue entry 存在，如果不存在則初始化
            issue_entry = report_entry.setdefault(issue_id, {
                "status": DEFAULT_STATUS,
                "screenshot_taken": False, # 這裡可能需要根據邏輯調整，通常上傳後應為 True?
                "note": DEFAULT_NOTE,
                "source": source.lower(),
                "screenshots_meta": {} # 確保 meta 初始化
            })
            # 確保 screenshots_meta 存在且為字典
            screenshots_meta = issue_entry.setdefault("screenshots_meta", {})
            if not isinstance(screenshots_meta, dict):
                app.logger.warning(f"重設問題 {issue_id} 的無效 screenshots_meta")
                screenshots_meta = {}
                issue_entry["screenshots_meta"] = screenshots_meta

            # 添加或更新此檔案的備註
            screenshots_meta[final_filename] = {"caption": image_caption}

            save_statuses(project_name, statuses)
            app.logger.info(f"已為截圖 '{final_filename}' 儲存備註 '{image_caption}'")

        except Exception as status_update_err:
            # 記錄錯誤，但不讓整個上傳失敗
            app.logger.error(f"上傳截圖後更新狀態檔時發生錯誤 ({final_filename}): {status_update_err}")

        # 返回檔名和備註 (MODIFIED)
        return jsonify({
            "message": "截圖已成功上傳。",
            "filename": final_filename,
            "caption": image_caption
        }), 200

    except Exception as e:
        app.logger.error(f"儲存專案 {project_name}/{report_filename}/{issue_id} 的截圖時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": f"伺服器儲存截圖時發生錯誤: {e}"}), 500


# API 路由：提供截圖檔案訪問
@app.route("/project/<project_name>/screenshots/<path:filename>")
def serve_screenshot(project_name, filename):
    """根據檔名，從專案的截圖資料夾安全地提供圖片檔案"""
    if not is_safe_project_name(project_name):
        abort(404)

    decoded_filename = unquote(filename)
    cleaned_filename = os.path.basename(decoded_filename)
    screenshot_dir = get_project_screenshot_folder(project_name)

    if not os.path.isdir(screenshot_dir):
        app.logger.error(f"找不到截圖目錄: {screenshot_dir}")
        abort(404)

    filename_pattern = r"^\d+-(?:APPSCAN|手動)-[\w\-\._~\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+-[\w\-\._~\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+-[\w\-\._~\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+-\d+\.(?:png|jpg|jpeg|gif|bmp)$"
    if not re.match(filename_pattern, cleaned_filename, re.IGNORECASE | re.UNICODE):
        app.logger.warning(f"無效的截圖檔名格式: '{filename}' -> '{cleaned_filename}'")
        abort(400, "無效的檔名格式。")

    file_path = os.path.abspath(os.path.join(screenshot_dir, cleaned_filename))

    if not file_path.startswith(os.path.abspath(screenshot_dir)):
        app.logger.error(f"拒絕遍歷嘗試: '{file_path}'")
        abort(403)

    if not os.path.isfile(file_path):
        app.logger.warning(f"找不到截圖檔案: '{file_path}'")
        abort(404)

    try:
        mimetype, _ = mimetypes.guess_type(file_path)
        return send_file(file_path, mimetype=mimetype, as_attachment=False)
    except Exception as e:
        app.logger.error(f"提供截圖 '{file_path}' 時發生錯誤: {e}")
        abort(500)


# API 路由：刪除截圖 (移至垃圾桶)
@app.route("/project/<project_name>/api/delete_screenshot", methods=["POST"])
def delete_screenshot(project_name):
    """將指定的截圖檔案移動到專案的截圖垃圾桶資料夾，並從狀態檔移除元數據"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.json
    filename_to_move = data.get("filename")
    issue_id = data.get("issueId")             # <<< 需要 issueId 來定位 meta
    report_filename = data.get("reportFilename") # <<< 需要 reportFilename 來定位 meta

    if not filename_to_move:
        return jsonify({"error": "缺少 'filename' 參數。"}), 400

    # --- NEW: 檢查 ID ---
    if not issue_id or not report_filename:
         app.logger.warning(f"刪除請求缺少 issueId 或 reportFilename (檔案: {filename_to_move})，將無法清理元數據。")
         # 可以選擇返回錯誤或繼續只移動檔案
         # return jsonify({"error": "缺少 issueId 或 reportFilename 參數。"}), 400
    # --- End NEW ---

    app.logger.info(f"API move_screenshot_to_trash: Proj='{project_name}', File='{filename_to_move}', Issue='{issue_id}', Report='{report_filename}'")

    decoded_filename = unquote(filename_to_move)
    cleaned_filename = os.path.basename(decoded_filename)

    filename_pattern = r"^\d+-(?:APPSCAN|手動)-[\w\-\._~\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+-[\w\-\._~\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+-[\w\-\._~\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+-\d+\.(?:png|jpg|jpeg|gif|bmp)$"
    if not re.match(filename_pattern, cleaned_filename, re.IGNORECASE | re.UNICODE):
        app.logger.warning(f"無效的移動目標檔名格式: '{filename_to_move}' -> '{cleaned_filename}'")
        return jsonify({"error": "無效的檔名格式。"}), 400

    source_path = None
    dest_path = None
    try:
        project_data_dir = get_project_data_folder(project_name, ensure_exists=True)
        screenshot_dir = os.path.join(project_data_dir, SCREENSHOTS_SUBFOLDER)
        trash_dir = os.path.join(project_data_dir, TRASH_SCREENSHOTS_SUBFOLDER)
        source_path = os.path.abspath(os.path.join(screenshot_dir, cleaned_filename))

        if not source_path.startswith(os.path.abspath(screenshot_dir)):
            app.logger.error(f"嘗試移動目錄外的檔案: '{source_path}'")
            return jsonify({"error": "拒絕存取。"}), 403

        if os.path.isfile(source_path):
            # --- 移動檔案邏輯 (不變) ---
            dest_filename = cleaned_filename
            dest_path = os.path.abspath(os.path.join(trash_dir, dest_filename))
            counter = 0
            while os.path.exists(dest_path):
                counter += 1
                name, ext = os.path.splitext(cleaned_filename)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                dest_filename = f"{name}_trashed_{timestamp}_{counter}{ext}"
                dest_path = os.path.abspath(os.path.join(trash_dir, dest_filename))
                if counter > 10:
                    app.logger.error(f"垃圾桶中檔案 {cleaned_filename} 的衝突次數過多。")
                    raise IOError("無法在垃圾桶中找到唯一的檔名。")

            shutil.move(source_path, dest_path)
            app.logger.info(f"已將 '{cleaned_filename}' 移動到垃圾桶: '{dest_path}'")

            # --- NEW: 從狀態檔移除截圖元數據 ---
            if issue_id and report_filename: # 確保有 ID 可以查找
                try:
                    statuses = load_statuses(project_name)
                    report_entry = statuses.get(report_filename)
                    issue_entry = report_entry.get(issue_id) if report_entry else None

                    if issue_entry and isinstance(issue_entry.get("screenshots_meta"), dict):
                        # 使用 pop 安全地移除 key，如果 key 不存在也不會報錯
                        removed_meta = issue_entry["screenshots_meta"].pop(cleaned_filename, None)
                        if removed_meta is not None: # 檢查是否真的移除了東西
                             app.logger.info(f"已移除已刪除截圖 '{cleaned_filename}' 的元數據")
                             save_statuses(project_name, statuses) # 儲存變更
                        else:
                             app.logger.warning(f"嘗試刪除元數據，但找不到鍵 '{cleaned_filename}' (Issue: {issue_id}, Report: {report_filename})")
                    elif issue_entry:
                         app.logger.warning(f"找不到或元數據格式錯誤 (screenshots_meta)，無法為 '{cleaned_filename}' 清理元數據")

                except Exception as meta_err:
                    app.logger.error(f"移除截圖元數據時發生錯誤 ('{cleaned_filename}'): {meta_err}")
            else:
                 app.logger.warning(f"缺少 issueId 或 reportFilename，無法為 '{cleaned_filename}' 清理元數據。")
            # --- End NEW ---

            return jsonify({"message": "截圖已移至垃圾桶。"}), 200
        else:
            app.logger.warning(f"找不到要移動的目標檔案: '{source_path}'")
            return jsonify({"error": "找不到截圖檔案。"}), 404

    except OSError as e:
        app.logger.error(f"將 '{source_path}' 移動到 '{dest_path}' 時發生 OS 錯誤: {e}")
        return jsonify({"error": f"無法移動檔案: {e}"}), 500
    except Exception as e:
        app.logger.error(f"將 '{source_path}' 移動到垃圾桶時發生未預期錯誤: {e}", exc_info=True)
        return jsonify({"error": f"伺服器錯誤: {e}"}), 500


# --- NEW: Update Screenshot Caption API ---
@app.route("/project/<project_name>/api/update_screenshot_caption", methods=["POST"])
def update_screenshot_caption(project_name):
    """更新特定截圖的備註。"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.get_json()
    report_filename = data.get("reportFilename")
    issue_id = data.get("issueId")
    screenshot_filename = data.get("screenshotFilename")
    new_caption = data.get("newCaption", "") # 預設為空字串

    if not all([report_filename, issue_id, screenshot_filename]):
        return jsonify({"error": "缺少必要參數 (reportFilename, issueId, screenshotFilename)"}), 400

    if not isinstance(new_caption, str):
        new_caption = str(new_caption) # 確保備註是字串

    app.logger.info(f"API update_caption: Proj='{project_name}', Report='{report_filename}', Issue='{issue_id}', File='{screenshot_filename}'")

    try:
        statuses = load_statuses(project_name)
        report_entry = statuses.get(report_filename)
        if not report_entry or not isinstance(report_entry, dict):
            return jsonify({"error": "找不到指定的報告狀態"}), 404

        issue_entry = report_entry.get(issue_id)
        if not issue_entry or not isinstance(issue_entry, dict):
            return jsonify({"error": "找不到指定的問題狀態"}), 404

        # 確保 screenshots_meta 存在且是字典
        screenshots_meta = issue_entry.setdefault("screenshots_meta", {})
        if not isinstance(screenshots_meta, dict):
             app.logger.warning(f"重設問題 {issue_id} 的無效 screenshots_meta (更新備註時)")
             screenshots_meta = {}
             issue_entry["screenshots_meta"] = screenshots_meta

        # 更新或建立特定截圖檔案的元數據字典
        screenshot_meta = screenshots_meta.setdefault(screenshot_filename, {})
        screenshot_meta["caption"] = new_caption # 設定新的備註

        save_statuses(project_name, statuses) # 儲存變更
        app.logger.info(f"已更新截圖 '{screenshot_filename}' 的備註。")
        return jsonify({"message": "圖片備註已更新。"}), 200

    except Exception as e:
        app.logger.error(f"更新截圖備註時出錯: {e}", exc_info=True)
        return jsonify({"error": "伺服器更新備註時發生錯誤。"}), 500
# --- End NEW API ---


# API 路由：開啟本地的 .scan 檔案
@app.route("/project/<project_name>/api/open_scan_file", methods=["POST"])
def open_scan_file(project_name):
    """根據 XML 檔名，嘗試使用作業系統預設程式開啟對應的 .scan 檔案"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    xml_filename = request.json.get("reportFilename")
    if not xml_filename:
        return jsonify({"error": "缺少 'reportFilename'。"}), 400

    if xml_filename.endswith("-找不到掃描檔"):
        return jsonify({"error": "無法為遺失的報告開啟 .scan 檔案。"}), 400

    secure_xml_fn = secure_filename(os.path.basename(xml_filename))
    original_basename = os.path.basename(xml_filename)

    if not original_basename.lower().endswith(".xml"):
        app.logger.warning(f"無效的 XML 檔名: '{xml_filename}'")
        return jsonify({"error": "無效的 XML 報告檔名格式。"}), 400

    if secure_xml_fn != original_basename:
        app.logger.warning(f"XML 檔名包含潛在問題字元: 原始='{original_basename}', 清理後='{secure_xml_fn}'")

    base_name, _ = os.path.splitext(original_basename)
    scan_filename = base_name + ".scan"
    report_dir = get_project_report_folder(project_name)
    scan_filepath = os.path.abspath(os.path.join(report_dir, scan_filename))

    if not scan_filepath.startswith(os.path.abspath(report_dir)):
        app.logger.error(f"嘗試開啟目錄外的檔案: '{scan_filepath}'")
        return jsonify({"error": "拒絕存取。"}), 403

    if not os.path.isfile(scan_filepath):
        app.logger.warning(f"找不到 .scan 檔案: '{scan_filepath}'")
        return jsonify({"error": f"找不到對應的 .scan 檔案: {scan_filename}"}), 404

    try:
        app.logger.info(f"嘗試開啟 .scan 檔案: '{scan_filepath}'")
        system = platform.system()
        if system == "Windows":
            os.startfile(scan_filepath)
        elif system == "Darwin":
            subprocess.run(["open", scan_filepath], check=True)
        else:
            subprocess.run(["xdg-open", scan_filepath], check=True)
        return jsonify({"message": f"已嘗試開啟 '{scan_filename}'。"}), 200
    except FileNotFoundError as e:
        app.logger.error(f"在 '{system}' 上找不到命令: {e}")
        return jsonify({"error": f"找不到命令 ('{system}' 作業系統)。"}), 500
    except subprocess.CalledProcessError as e:
        app.logger.error(f"執行命令時發生錯誤: {e}")
        return jsonify({"error": f"開啟檔案時發生錯誤: {e}"}), 500
    except Exception as e:
        app.logger.error(f"開啟檔案時發生未預期錯誤: {e}")
        return jsonify({"error": f"未預期錯誤: {e}"}), 500


# 路由：全域設定頁面 (目前只有排除規則)
@app.route("/global_settings", methods=["GET", "POST"])
def global_settings_page():
    """顯示和處理全域設定，目前主要是排除規則的新增和刪除"""
    global exclusion_rules

    if request.method == "POST":
        action = request.form.get("action")

        if action == "add_rule":
            match_type = request.form.get("match_type", "entity_starts_with")
            issue_type = request.form.get("issue_type", "").strip()
            entity_pattern = request.form.get("entity_pattern", "").strip()
            new_rule = {}
            valid = False
            error_message = None

            if not issue_type:
                error_message = "新增規則失敗：「弱點類型」為必填欄位。"
            elif match_type == "issue_type_only":
                new_rule = {"match_type": match_type, "issue_type": issue_type}
                valid = True
            elif match_type in ["entity_starts_with", "entity_contains"]:
                if not entity_pattern:
                    error_message = f"新增規則失敗：選擇「{match_type}」模式時，「實體 Pattern」為必填欄位。"
                else:
                    new_rule = {"match_type": match_type, "issue_type": issue_type, "entity_pattern": entity_pattern}
                    valid = True
            else:
                error_message = "新增規則失敗：無效的匹配模式。"

            if error_message:
                flash(error_message, "danger")
            elif valid:
                is_duplicate = any(rule == new_rule for rule in exclusion_rules)
                if not is_duplicate:
                    exclusion_rules.append(new_rule)
                    save_rules()
                    flash(f"通用排除規則已成功新增！ ({new_rule})", "success")
                    app.logger.info(f"已新增全域規則: {new_rule}")
                else:
                    flash("此通用規則已存在。", "warning")

        elif action == "delete_rule":
            rule_index_str = request.form.get("rule_index")
            if rule_index_str is not None:
                try:
                    rule_index = int(rule_index_str)
                    if 0 <= rule_index < len(exclusion_rules):
                        deleted_rule = exclusion_rules.pop(rule_index)
                        save_rules()
                        flash(f"通用排除規則已成功刪除： {deleted_rule}", "success")
                        app.logger.info(f"已刪除全域規則索引 {rule_index}: {deleted_rule}")
                    else:
                        flash("刪除失敗：無效索引。", "danger")
                        app.logger.warning(f"嘗試刪除無效的規則索引: {rule_index_str}")
                except ValueError:
                    flash("刪除失敗：索引格式錯誤。", "danger")
                    app.logger.warning(f"非整數的規則索引: {rule_index_str}")
                except Exception as e:
                    flash(f"刪除規則時出錯：{e}", "danger")
                    app.logger.error(f"刪除規則索引 {rule_index_str} 時發生錯誤: {e}")
            else:
                flash("刪除失敗：缺少索引。", "danger")
                app.logger.warning("刪除請求中缺少 rule_index。")

        return redirect(url_for("global_settings_page"))

    return render_template("settings.html", rules=list(exclusion_rules), is_global=True)


# 路由：匯出已確認弱點為 Excel
@app.route("/project/<project_name>/export/confirmed_vulnerabilities")
def export_confirmed_vulnerabilities(project_name):
    """產生並提供一個 Excel 檔案，包含所選專案中所有標記為「已確認弱點」的問題"""
    if not is_safe_project_name(project_name):
        abort(400, "無效的專案名稱。")

    try:
        config = load_project_config(project_name)
        display_name = config.get("project_display_name", project_name)
    except Exception as e:
        app.logger.error(f"匯出時載入專案設定錯誤: {e}")
        display_name = project_name

    merge_duplicates = request.args.get("merge_duplicates", "false").lower() == "true"
    include_notes = request.args.get("include_notes", "false").lower() == "true"
    app.logger.info(f"開始匯出已確認弱點 (合併={merge_duplicates}, 筆記={include_notes}) for '{project_name}'")

    target_details = read_target_details_from_excel(project_name)
    if not target_details:
        error_message = f"無法匯出 '{display_name}': 缺少 '{TARGET_LIST_FILENAME}'。"
        flash(error_message, "danger")
        return redirect(url_for("project_index", project_name=project_name))

    expected_numbers = sorted(list(target_details.keys()))

    try:
        project_statuses = load_statuses(project_name)
    except Exception as e:
        app.logger.error(f"匯出時載入狀態檔錯誤: {e}")
        flash("讀取狀態檔時發生錯誤。", "danger")
        return redirect(url_for("project_index", project_name=project_name))

    report_folder = get_project_report_folder(project_name)

    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        safe_sheet_title = re.sub(r"[\\/*?:\[\]]", "_", f"{display_name} 已確認弱點")[:31]
        worksheet.title = safe_sheet_title
        headers = [ "編號", "掃描狀態", "檔案名稱", "標的名稱", "!!!標的!!!", "!!!弱點1!!!", "!!!弱點2!!!", "!!!弱點3!!!", "!!!弱點4!!!", "!!!弱點5!!!", "!!!弱點6!!!", "!!!弱點7!!!" ]
        max_vuln_columns = 7
        vulnerability_start_index = 5

        worksheet.append([None] * len(headers))
        worksheet.append(headers)

        found_files_map = {}
        if os.path.isdir(report_folder):
            try:
                xml_files = [f for f in os.listdir(report_folder) if f.lower().endswith(".xml") and os.path.isfile(os.path.join(report_folder, f))]
                for fn in xml_files:
                    match = re.match(r"(\d+)-.*\.xml", fn, re.IGNORECASE)
                    if match:
                        try:
                            found_files_map[int(match.group(1))] = fn
                        except ValueError:
                            pass
            except Exception as e:
                app.logger.error(f"匯出時列出 XML 檔案錯誤: {e}")
        else:
            app.logger.warning(f"匯出時找不到報告資料夾: {report_folder}")

        manual_override_statuses = { STATUS_OPTIONS["誤判"], STATUS_OPTIONS["已確認弱點"], STATUS_OPTIONS["人工審查中"] }

        for report_num in sorted(expected_numbers):
            row_data = [report_num] + [""] * (len(headers) - 1)
            report_filename = found_files_map.get(report_num)
            file_path = os.path.join(report_folder, report_filename) if report_filename else None
            scan_status_display = FILE_NOT_FOUND_STATUS
            filename_display = f"{report_num}-找不到掃描檔"
            all_confirmed_issues_for_report = []
            report_file_statuses = project_statuses.get(report_filename, {})
            target_info = target_details.get(report_num, {"name": "N/A (不在目標列表)", "url": "N/A (不在目標列表)"})
            target_name = target_info["name"]
            target_url = target_info["url"]

            if report_filename and file_path and os.path.isfile(file_path):
                filename_display = report_filename
                try:
                    scan_status_display = get_scan_status(file_path)
                except Exception:
                    scan_status_display = READ_ERROR_STATUS

                if scan_status_display not in [FILE_NOT_FOUND_STATUS, PARSE_ERROR_STATUS, READ_ERROR_STATUS, INCOMPLETE_STATUS]:
                    try:
                        parsed_data = parse_appscan_xml(project_name, file_path, report_filename)
                        if parsed_data and parsed_data.get("issues"):
                            for issue in parsed_data["issues"]:
                                issue_id = issue.get("id")
                                if not issue_id or issue_id.startswith("error_item_"):
                                    continue

                                status_info = report_file_statuses.get(issue_id, {})
                                saved_status = status_info.get("status", DEFAULT_STATUS) if isinstance(status_info, dict) else str(status_info or DEFAULT_STATUS)
                                display_status = saved_status

                                if saved_status not in manual_override_statuses:
                                    issue_type = issue.get("issue_type")
                                    entity_name = issue.get("entity_name", "")
                                    matches_a_rule = False
                                    for rule in exclusion_rules:
                                        rule_match_type = rule.get("match_type", "entity_starts_with")
                                        rule_issue_type = rule.get("issue_type")
                                        if not rule_issue_type or issue_type != rule_issue_type:
                                            continue

                                        rule_applied = False
                                        if rule_match_type == "issue_type_only":
                                            rule_applied = True
                                        elif rule_match_type == "entity_starts_with":
                                            pattern = rule.get("entity_pattern")
                                            if pattern and entity_name is not None and entity_name.startswith(pattern):
                                                rule_applied = True
                                        elif rule_match_type == "entity_contains":
                                            pattern = rule.get("entity_pattern")
                                            if pattern and entity_name is not None and pattern in entity_name:
                                                rule_applied = True

                                        if rule_applied:
                                            matches_a_rule = True
                                            break

                                    if matches_a_rule:
                                        display_status = AUTO_EXCLUDED_STATUS
                                    elif saved_status == AUTO_EXCLUDED_STATUS:
                                        display_status = DEFAULT_STATUS

                                if display_status == STATUS_OPTIONS["已確認弱點"]:
                                    issue["status"] = display_status
                                    issue["screenshot_taken"] = status_info.get("screenshot_taken", False) if isinstance(status_info, dict) else False
                                    issue["note"] = status_info.get("note", DEFAULT_NOTE) if isinstance(status_info, dict) else DEFAULT_NOTE
                                    issue["source"] = "appscan"
                                    if "severity_key" not in issue:
                                        issue["severity_key"] = "unknown"
                                        app.logger.warning(f"Issue {issue_id} in {report_filename} missing severity_key, defaulting to 'unknown'.")
                                    if "severity_display" not in issue:
                                        issue["severity_display"] = SEVERITY_DISPLAY_MAP.get(issue.get("severity_key", "unknown"), issue.get("severity_key", "unknown").capitalize())
                                    all_confirmed_issues_for_report.append(issue)
                    except Exception as e:
                        app.logger.error(f"匯出時處理 AppScan 問題錯誤 {file_path}: {e}", exc_info=True)
                        scan_status_display = "弱點處理錯誤"

            for issue_id, status_data in report_file_statuses.items():
                if issue_id.startswith("_manual_") and isinstance(status_data, dict):
                    manual_details = status_data.get("manual_details")
                    current_status = status_data.get("status", DEFAULT_STATUS)
                    if manual_details and current_status == STATUS_OPTIONS["已確認弱點"]:
                        severity_key = manual_details.get("severity_key", "medium")
                        severity_display = manual_details.get("severity_display", SEVERITY_DISPLAY_MAP.get(severity_key, "未知"))
                        manual_issue = {
                            "id": issue_id,
                            "status": current_status,
                            "screenshot_taken": status_data.get("screenshot_taken", False),
                            "note": status_data.get("note", DEFAULT_NOTE),
                            "source": status_data.get("source", "manual"),
                            **manual_details,
                            "severity_key": severity_key,
                            "severity_display": severity_display,
                            "cvss_score": "N/A",
                            "cve_name": None,
                            "cve_url": None,
                            "http_traffic": "N/A"
                        }
                        all_confirmed_issues_for_report.append(manual_issue)

            confirmed_vulnerabilities_output = []
            if all_confirmed_issues_for_report:
                 try:
                     if merge_duplicates:
                         merged_data = {}
                         for issue in all_confirmed_issues_for_report:
                             issue_type = issue.get("issue_type", "N/A")
                             entity_name = issue.get("entity_name", "N/A")
                             effective_merge_key = issue_type
                             if issue_type == "有弱點的元件":
                                 cleaned_entity_name = entity_name
                                 if entity_name and entity_name != 'N/A':
                                     cleaned_entity_name = re.sub(r'\s*\([^)]*\)\s*$', '', entity_name).strip()
                                 if cleaned_entity_name and cleaned_entity_name != 'N/A':
                                     effective_merge_key = f"{issue_type} {cleaned_entity_name}"

                             if effective_merge_key not in merged_data:
                                 merged_data[effective_merge_key] = {
                                     "representative_url": issue.get("url", "N/A"),
                                     "any_screenshot_taken": False,
                                     "combined_notes": [],
                                     "entities_details": [],
                                     "sources": set()
                                 }

                             merged_data[effective_merge_key]['sources'].add(issue.get("source", "unknown").lower())
                             merged_data[effective_merge_key]["entities_details"].append({
                                 "name": entity_name,
                                 "type": issue.get("entity_type", "N/A"),
                                 "severity_display": issue.get("severity_display", "未知"),
                                 "severity_key": issue.get("severity_key", "unknown")
                             })
                             note = issue.get("note", "").strip()
                             if note:
                                 merged_data[effective_merge_key]["combined_notes"].append(note)
                             merged_data[effective_merge_key]["any_screenshot_taken"] |= issue.get("screenshot_taken", False)

                         for effective_key, data in merged_data.items():
                             screenshot_text = "已截圖" if data["any_screenshot_taken"] else "無"
                             source_tag_text = "混合" if len(data['sources']) > 1 else ("手動" if "manual" in data['sources'] else "AppScan")
                             highest_overall_severity_level = -1
                             highest_overall_severity_display = "未知"
                             for entity_detail in data["entities_details"]:
                                 level = SEVERITY_LEVELS.get(entity_detail["severity_key"], -1)
                                 if level > highest_overall_severity_level:
                                     highest_overall_severity_level = level
                                     highest_overall_severity_display = entity_detail["severity_display"]

                             source_tag_with_severity = f"({source_tag_text}-{highest_overall_severity_display})"
                             first_line = effective_key
                             formatted_string = f"{first_line}\n{source_tag_with_severity}({screenshot_text})\n{data['representative_url']}\n\n"
                             highest_severity_for_entity = {}

                             for entity_detail in data["entities_details"]:
                                 name = entity_detail["name"]
                                 etype = entity_detail["type"]
                                 if name == "N/A" and etype == "N/A":
                                     continue

                                 entity_key = (name, etype)
                                 severity_key = entity_detail["severity_key"]
                                 severity_display = entity_detail["severity_display"]
                                 current_level = SEVERITY_LEVELS.get(severity_key, -1)
                                 stored_data = highest_severity_for_entity.get(entity_key)
                                 stored_level = stored_data[0] if stored_data else -1

                                 if current_level > stored_level or stored_data is None:
                                     highest_severity_for_entity[entity_key] = (current_level, severity_display, name, etype)

                             sorted_processed_entities = sorted(highest_severity_for_entity.values(), key=lambda x: (x[2], x[3]))
                             entity_lines = [ f"{name} ({etype}) ({sev_display})" for _level, sev_display, name, etype in sorted_processed_entities ]
                             formatted_string += "\n".join(entity_lines)

                             if include_notes and data["combined_notes"]:
                                 unique_notes = sorted(list(set(data["combined_notes"])))
                                 formatted_string += f"\n\n--- 筆記 ---\n" + "\n---\n".join(unique_notes)

                             confirmed_vulnerabilities_output.append(formatted_string)
                     else:
                         highest_severity_issues = {}
                         for issue in all_confirmed_issues_for_report:
                             group_key = (
                                 issue.get("issue_type", "N/A"),
                                 issue.get("url", "N/A"),
                                 issue.get("entity_name", "N/A"),
                                 issue.get("entity_type", "N/A"),
                                 issue.get("source", "unknown").lower()
                             )
                             current_severity_level = SEVERITY_LEVELS.get(issue.get("severity_key"), -1)

                             if group_key not in highest_severity_issues:
                                 highest_severity_issues[group_key] = issue
                             else:
                                 stored_issue = highest_severity_issues[group_key]
                                 stored_severity_level = SEVERITY_LEVELS.get(stored_issue.get("severity_key"), -1)
                                 if current_severity_level > stored_severity_level:
                                     highest_severity_issues[group_key] = issue

                         confirmed_vulnerabilities_output = []
                         for final_issue in highest_severity_issues.values():
                             screenshot_text = "已截圖" if final_issue.get("screenshot_taken", False) else "無"
                             note = final_issue.get("note", "").strip()
                             issue_type = final_issue.get('issue_type', 'N/A')
                             entity_name = final_issue.get('entity_name', 'N/A')
                             entity_type = final_issue.get('entity_type', 'N/A')
                             url = final_issue.get('url', 'N/A')
                             severity_display = final_issue.get("severity_display", "未知")
                             source_raw = final_issue.get("source", "unknown").lower()
                             source_base_text = "手動" if source_raw == "manual" else "AppScan"
                             source_tag_with_severity = f"({source_base_text}-{severity_display})"

                             if issue_type == "有弱點的元件":
                                 cleaned_entity_name = entity_name
                                 if entity_name and entity_name != 'N/A':
                                     cleaned_entity_name = re.sub(r'\s*\([^)]*\)\s*$', '', entity_name).strip()
                                 first_line = issue_type
                                 if cleaned_entity_name and cleaned_entity_name != 'N/A':
                                     first_line += f" {cleaned_entity_name}"
                                 formatted_string = f"{first_line}\n{source_tag_with_severity}({screenshot_text})\n{url}\n類型: {entity_type}"
                             else:
                                 formatted_string = f"{issue_type}\n{source_tag_with_severity}({screenshot_text})\n{url}\n\n{entity_name} ({entity_type})"

                             if include_notes and note:
                                 formatted_string += f"\n\n--- 筆記 ---\n{note}"

                             confirmed_vulnerabilities_output.append(formatted_string)
                 except Exception as format_err:
                     scan_status_display = "弱點格式化錯誤"
                     app.logger.error(f"格式化報告 {report_filename} 的已確認弱點時發生錯誤: {format_err}", exc_info=True)
                     confirmed_vulnerabilities_output = ["格式化錯誤"]

            row_data[1] = scan_status_display
            row_data[2] = filename_display
            row_data[3] = target_name
            row_data[4] = target_url

            for i, vuln_text in enumerate(confirmed_vulnerabilities_output[:max_vuln_columns]):
                try:
                    cell_text = str(vuln_text) if vuln_text is not None else ""
                    if len(cell_text) > 32767:
                        app.logger.warning(f"單元格文本過長 (報告 {report_num}, 弱點 {i+1})，已截斷。")
                        cell_text = cell_text[:32767]
                    row_data[i + vulnerability_start_index] = cell_text
                except Exception as cell_write_err:
                    app.logger.error(f"寫入弱點文本到 row_data 時出錯: {cell_write_err}", exc_info=True)
                    row_data[i + vulnerability_start_index] = "寫入錯誤"

            try:
                final_row_data = [str(item) if item is not None else "" for item in row_data]
                worksheet.append(final_row_data)
            except Exception as append_err:
                app.logger.error(f"將行數據附加到工作表時出錯 (報告 {report_num}): {append_err}", exc_info=True)

        for col_idx, column_cells in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            header_name = headers[col_idx] if col_idx < len(headers) else ""
            is_vulnerability_column = header_name.startswith("!!!弱點")

            for cell in column_cells[1:]:
                try:
                    cell_value_str = str(cell.value or "")
                    line_max_len = 0
                    visual_length = 0
                    if cell_value_str:
                        lines = cell_value_str.split('\n')
                        line_max_len = max(len(line) for line in lines) if lines else 0
                        cjk_chars = len(re.findall(r'[\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]', cell_value_str))
                        non_cjk_in_max_line = 0
                        cjk_in_max_line = 0
                        if lines:
                            max_line_str = max(lines, key=len)
                            non_cjk_in_max_line = len(max_line_str) - len(re.findall(r'[\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]', max_line_str))
                            cjk_in_max_line = len(max_line_str) - non_cjk_in_max_line
                        visual_length = non_cjk_in_max_line + (cjk_in_max_line * 1.9)
                        max_length = max(max_length, visual_length)
                except Exception as e:
                    app.logger.warning(f"計算儲存格 {cell.coordinate} 的寬度時發生錯誤: {e}")

                try:
                    needs_wrap = "\n" in str(cell.value or "") or is_vulnerability_column or header_name == "標的名稱" or header_name == "!!!標的!!!"
                    cell.alignment = Alignment(wrap_text=needs_wrap, vertical="top", horizontal="left")
                except Exception as e:
                    app.logger.warning(f"設定儲存格 {cell.coordinate} 的對齊方式時發生錯誤: {e}")

            if header_name == "編號":
                adjusted_width = max(max_length, 8)
            elif header_name == "掃描狀態":
                adjusted_width = max(max_length, 15)
            elif header_name == "檔案名稱":
                adjusted_width = min(max(max_length + 2, 40), 70)
            elif header_name == "標的名稱":
                adjusted_width = min(max(max_length + 2, 25), 60)
            elif header_name == "!!!標的!!!":
                adjusted_width = min(max(max_length + 2, 30), 70)
            elif is_vulnerability_column:
                adjusted_width = min(max(max_length + 2, 35), 80)
            else:
                adjusted_width = max_length + 2

            worksheet.column_dimensions[column_letter].width = adjusted_width + 1

        excel_stream = io.BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)

        safe_display_name = secure_filename(display_name) or "project_export"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        merge_suffix = "_merged" if merge_duplicates else ""
        notes_suffix = "_with_notes" if include_notes else ""
        download_filename = f"{safe_display_name}_{timestamp}_confirmed{merge_suffix}{notes_suffix}.xlsx"

        app.logger.info(f"已確認弱點 Excel 產生完成。正在發送: {download_filename}")

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=download_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"匯出已確認弱點時發生嚴重錯誤: {traceback.format_exc()}")
        flash(f"匯出已確認弱點時出錯: {e}", "danger")
        return redirect(url_for("project_index", project_name=project_name))


# 路由：匯出所有問題的筆記為 Excel
@app.route("/project/<project_name>/export/all_notes")
def export_all_notes(project_name):
    """產生並提供一個 Excel 檔案，包含所選專案中所有報告的所有問題及其筆記"""
    if not is_safe_project_name(project_name):
        abort(400, "無效的專案名稱。")

    try:
        config = load_project_config(project_name)
        display_name = config.get("project_display_name", project_name)
    except Exception as e:
        app.logger.error(f"匯出時載入專案設定錯誤: {e}")
        display_name = project_name

    app.logger.info(f"開始匯出專案 '{project_name}' 的所有筆記")

    try:
        project_statuses = load_statuses(project_name)
    except Exception as e:
        app.logger.error(f"匯出所有筆記時載入狀態檔錯誤: {e}")
        flash("讀取狀態檔時發生錯誤。", "danger")
        return jsonify({"error": "伺服器載入狀態數據時發生錯誤。"}), 500

    report_folder = get_project_report_folder(project_name)

    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        safe_sheet_title = re.sub(r'[\\/*?:\[\]]', '_', f"{display_name} 所有筆記")[:31]
        worksheet.title = safe_sheet_title
        headers = [ "報告檔案", "問題 ID", "來源", "嚴重性", "狀態", "截圖完成", "弱點類型", "URL", "實體名稱", "實體類型", "筆記" ]
        worksheet.append(headers)

        processed_files = 0
        total_issues_exported = 0
        xml_files = []

        if os.path.isdir(report_folder):
            xml_files = sorted([f for f in os.listdir(report_folder) if f.lower().endswith(".xml") and os.path.isfile(os.path.join(report_folder, f))])
        else:
            app.logger.warning(f"匯出所有筆記時找不到報告資料夾 '{report_folder}'。")

        all_report_keys = set(xml_files) | set(project_statuses.keys())
        all_report_files = {k for k in all_report_keys if k != REPORT_COMPLETED_KEY}
        manual_override_statuses = { STATUS_OPTIONS["誤判"], STATUS_OPTIONS["已確認弱點"], STATUS_OPTIONS["人工審查中"] }

        for report_filename in sorted(list(all_report_files)):
            file_path = os.path.join(report_folder, report_filename)
            app.logger.debug(f"正在為所有筆記匯出處理報告: {report_filename}")
            processed_files += 1
            all_issues_for_report = []
            report_file_statuses = project_statuses.get(report_filename, {})
            parsed_data = None
            is_placeholder = report_filename.endswith("-找不到掃描檔")
            xml_file_exists = os.path.isfile(file_path)

            if not is_placeholder and xml_file_exists:
                try:
                    parsed_data = parse_appscan_xml(project_name, file_path, report_filename)
                except Exception as parse_err:
                    app.logger.error(f"解析 {report_filename} 時發生錯誤: {parse_err}")
            elif not xml_file_exists and report_filename in xml_files:
                app.logger.warning(f"XML 檔案 {report_filename} 在列表中但找不到。")

            if parsed_data and parsed_data.get("issues"):
                for issue in parsed_data["issues"]:
                    issue_id = issue.get("id", "")
                    if issue.get("source") == "error":
                        issue["export_status"] = issue.get("status", "處理錯誤")
                        issue["export_screenshot_taken"] = issue.get("screenshot_taken", False)
                        issue["export_note"] = issue.get("note", "XML解析錯誤")
                        all_issues_for_report.append(issue)
                        continue

                    if not issue_id:
                        continue

                    status_info = report_file_statuses.get(issue_id, {})
                    saved_status = status_info.get("status", DEFAULT_STATUS) if isinstance(status_info, dict) else str(status_info or DEFAULT_STATUS)
                    saved_screenshot_taken = status_info.get("screenshot_taken", False) if isinstance(status_info, dict) else False
                    saved_note = status_info.get("note", DEFAULT_NOTE) if isinstance(status_info, dict) else DEFAULT_NOTE
                    display_status = saved_status
                    display_screenshot_taken = saved_screenshot_taken

                    if saved_status not in manual_override_statuses:
                        issue_type = issue.get("issue_type")
                        entity_name = issue.get("entity_name", "")
                        matches_a_rule = False
                        for rule in exclusion_rules:
                            rule_match_type = rule.get("match_type", "entity_starts_with")
                            rule_issue_type = rule.get("issue_type")
                            if not rule_issue_type or issue_type != rule_issue_type:
                                continue

                            rule_applied = False
                            if rule_match_type == "issue_type_only":
                                rule_applied = True
                            elif rule_match_type == "entity_starts_with":
                                pattern = rule.get("entity_pattern")
                                if pattern and entity_name is not None and entity_name.startswith(pattern):
                                    rule_applied = True
                            elif rule_match_type == "entity_contains":
                                pattern = rule.get("entity_pattern")
                                if pattern and entity_name is not None and pattern in entity_name:
                                    rule_applied = True

                            if rule_applied:
                                matches_a_rule = True
                                break

                        if matches_a_rule:
                            display_status = AUTO_EXCLUDED_STATUS
                            display_screenshot_taken = True
                        elif saved_status == AUTO_EXCLUDED_STATUS:
                            display_status = DEFAULT_STATUS
                            display_screenshot_taken = False

                    issue["export_status"] = display_status
                    issue["export_screenshot_taken"] = display_screenshot_taken
                    issue["export_note"] = saved_note
                    issue["source"] = "AppScan"
                    all_issues_for_report.append(issue)

            for issue_id, status_data in report_file_statuses.items():
                 if issue_id == REPORT_COMPLETED_KEY:
                     continue

                 if issue_id.startswith("_manual_") and isinstance(status_data, dict):
                    manual_details = status_data.get("manual_details")
                    if manual_details and isinstance(manual_details, dict):
                        manual_issue = {
                            "id": issue_id,
                            "export_status": status_data.get("status", DEFAULT_STATUS),
                            "export_screenshot_taken": status_data.get("screenshot_taken", False),
                            "export_note": status_data.get("note", DEFAULT_NOTE),
                            "source": status_data.get("source", "manual"),
                            **manual_details,
                            "severity_display": manual_details.get("severity_display", "未知"),
                            "issue_type": manual_details.get("issue_type", "N/A"),
                            "url": manual_details.get("url", "N/A"),
                            "entity_name": manual_details.get("entity_name", "N/A"),
                            "entity_type": manual_details.get("entity_type", "N/A")
                        }
                        all_issues_for_report.append(manual_issue)

            if not all_issues_for_report:
                 status_to_show, note_to_show, source_to_show, issue_id_to_show = "", "", "-", "INFO"
                 if not is_placeholder and not xml_file_exists and report_filename in xml_files:
                     status_to_show, note_to_show = FILE_NOT_FOUND_STATUS, "原始報告檔遺失"
                 elif parsed_data is None and not is_placeholder and xml_file_exists:
                     status_to_show, note_to_show, issue_id_to_show = PARSE_ERROR_STATUS, "XML 無法解析", "PARSE_ERROR"
                 elif is_placeholder:
                     status_to_show, note_to_show, issue_id_to_show = FILE_NOT_FOUND_STATUS, "來自標的列表，無報告檔", "PLACEHOLDER"

                 if status_to_show:
                     worksheet.append([ report_filename, issue_id_to_show, source_to_show, "-", status_to_show, "-", "-", "-", "-", "-", note_to_show ])
                     total_issues_exported += 1

            for issue in all_issues_for_report:
                screenshot_text = "是" if issue.get("export_screenshot_taken", False) else "否"
                note_text = issue.get("export_note", DEFAULT_NOTE)
                source = issue.get("source", "unknown").lower()
                source_text = "手動" if source == "manual" else ("錯誤" if source == "error" else "AppScan")
                row_data = [
                    report_filename,
                    issue.get("id", "N/A"),
                    source_text,
                    issue.get("severity_display", "未知"),
                    issue.get("export_status", DEFAULT_STATUS),
                    screenshot_text,
                    issue.get("issue_type", "N/A"),
                    issue.get("url", "N/A"),
                    issue.get("entity_name", "N/A"),
                    issue.get("entity_type", "N/A"),
                    note_text
                ]
                worksheet.append(row_data)
                total_issues_exported += 1

        for col_idx, column_cells in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            note_col_idx = headers.index("筆記") + 1
            url_col_idx = headers.index("URL") + 1
            is_note_column = col_idx + 1 == note_col_idx
            is_url_column = col_idx + 1 == url_col_idx

            for cell in column_cells:
                try:
                    cell_value_str = str(cell.value or "")
                    current_max = 0
                    if cell_value_str:
                        current_max = max(len(line) for line in cell_value_str.split('\n'))
                        max_length = max(max_length, current_max)
                    needs_wrap = is_note_column or is_url_column or "\n" in cell_value_str
                    cell.alignment = Alignment(wrap_text=needs_wrap, vertical="top", horizontal="left")
                except Exception as e:
                    app.logger.warning(f"格式化儲存格 {cell.coordinate} (所有筆記) 時發生錯誤: {e}")

            long_text_limit, medium_text_limit, default_limit = 70, 50, 30
            header_name = headers[col_idx] if col_idx < len(headers) else ""

            if header_name == "筆記":
                adjusted_width = long_text_limit
            elif header_name in ["URL", "弱點類型", "實體名稱", "報告檔案"]:
                adjusted_width = medium_text_limit
            elif header_name == "問題 ID":
                adjusted_width = 18
            elif header_name == "來源":
                adjusted_width = 10
            elif header_name == "嚴重性":
                adjusted_width = 10
            elif header_name == "狀態":
                adjusted_width = 12
            elif header_name == "截圖完成":
                adjusted_width = 10
            elif header_name == "實體類型":
                adjusted_width = 15
            else:
                adjusted_width = default_limit

            adjusted_width = min(max(adjusted_width, max_length + 2), 100)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        excel_stream = io.BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)

        safe_display_name = secure_filename(display_name) or "project_export"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"{safe_display_name}_{timestamp}_all_notes.xlsx"

        app.logger.info(f"所有筆記匯出已產生 ({processed_files} 個檔案, {total_issues_exported} 個問題)。正在發送: {download_filename}")

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=download_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"匯出所有筆記時發生嚴重錯誤: {traceback.format_exc()}")
        return jsonify({"error": f"匯出所有筆記時錯誤: {e}"}), 500


# 路由：匯出異常報告為 Excel
@app.route("/project/<project_name>/export/abnormal_reports")
def export_abnormal_reports(project_name):
    """產生並提供一個 Excel 檔案，包含所選專案中所有掃描狀態異常或遺失的報告"""
    if not is_safe_project_name(project_name):
        abort(400, "無效的專案名稱。")

    app.logger.info(f"開始匯出專案 '{project_name}' 的異常報告")

    try:
        target_details = read_target_details_from_excel(project_name)
        if not target_details:
            flash(f"無法匯出異常報告：專案 '{project_name}' 缺少 '{TARGET_LIST_FILENAME}'。", "danger")
            return redirect(url_for("project_index", project_name=project_name))
        expected_numbers = sorted(list(target_details.keys()))
        expected_set = set(expected_numbers)
    except Exception as e:
        flash(f"讀取 target.xlsx 錯誤: {e}", "danger")
        app.logger.error(f"讀取目標細節時發生錯誤: {e}")
        return redirect(url_for("project_index", project_name=project_name))

    report_folder = get_project_report_folder(project_name)
    abnormal_reports_data = []
    processed_numbers = set()
    problematic_statuses_list = [
        SCAN_STATUS_MAP.get("Failed", "失敗"), PARSE_ERROR_STATUS, READ_ERROR_STATUS,
        INCOMPLETE_STATUS, DEFAULT_SCAN_STATUS, SCAN_STATUS_MAP.get("Aborted", "已中斷")
    ]

    if os.path.isdir(report_folder):
        try:
            xml_files = [f for f in os.listdir(report_folder) if f.lower().endswith(".xml") and os.path.isfile(os.path.join(report_folder, f))]
            for filename in xml_files:
                filepath = os.path.join(report_folder, filename)
                file_number = None
                match = re.match(r"(\d+)-.*\.xml", filename, re.IGNORECASE)
                if match:
                    try:
                        file_number = int(match.group(1))
                        if file_number in expected_set:
                            processed_numbers.add(file_number)
                        else:
                            app.logger.warning(f"異常報告匯出: 檔案 '{filename}' 編號 {file_number} 不在目標清單中。")
                    except ValueError:
                        pass

                if file_number is None:
                    file_number = "無編號"

                scan_status = get_scan_status(filepath)
                if scan_status in problematic_statuses_list:
                    target_info = {"url": "N/A", "name": "N/A"}
                    if isinstance(file_number, int):
                        target_info = target_details.get(file_number, {"url": "N/A (未於列表)", "name": "N/A (未於列表)"})
                    elif file_number == "無編號":
                        target_info = {"url": "N/A (無編號)", "name": "N/A (無編號)"}
                    abnormal_reports_data.append({
                        "number": file_number,
                        "status": scan_status,
                        "url": target_info["url"],
                        "name": target_info["name"],
                        "filename": filename
                    })
        except Exception as e:
            app.logger.error(f"匯出異常報告時列出/處理 XML 錯誤: {e}")

    missing_numbers = sorted(list(expected_set - processed_numbers))
    for m_num in missing_numbers:
        target_info = target_details.get(m_num, {"url": "N/A", "name": "N/A"})
        abnormal_reports_data.append({
            "number": m_num,
            "status": FILE_NOT_FOUND_STATUS,
            "url": target_info["url"],
            "name": target_info["name"],
            "filename": f"{m_num}-找不到掃描檔"
        })

    abnormal_reports_data.sort(key=lambda item: item["number"] if isinstance(item["number"], int) else float("inf"))

    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        config = load_project_config(project_name)
        display_name = config.get("project_display_name", project_name)
        safe_sheet_title = re.sub(r'[\\/*?:\[\]]', '_', f"{display_name} 異常報告")[:31]
        worksheet.title = safe_sheet_title
        headers = ["編號", "檔案名稱", "URL", "標的名稱", "報告狀態"]
        worksheet.append(headers)

        for report_info in abnormal_reports_data:
            worksheet.append([
                report_info["number"],
                report_info["filename"],
                report_info["url"],
                report_info["name"],
                report_info["status"]
            ])

        for col_idx, column_cells in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            header_name = headers[col_idx] if col_idx < len(headers) else ""
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value or "")))
                except:
                    pass

            adjusted_width = max_length + 2
            if header_name == "編號":
                adjusted_width = max(adjusted_width, 8)
            elif header_name == "檔案名稱":
                adjusted_width = min(max(adjusted_width, 40), 70)
            elif header_name == "URL":
                adjusted_width = min(max(adjusted_width, 30), 70)
            elif header_name == "標的名稱":
                adjusted_width = min(max(adjusted_width, 25), 60)
            elif header_name == "報告狀態":
                adjusted_width = max(adjusted_width, 15)

            worksheet.column_dimensions[column_letter].width = adjusted_width

        excel_stream = io.BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)

        safe_display_name = secure_filename(display_name) or "project_export"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"{safe_display_name}_{timestamp}_abnormal_reports.xlsx"

        app.logger.info(f"異常報告 Excel 已產生 ({len(abnormal_reports_data)} 個條目)。正在發送: {download_filename}")

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=download_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"產生異常報告匯出時發生嚴重錯誤: {traceback.format_exc()}")
        flash(f"匯出異常報告時發生嚴重錯誤: {e}", "danger")
        return redirect(url_for("project_index", project_name=project_name))


# --- JSON Export Route (Merged Logic) ---
@app.route("/project/<project_name>/export/confirmed_json")
def export_confirmed_json(project_name):
    """
    產生並提供 JSON 檔案，其中包含所有 '已確認弱點' 的問題。
    按標的網址分組，然後按弱點名稱合併。
    對於 URL 和實體名稱完全相同的重複項，只保留嚴重性最高的實例。
    包含所有相關截圖及其備註。
    """
    if not is_safe_project_name(project_name):
        abort(400, "無效的專案名稱。")

    try:
        config = load_project_config(project_name)
        display_name = config.get("project_display_name", project_name)
        app.logger.info(f"開始 JSON 匯出 (合併/最高嚴重性/含備註) - 已確認弱點於 '{display_name}'")
    except Exception as e:
        app.logger.error(f"JSON 匯出時載入專案設定錯誤: {e}")
        display_name = project_name # 出錯時使用內部名稱

    # 載入目標詳細資訊
    target_details = read_target_details_from_excel(project_name)
    if not target_details:
        error_message = f"無法匯出 '{display_name}': 缺少必要的 '{TARGET_LIST_FILENAME}' 檔案。"
        flash(error_message, "danger")
        app.logger.error(f"JSON 匯出失敗 for '{project_name}': 缺少目標清單。")
        return redirect(url_for("project_index", project_name=project_name))

    # 建立檔名到目標資訊的映射 (URL 和報告編號)
    filename_to_target_map = {}
    report_folder = get_project_report_folder(project_name)
    if os.path.isdir(report_folder):
        for report_num, details in target_details.items():
            potential_files = glob.glob(os.path.join(report_folder, f"{report_num}-*.xml"))
            if potential_files:
                # 儲存檔名、URL 和報告編號
                filename_to_target_map[os.path.basename(potential_files[0])] = {
                    "url": details.get("url", "N/A"),
                    "number": report_num
                }
            else:
                 # 即使找不到 XML，也為佔位符建立映射，以便處理關聯的手動弱點
                 placeholder_filename = f"{report_num}-找不到掃描檔"
                 filename_to_target_map[placeholder_filename] = {
                     "url": details.get("url", "N/A"),
                     "number": report_num
                 }

    # 載入狀態
    try:
        project_statuses = load_statuses(project_name)
    except Exception as e:
        app.logger.error(f"JSON 匯出時載入狀態檔錯誤: {e}")
        flash("讀取狀態檔時發生錯誤。", "danger")
        return redirect(url_for("project_index", project_name=project_name))

    # --- 資料收集與合併 ---
    # temp_instances[target_url][effective_vuln_name] = list of all raw instance dicts before deduplication
    temp_instances = defaultdict(lambda: defaultdict(list))
    processed_issue_ids = set() # 追蹤已處理的 Issue ID，避免重複處理
    manual_override_statuses = { STATUS_OPTIONS["誤判"], STATUS_OPTIONS["已確認弱點"], STATUS_OPTIONS["人工審查中"] }

    # --- 第一步：收集所有確認的弱點實例的詳細資料 ---
    # 處理 AppScan 問題
    for report_filename, target_info in filename_to_target_map.items():
        # 跳過佔位符檔案，只處理實際存在的 XML
        if report_filename.endswith("-找不到掃描檔"):
            continue
        file_path = os.path.join(report_folder, report_filename)
        if not os.path.isfile(file_path):
            app.logger.warning(f"JSON Export: 找不到檔案 {report_filename}，跳過。")
            continue

        target_url = target_info["url"]
        report_num = target_info["number"]
        report_file_statuses = project_statuses.get(report_filename, {})
        parsed_data = None
        try:
            parsed_data = parse_appscan_xml(project_name, file_path, report_filename)
        except Exception as parse_err:
            app.logger.error(f"JSON Export: 解析 {report_filename} 時發生錯誤: {parse_err}")
            continue # 跳過此報告

        if parsed_data and parsed_data.get("issues"):
            for issue in parsed_data["issues"]:
                issue_id = issue.get("id")
                if not issue_id or issue_id.startswith("error_item_"):
                    continue # 跳過無效或錯誤的 issue

                # 標記此 issue 已從 XML 處理過
                processed_issue_ids.add((report_filename, issue_id))

                # 獲取並計算最終狀態 (含規則)
                status_info = report_file_statuses.get(issue_id, {})
                saved_status = status_info.get("status", DEFAULT_STATUS) if isinstance(status_info, dict) else str(status_info or DEFAULT_STATUS)
                final_status = saved_status
                if saved_status not in manual_override_statuses:
                    issue_type_name = issue.get("issue_type")
                    entity_name_raw = issue.get("entity_name", "")
                    matches_a_rule = False
                    for rule in exclusion_rules:
                        rule_match_type = rule.get("match_type", "entity_starts_with")
                        rule_issue_type = rule.get("issue_type")
                        if not rule_issue_type or issue_type_name != rule_issue_type:
                            continue
                        rule_applied = False
                        if rule_match_type == "issue_type_only":
                            rule_applied = True
                        elif rule_match_type == "entity_starts_with":
                            pattern = rule.get("entity_pattern")
                            if pattern and entity_name_raw is not None and entity_name_raw.startswith(pattern):
                                rule_applied = True
                        elif rule_match_type == "entity_contains":
                            pattern = rule.get("entity_pattern")
                            if pattern and entity_name_raw is not None and pattern in entity_name_raw:
                                rule_applied = True
                        if rule_applied:
                            matches_a_rule = True
                            break
                    if matches_a_rule:
                        final_status = AUTO_EXCLUDED_STATUS
                    elif saved_status == AUTO_EXCLUDED_STATUS:
                        final_status = DEFAULT_STATUS

                # 只處理確認的弱點
                if final_status == STATUS_OPTIONS["已確認弱點"]:
                    # 決定合併用的弱點名稱
                    issue_type = issue.get("issue_type", "N/A")
                    entity_name = issue.get("entity_name", "N/A")
                    effective_vuln_name = issue_type
                    if issue_type == "有弱點的元件":
                        cleaned_entity_name = re.sub(r'\s*\([^)]*\)\s*$', '', entity_name).strip()
                        if cleaned_entity_name and cleaned_entity_name != 'N/A':
                            effective_vuln_name = f"{issue_type} {cleaned_entity_name}"

                    # 收集原始實例資料以供後續去重和篩選
                    instance_dict = {
                        "Source": APPSCAN_SOURCE_LABEL,
                        "Severity": issue.get("severity_display", "未知"),
                        "Vulnerability URL": issue.get("url", "N/A"),
                        "Vulnerability Entity": f"{issue.get('entity_name', 'N/A')} ({issue.get('entity_type', 'N/A')})",
                        "_severity_key": issue.get("severity_key", "unknown"), # 內部使用
                        "_report_num": report_num,                             # 內部使用
                        "_issue_id": issue_id,                                 # 內部使用
                        "_report_filename": report_filename,                   # 內部使用
                        "_original_issue_type": issue.get("issue_type", "N/A"),# 內部使用 (用於截圖查找)
                        "_original_entity_name": issue.get("entity_name", "N/A")# 內部使用 (用於截圖查找)
                    }
                    target_key = target_url if target_url != "N/A" else f"未指定目標 ({report_filename})"
                    temp_instances[target_key][effective_vuln_name].append(instance_dict)

    # 處理手動問題 (從狀態檔)
    for report_filename, report_data in project_statuses.items():
         if report_filename == REPORT_COMPLETED_KEY:
             continue
         target_info = filename_to_target_map.get(report_filename)
         # 如果狀態檔中的報告不在 target.xlsx 或找不到對應 XML，則無法確定標的 URL，給予預設值
         if not target_info:
              target_url = f"未知目標 ({report_filename})"
              report_num = "unknown"
              app.logger.warning(f"JSON Export: 找不到 {report_filename} 的目標資訊，將使用預設目標鍵。")
         else:
              target_url = target_info["url"]
              report_num = target_info["number"]

         for issue_id, status_data in report_data.items():
            # 跳過非手動、已處理或特殊鍵
            if not issue_id.startswith("_manual_") or issue_id == REPORT_COMPLETED_KEY or (report_filename, issue_id) in processed_issue_ids:
                continue

            # 檢查是否為確認的手動弱點
            if isinstance(status_data, dict) and status_data.get("status") == STATUS_OPTIONS["已確認弱點"]:
                manual_details = status_data.get("manual_details")
                if manual_details and isinstance(manual_details, dict):
                    processed_issue_ids.add((report_filename, issue_id)) # 標記為已處理

                    # 決定合併用的弱點名稱
                    issue_type = manual_details.get("issue_type", "N/A")
                    entity_name = manual_details.get("entity_name", "N/A")
                    effective_vuln_name = issue_type
                    if issue_type == "有弱點的元件":
                         cleaned_entity_name = re.sub(r'\s*\([^)]*\)\s*$', '', entity_name).strip()
                         if cleaned_entity_name and cleaned_entity_name != 'N/A':
                             effective_vuln_name = f"{issue_type} {cleaned_entity_name}"

                    # 收集原始實例資料
                    instance_dict = {
                        "Source": MANUAL_SOURCE_LABEL,
                        "Severity": manual_details.get("severity_display", "未知"),
                        "Vulnerability URL": manual_details.get("url", "N/A"),
                        "Vulnerability Entity": f"{manual_details.get('entity_name', 'N/A')} ({manual_details.get('entity_type', MANUAL_ENTITY_TYPE_LABEL)})",
                        "_severity_key": manual_details.get("severity_key", "medium"),
                        "_report_num": report_num,
                        "_issue_id": issue_id,
                        "_report_filename": report_filename,
                        "_original_issue_type": manual_details.get("issue_type", "N/A"),
                        "_original_entity_name": manual_details.get("entity_name", "N/A")
                    }
                    target_key = target_url if target_url != "N/A" else f"未指定目標 ({report_filename})"
                    temp_instances[target_key][effective_vuln_name].append(instance_dict)

    # --- 第二步：根據 URL+實體 篩選每個群組，保留最高嚴重性，並附加截圖 ---
    final_json_data = {}
    # 遍歷每個標的 URL
    for target_key, vuln_groups in temp_instances.items():
        target_vuln_list = []
        # 遍歷該標的下的每個弱點名稱群組
        for vuln_name, instances in vuln_groups.items():
            # dedup_instances 結構: {(url, entity): highest_severity_instance_dict}
            dedup_instances = {}
            # 第一次遍歷：找出每個 (URL, 實體) 組合中嚴重性最高的實例
            for instance in instances:
                # 使用弱點 URL 和弱點實體作為去重的鍵
                instance_key = (instance["Vulnerability URL"], instance["Vulnerability Entity"])
                current_level = SEVERITY_LEVELS.get(instance["_severity_key"], -1)

                # 如果是新的 URL+實體組合，或目前這筆嚴重性更高，則取代
                # Also pass along necessary original info for screenshot lookup
                if instance_key not in dedup_instances or \
                   current_level > SEVERITY_LEVELS.get(dedup_instances[instance_key]["_severity_key"], -1):
                    # 儲存這個最高嚴重性的實例，包含查找截圖需要的原始信息
                    dedup_instances[instance_key] = instance

            # --- 現在 dedup_instances 中儲存了每個 (URL, 實體) 組合的代表性實例 ---
            # --- 我們需要將這些代表性實例整理成最終的輸出格式 ---
            final_instances_list = []

            # 遍歷去重後的代表性實例
            for instance_key, final_instance in dedup_instances.items():
                # 獲取截圖 (基於最終保留的實例的資訊)
                report_num = final_instance.get("_report_num", "unknown")
                issue_id = final_instance.get("_issue_id")
                report_filename = final_instance.get("_report_filename")
                source_label = final_instance.get("Source") # 直接使用已存的 Source

                screenshots_meta = {}
                if report_filename and issue_id and report_filename in project_statuses and issue_id in project_statuses[report_filename]:
                     meta_candidate = project_statuses[report_filename][issue_id].get("screenshots_meta", {})
                     if isinstance(meta_candidate, dict):
                         screenshots_meta = meta_candidate

                screenshots_data = []
                if report_num != "unknown" and source_label:
                    # 使用儲存的原始弱點類型和實體名稱來查找截圖
                    original_issue_type = final_instance.get("_original_issue_type", "Unknown")
                    original_url = final_instance.get("Vulnerability URL", "N/A") # 用最終的URL
                    original_entity_name = final_instance.get("_original_entity_name", "N/A")

                    sanitized_name = sanitize_filename_part(original_issue_type)
                    sanitized_url = sanitize_filename_part(original_url, 30)
                    sanitized_entity = sanitize_filename_part(original_entity_name, 30)

                    # 檢查清理後的名稱是否有效，避免產生無效前綴
                    if sanitized_name != "na" and sanitized_name != "sanitized_empty":
                        screenshot_prefix = construct_screenshot_filename_prefix(report_num, source_label, sanitized_name, sanitized_url, sanitized_entity)
                        screenshot_files = get_existing_screenshots(project_name, screenshot_prefix)
                        screenshot_dir = get_project_screenshot_folder(project_name)
                        for ss_filename in screenshot_files:
                             ss_filepath = os.path.join(screenshot_dir, ss_filename)
                             if os.path.isfile(ss_filepath):
                                 encoded_string = ""
                                 caption = screenshots_meta.get(ss_filename, {}).get("caption", "")
                                 try:
                                     with open(ss_filepath, "rb") as image_file:
                                         encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                                     screenshots_data.append({ "Base64": encoded_string, "Caption": caption })
                                 except Exception as img_err:
                                     app.logger.error(f"JSON Export (Dedup): 讀取/編碼截圖時發生錯誤 {ss_filepath}: {img_err}")
                                     screenshots_data.append({ "Base64": None, "Caption": f"錯誤：無法載入 {ss_filename}: {img_err}" })
                    else:
                        app.logger.warning(f"JSON Export (Dedup): 無效的截圖查找名稱 (Issue: {issue_id}, Type: {original_issue_type})")


                # 創建最終要輸出的實例字典 (移除內部鍵)
                output_instance = {
                    "Source": final_instance["Source"],
                    "Severity": final_instance["Severity"],
                    "Vulnerability URL": final_instance["Vulnerability URL"],
                    "Vulnerability Entity": final_instance["Vulnerability Entity"],
                    "Screenshots": screenshots_data # 添加處理好的截圖列表
                }
                final_instances_list.append(output_instance)

            # 如果該弱點名稱下有保留的實例，則加入最終列表
            if final_instances_list:
                # 按 URL, Entity 排序實例列表 (可選)
                final_instances_list.sort(key=lambda x: (x["Vulnerability URL"], x["Vulnerability Entity"]))
                target_vuln_list.append({
                    "Vulnerability Name": vuln_name,
                    "Instances": final_instances_list
                })

        # 將此標的下的弱點按名稱排序
        final_json_data[target_key] = sorted(target_vuln_list, key=lambda x: x["Vulnerability Name"])

    # --- 準備 JSON 回應 ---
    try:
        json_output = json.dumps(final_json_data, ensure_ascii=False, indent=4)
        safe_display_name = secure_filename(display_name) or "project_export"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"{safe_display_name}_{timestamp}_confirmed_vulnerabilities_dedup.json" # 更新檔名

        app.logger.info(f"去重後的 JSON 匯出已產生 ({len(final_json_data)} 個標的)。正在發送: {download_filename}")
        return Response(
            json_output,
            mimetype="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=\"{download_filename}\"",
                "Content-Type": "application/json; charset=utf-8"
            }
        )
    except Exception as e:
        app.logger.error(f"產生/發送去重後 JSON 匯出時發生錯誤: {traceback.format_exc()}")
        flash(f"產生去重後 JSON 匯出時發生錯誤: {e}", "danger")
        return redirect(url_for("project_index", project_name=project_name))
# --- End Deduplicated JSON Export ---

# --- Selenium 驗證邏輯 ---
@app.route("/project/<project_name>/api/verify_vulnerability", methods=["POST"])
def verify_vulnerability_via_selenium(project_name):
    global selenium_driver_instance

    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.get_json()
    issue_url = data.get("issueUrl")
    entity_name = data.get("entityName")
    entity_type = data.get("entityType")
    reasoning = data.get("reasoning")
    component_name = data.get("componentName")
    component_version = data.get("componentVersion")

    target_url_to_view, search_term_primary, search_term_secondary = None, None, None
    is_external_link_issue = False

    if reasoning == EXTERNAL_LINK_REASONING and entity_type and "link" in entity_type.lower():
        is_external_link_issue = True
        target_url_to_view = issue_url
        if entity_name and entity_name.lower() != 'n/a':
            try:
                parsed_entity_url = urlparse(entity_name)
                search_term_primary = parsed_entity_url.netloc or entity_name
            except ValueError:
                search_term_primary = entity_name
        else:
            app.logger.warning("外部連結問題，但實體名稱遺失或為 N/A。")
    else:
        target_url_to_view = issue_url
        search_term_primary = component_version if component_version and component_version != "N/A" else None
        search_term_secondary = component_name if component_name and component_name != "N/A" else None
        if not search_term_primary and not search_term_secondary:
            search_term_primary = entity_name if entity_name and entity_name != "N/A" else None

    if not target_url_to_view or target_url_to_view.lower() == "n/a":
        return jsonify({"error": "無法確定目標 URL。"}), 400

    try:
        parsed_target = urlparse(target_url_to_view)
        if not parsed_target.scheme or parsed_target.scheme not in ["http", "https"]:
            return jsonify({"error": f"目標 URL '{target_url_to_view}' 無效。"}), 400
        if parsed_target.scheme == "file":
            return jsonify({"error": "不允許本地路徑。"}), 400
    except ValueError:
        return jsonify({"error": f"無法解析目標 URL '{target_url_to_view}'。"}), 400

    app.logger.info(f"Selenium 驗證 (等待跳轉, 不區分大小寫): 初始URL='{target_url_to_view}', 主要查找='{search_term_primary}', 次要查找='{search_term_secondary}'")
    driver = None
    page_load_timeout = 20

    try:
        with selenium_driver_lock:
            if selenium_driver_instance:
                try:
                    _ = selenium_driver_instance.current_url
                    app.logger.info("正在重用 Selenium driver。")
                    driver = selenium_driver_instance
                except WebDriverException as reuse_e:
                    app.logger.warning(f"現有的 driver 無法使用 ({reuse_e})，正在建立新的。")
                    try:
                        selenium_driver_instance.quit()
                    except Exception:
                        pass
                    selenium_driver_instance = None

            if driver is None:
                 chrome_options = webdriver.ChromeOptions()
                 chrome_options.add_argument("--ignore-certificate-errors")
                 chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
                 try:
                     if WEBDRIVER_MANAGER_AVAILABLE:
                         service = ChromeService(ChromeDriverManager().install())
                         driver = webdriver.Chrome(service=service, options=chrome_options)
                         app.logger.info("已透過 webdriver-manager 建立新的 ChromeDriver。")
                     else:
                         try:
                             driver = webdriver.Chrome(options=chrome_options)
                             app.logger.info("已透過系統 PATH 建立新的 ChromeDriver。")
                         except WebDriverException as path_e:
                             app.logger.error(f"找不到 ChromeDriver: {path_e}")
                             raise WebDriverException("無法啟動 Chrome。")
                     selenium_driver_instance = driver
                 except WebDriverException as e:
                     app.logger.error(f"WebDriver 設置錯誤: {e}")
                     err_msg = f"無法啟動瀏覽器：請確認 Chrome 已安裝。"
                     if WEBDRIVER_MANAGER_AVAILABLE:
                         err_msg += " 請確認 webdriver-manager 可連網。"
                     else:
                         err_msg += " 請確認 ChromeDriver 在 PATH。"
                     err_msg += f" (錯誤: {e})"
                     selenium_driver_instance = None
                     return jsonify({"error": err_msg}), 500
                 except Exception as setup_e:
                     app.logger.error(f"未預期的 WebDriver 設置錯誤: {setup_e}", exc_info=True)
                     selenium_driver_instance = None
                     return jsonify({"error": f"啟動瀏覽器出錯: {setup_e}"}), 500

            app.logger.info(f"Selenium 正在導航至目標頁面: {target_url_to_view}")
            try:
                driver.get(target_url_to_view)
                app.logger.info(f"等待頁面穩定 (最多 {page_load_timeout} 秒)...")
                WebDriverWait(driver, page_load_timeout).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )
                app.logger.info("頁面 readyState 為 complete。")
            except TimeoutException:
                final_url_on_timeout = driver.current_url
                app.logger.error(f"頁面在 {page_load_timeout} 秒內未能穩定 (readyState != 'complete')。當前 URL: {final_url_on_timeout}")
                # 即使超時，仍然繼續嘗試查看原始碼
            except WebDriverException as nav_e:
                app.logger.error(f"Selenium 導航至目標頁面失敗: {nav_e}")
                try:
                    driver.quit()
                except Exception:
                    pass
                selenium_driver_instance = None
                return jsonify({"error": f"無法載入目標頁面: {nav_e}"}), 500

            final_url = driver.current_url
            app.logger.info(f"頁面穩定後的最終 URL: {final_url}")
            view_source_final_url = f"view-source:{final_url}"
            app.logger.info(f"Selenium 正在導航至原始碼頁面: {view_source_final_url}")
            try:
                driver.get(view_source_final_url)
                time.sleep(0.5) # 給 view-source 一點載入時間
            except WebDriverException as vs_nav_e:
                app.logger.error(f"Selenium 導航至 view-source 失敗: {vs_nav_e}")
                # 不退出 driver，因為瀏覽器可能還在，只是 view-source 打不開
                return jsonify({"error": f"無法載入最終頁面的原始碼: {vs_nav_e}"}), 500

            found = False
            found_term_used = None
            page_source_lower = None

            if search_term_primary:
                escaped_search_term = search_term_primary.replace('\'', '\\\'')
                # 使用不區分大小寫查找
                js_script_find = f"return window.find('{escaped_search_term}', false, false, true, false, true, false);"
                try:
                    if driver.execute_script(js_script_find):
                        found = True
                        found_term_used = search_term_primary
                        app.logger.info(f"透過 find() 找到主要字串(不區分大小寫) '{search_term_primary}'。")
                    else:
                        # 備用：檢查原始碼（轉小寫）
                        if page_source_lower is None:
                            page_source_lower = driver.page_source.lower()
                        if search_term_primary.lower() in page_source_lower:
                            found = True
                            found_term_used = search_term_primary
                            app.logger.info(f"在 page_source 中找到主要字串(不區分大小寫) '{search_term_primary}'。")
                except Exception as find_e:
                    app.logger.warning(f"查找主要字串 '{search_term_primary}' 時發生錯誤: {find_e}")

            if not found and search_term_secondary:
                escaped_search_term_secondary = search_term_secondary.replace('\'', '\\\'')
                js_script_find_sec = f"return window.find('{escaped_search_term_secondary}', false, false, true, false, true, false);"
                try:
                    if driver.execute_script(js_script_find_sec):
                        found = True
                        found_term_used = search_term_secondary
                        app.logger.info(f"透過 find() 找到次要字串(不區分大小寫) '{search_term_secondary}'。")
                    else:
                        if page_source_lower is None:
                            page_source_lower = driver.page_source.lower()
                        if search_term_secondary.lower() in page_source_lower:
                            found = True
                            found_term_used = search_term_secondary
                            app.logger.info(f"在 page_source 中找到次要字串(不區分大小寫) '{search_term_secondary}'。")
                except Exception as find_e:
                    app.logger.warning(f"查找次要字串 '{search_term_secondary}' 時發生錯誤: {find_e}")

            if found:
                message = f"已在最終頁面 ({final_url}) 的原始碼中找到(不區分大小寫) '{found_term_used}'。"
                app.logger.info(message)
                try:
                    driver.switch_to.window(driver.current_window_handle) # 嘗試切換回視窗焦點
                except Exception:
                    pass
                return jsonify({"message": message, "status": "found"}), 200
            else:
                search_terms_tried = [f"'{t}'" for t in [search_term_primary, search_term_secondary] if t]
                terms_str = " 或 ".join(search_terms_tried) or "指定字串"
                message = f"無法在最終頁面 ({final_url}) 的原始碼中自動定位 {terms_str} (不區分大小寫)。已開啟原始碼頁面供手動檢閱。"
                app.logger.warning(f"無法在 {final_url} 的原始碼中找到 {terms_str} (不區分大小寫)")
                try:
                    driver.switch_to.window(driver.current_window_handle) # 嘗試切換回視窗焦點
                except Exception:
                    pass
                return jsonify({"message": message, "status": "not_found"}), 200

    except WebDriverException as e:
        app.logger.error(f"驗證時發生 WebDriverException: {e}", exc_info=True)
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        with selenium_driver_lock:
            selenium_driver_instance = None
        return jsonify({"error": f"瀏覽器操作失敗: {e}"}), 500
    except Exception as e:
        app.logger.error(f"驗證時發生未預期錯誤: {e}", exc_info=True)
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        with selenium_driver_lock:
            selenium_driver_instance = None
        return jsonify({"error": f"驗證時出錯: {e}"}), 500


# API 路由：新增自訂弱點
@app.route("/project/<project_name>/api/add_custom_vulnerability", methods=["POST"])
def add_custom_vulnerability(project_name):
    """處理前端發送的新增手動弱點請求"""
    if not is_safe_project_name(project_name):
        return jsonify({"error": "無效的專案名稱"}), 400

    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.get_json()
    report_filename = data.get("reportFilename")
    issue_name = data.get("issueName", "").strip()
    severity_key = data.get("severity", "medium").lower()
    url = data.get("url", "").strip() or "N/A"
    entity_name = data.get("entityName", "").strip() or "N/A"
    note = data.get("note", "").strip()

    if not report_filename:
        return jsonify({"error": "缺少報告檔案名稱。"}), 400
    if not issue_name:
        return jsonify({"error": "弱點名稱為必填。"}), 400
    if severity_key not in SEVERITY_LEVELS:
        return jsonify({"error": f"無效的嚴重性等級: {severity_key}"}), 400

    manual_id = f"_manual_{uuid.uuid4().hex[:12]}"

    try:
        statuses = load_statuses(project_name)
        report_entry = statuses.setdefault(report_filename, {})

        while manual_id in report_entry:
            app.logger.warning(f"手動弱點 ID 衝突: {manual_id}。正在重新產生...")
            manual_id = f"_manual_{uuid.uuid4().hex[:12]}"

        new_vuln_data = {
            "status": DEFAULT_STATUS,
            "screenshot_taken": False,
            "note": note,
            "source": "manual",
            "screenshots_meta": {}, # Initialize meta
            "manual_details": {
                "issue_type": issue_name,
                "severity_key": severity_key,
                "url": url,
                "entity_name": entity_name,
                "entity_type": MANUAL_ENTITY_TYPE_LABEL,
                "reasoning": "(手動新增)",
                "severity_display": SEVERITY_DISPLAY_MAP.get(severity_key, severity_key.capitalize()),
            }
        }
        report_entry[manual_id] = new_vuln_data
        save_statuses(project_name, statuses)
        app.logger.info(f"已將手動弱點 '{issue_name}' ({manual_id}) 新增至報告 '{report_filename}'")

        display_issue = {
            "id": manual_id,
            "status": new_vuln_data["status"],
            "screenshot_taken": new_vuln_data["screenshot_taken"],
            "note": new_vuln_data["note"],
            "source": new_vuln_data["source"],
            **new_vuln_data["manual_details"],
            "screenshots_detailed": [],
            "cvss_score": "N/A",
            "cve_name": None,
            "cve_url": None,
            "http_traffic": "N/A",
            "scan_info": {}
        }
        return jsonify({"message": "自訂弱點已成功新增！", "new_issue": display_issue}), 201

    except Exception as e:
        app.logger.error(f"新增自訂弱點時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "新增自訂弱點時出錯。"}), 500


# --- 伺服器關閉相關函式 ---
def shutdown_server():
    """嘗試呼叫 Werkzeug 內建的關閉函式"""
    func = request.environ.get("werkzeug.server.shutdown")
    if func is None:
        app.logger.error("找不到 Werkzeug 關閉函式。")
        return False

    try:
        func() # 執行關閉函式
        app.logger.info("已呼叫 Werkzeug 關閉函式。")
        return True
    except Exception as e:
        app.logger.error(f"呼叫關閉函式時發生錯誤: {e}")
        return False


# API 路由：關閉伺服器 (從 GUI 或信號觸發)
@app.route("/shutdown", methods=["POST"])
def shutdown():
    """接收關閉請求並嘗試關閉伺服器"""
    app.logger.info("收到 POST /shutdown 請求。")
    if shutdown_server():
        return "伺服器正在關閉..."
    else:
        return "無法觸發關閉。", 500


# --- Flask 伺服器啟動函式 ---
def run_flask_app():
    """在單獨的線程中運行 Flask 應用程式"""
    global server_running, server_port, status_window_root # 使用全域變數

    try:
        load_server_config() # 載入伺服器設定 (Port)
        app.logger.info(f"正在啟動 Flask 伺服器於 http://0.0.0.0:{server_port}/")
        server_running = True # 標記伺服器正在運行
        # 啟動 Flask 伺服器
        app.run(host="0.0.0.0", port=server_port, debug=False, use_reloader=False, threaded=True)
        # --- 當 app.run() 結束後 (例如被 shutdown) ---
        app.logger.info("Flask app.run() 已結束。")
    except OSError as e:
        # 處理作業系統錯誤，最常見的是 Port 被佔用
        server_running = False # 標記伺服器未運行
        error_message = f"作業系統錯誤: {e}"
        if "Address already in use" in str(e) or "Only one usage" in str(e):
            error_message = f"啟動失敗：Port {server_port} 已被佔用。"
        app.logger.error(error_message)
        # 如果 GUI 視窗存在，顯示錯誤訊息框
        if status_window_root and status_window_root.winfo_exists():
            status_window_root.after(0, lambda: messagebox.showerror("啟動錯誤", error_message, parent=status_window_root))
        else:
            # 否則直接印到控制台
            print(f"錯誤: {error_message}", file=sys.stderr)
    except Exception as e:
        # 處理其他未預期啟動錯誤
        server_running = False # 標記伺服器未運行
        error_message = f"啟動 Flask 時發生未知錯誤: {e}"
        app.logger.error(f"未知的啟動錯誤: {traceback.format_exc()}")
        # 同上，嘗試在 GUI 顯示錯誤
        if status_window_root and status_window_root.winfo_exists():
            status_window_root.after(0, lambda: messagebox.showerror("啟動錯誤", error_message, parent=status_window_root))
        else:
            print(f"錯誤: {error_message}", file=sys.stderr)
    finally:
        # 無論成功或失敗，最後標記伺服器為停止狀態
        app.logger.info(f"Flask 線程結束 (運行狀態={server_running})。")
        server_running = False


# --- Tkinter GUI 類別 (Tkinter GUI Classes) ---
class ServerControlGUI:
    """伺服器控制台 GUI 類別"""
    def __init__(self, master):
        """初始化伺服器控制台 GUI"""
        self.master = master # Tkinter 的根視窗
        master.title("AppScan-Report-Manager") # 視窗標題
        master.geometry("650x580") # 設定視窗大小
        master.protocol("WM_DELETE_WINDOW", self.on_closing) # 設定關閉視窗按鈕的回呼函式

        # --- 設定視窗圖示 ---
        try:
            icon_path = resource_path("icon.ico") # 使用 resource_path
            if os.path.exists(icon_path):
                master.iconbitmap(icon_path)
                app.logger.info(f"已成功載入圖示: {icon_path}")
            else:
                app.logger.warning(f"找不到圖示檔案: {icon_path}")
        except Exception as e:
            app.logger.error(f"載入圖示時發生錯誤: {e}")

        # --- 設定 GUI 樣式 ---
        style = ttk.Style()
        try:
            style.theme_use("clam") # 嘗試使用 clam 主題
        except tk.TclError:
            app.logger.warning("無法使用 Clam 主題。")

        master.config(bg="black") # 設定背景色為黑色
        style.configure("TLabel", background="black", foreground="white", font=("Segoe UI", 11))
        style.configure("TButton", font=("Segoe UI", 10), padding=5)
        style.map("TButton", background=[('active', '#444444'), ('disabled', '#333333')], foreground=[('disabled', '#888888')])
        style.configure("Status.TLabel", font=("Segoe UI", 12, "bold"))

        # --- 建立 GUI 元件 ---
        self.status_label = ttk.Label(master, text="伺服器狀態：檢查中...", style="Status.TLabel", anchor="center")
        self.status_label.pack(pady=15, fill=tk.X, padx=10)

        self.open_ui_button = ttk.Button(master, text="開啟使用者介面 (瀏覽器)", command=self.open_browser_ui, state=tk.DISABLED)
        self.open_ui_button.pack(pady=(0, 10))

        self.log_text = scrolledtext.ScrolledText(master, wrap=tk.WORD, height=20, font=("Consolas", 9), bg="#1e1e1e", fg="#d4d4d4", insertbackground="white", state="disabled", relief=tk.FLAT, borderwidth=0)
        self.log_text.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)

        # 設定不同日誌級別的顏色標籤
        self.log_text.tag_config("INFO", foreground="#cccccc")
        self.log_text.tag_config("WARNING", foreground="#ffd700")
        self.log_text.tag_config("ERROR", foreground="#f44747")
        self.log_text.tag_config("CRITICAL", foreground="#ff6a6a", font=("Consolas", 9, "bold"))
        self.log_text.tag_config("DEBUG", foreground="#888888")

        self.shutdown_button = ttk.Button(master, text="關閉伺服器並離開程式", command=self.handle_shutdown_button_click, state=tk.DISABLED)
        self.shutdown_button.pack(pady=15)

        # 啟動定時更新 GUI 的函式
        self.master.after(100, self.update_status_label) # 更新狀態標籤
        self.master.after(100, self.update_log)          # 更新日誌區域


    def update_status_label(self):
        """定時更新伺服器狀態標籤和按鈕狀態"""
        if not self.master.winfo_exists():
            return # 如果視窗已關閉，則停止

        try:
            is_running = server_running # 獲取伺服器運行狀態
            status_text = f"伺服器狀態：運行中 (Port: {server_port})" if is_running else "伺服器狀態：已停止"
            status_color = "lime green" if is_running else "tomato"
            self.status_label.config(text=status_text, foreground=status_color)

            target_shutdown_state = 'disabled' if not is_running else '!disabled' # '!' 表示啟用
            if target_shutdown_state not in self.shutdown_button.state():
                self.shutdown_button.state([target_shutdown_state])

            target_open_state = 'disabled' if not is_running else '!disabled'
            if target_open_state not in self.open_ui_button.state():
                self.open_ui_button.state([target_open_state])

            # 設定下一次更新 (1 秒後)
            self.master.after(1000, self.update_status_label)
        except tk.TclError as e:
            app.logger.warning(f"更新 GUI 狀態時發生 TclError: {e}")
        except Exception as e:
            app.logger.error(f"更新 GUI 狀態時發生錯誤: {e}")
            self.master.after(5000, self.update_status_label) # 出錯後延遲更長時間再試


    def update_log(self):
        """定時從佇列中讀取日誌並顯示在文字框中"""
        if not self.master.winfo_exists():
            return # 視窗關閉則停止

        try:
            processed_count = 0 # 記錄本次處理的日誌數量
            max_logs_per_cycle = 50 # 每次最多處理 50 條，避免卡頓
            while status_window_log_queue and processed_count < max_logs_per_cycle:
                try:
                    level, record_text = status_window_log_queue.pop(0)
                except IndexError:
                    break # 佇列為空則跳出

                processed_count += 1
                if not self.log_text.winfo_exists():
                    break # 文字框不存在則跳出

                self.log_text.config(state="normal") # 允許編輯
                log_tag = level if level in ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"] else "INFO"
                self.log_text.insert(tk.END, record_text, (log_tag,)) # 插入文字並應用標籤
                self.log_text.config(state="disabled") # 禁止編輯

            if processed_count > 0 and self.log_text.winfo_exists():
                self.log_text.see(tk.END) # 滾動到最下方

            # 設定下一次更新 (0.25 秒後)
            self.master.after(250, self.update_log)
        except tk.TclError as e:
            app.logger.warning(f"更新 GUI 日誌時發生 TclError: {e}")
        except Exception as e:
            app.logger.error(f"更新 GUI 日誌時發生錯誤: {e}")
            self.master.after(1000, self.update_log) # 出錯後延遲更新


    def open_browser_ui(self):
        """開啟預設瀏覽器訪問應用程式首頁"""
        global server_running, server_port # 使用全域變數

        if server_running:
            url = f"http://127.0.0.1:{server_port}/" # 構造 URL
            app.logger.info(f"正在開啟瀏覽器: {url}")
            try:
                webbrowser.open(url, new=2) # 開啟新分頁
            except Exception as e:
                error_message = f"無法自動開啟瀏覽器：\n{e}\n\n請手動訪問:\n{url}"
                app.logger.error(f"無法開啟瀏覽器: {e}")
                parent = self.master if self.master.winfo_exists() else None
                messagebox.showerror("錯誤", error_message, parent=parent)
        else:
            app.logger.warning("無法開啟 UI: 伺服器未運行。")
            parent = self.master if self.master.winfo_exists() else None
            messagebox.showwarning("未運行", "伺服器未運行。", parent=parent)


    def _request_server_shutdown(self):
        """向 Flask 伺服器發送關閉請求"""
        global server_running, server_port # 使用全域變數

        if not server_running:
            app.logger.info("忽略關閉請求: 伺服器未運行。")
            return True # 直接返回成功

        shutdown_url = f"http://127.0.0.1:{server_port}/shutdown" # 關閉 API 的 URL
        app.logger.info(f"正在發送關閉請求至: {shutdown_url}")

        # 禁用按鈕防止重複點擊
        if self.master.winfo_exists():
            try:
                if self.shutdown_button.winfo_exists():
                    self.shutdown_button.state(["disabled"])
                if self.open_ui_button.winfo_exists():
                    self.open_ui_button.state(["disabled"])
            except tk.TclError:
                pass

        try:
            request_obj = urllib.request.Request(shutdown_url, method="POST")
            with urllib.request.urlopen(request_obj, timeout=2) as response: # 設定超時
                app.logger.info(f"關閉請求已發送。回應狀態: {response.status}")
            return True
        except urllib.error.URLError as e:
            # 如果請求失敗 (可能是伺服器已關閉)，則更新狀態並返回成功
            app.logger.warning(f"關閉 URL 錯誤 (伺服器已停止?): {e}")
            server_running = False # 更新伺服器狀態
            if self.master.winfo_exists():
                self.master.after(0, self.update_status_label) # 立即更新 GUI 狀態
            return True
        except Exception as e:
            # 其他錯誤，顯示錯誤訊息框
            app.logger.error(f"發送關閉請求時發生錯誤: {e}")
            parent = self.master if self.master.winfo_exists() else None
            messagebox.showerror("關閉錯誤", f"發送關閉請求時出錯：\n{e}", parent=parent)
            # 如果伺服器仍在運行，重新啟用按鈕
            if server_running and self.master.winfo_exists():
                try:
                    if self.shutdown_button.winfo_exists():
                        self.shutdown_button.state(["!disabled"])
                    if self.open_ui_button.winfo_exists():
                        self.open_ui_button.state(["!disabled"])
                except tk.TclError:
                    pass
            return False # 返回失敗


    def handle_shutdown_button_click(self):
        """處理關閉按鈕的點擊事件"""
        app.logger.info("透過 GUI 按鈕觸發關閉。")
        self._request_server_shutdown() # 發送關閉請求
        # 稍後關閉 GUI 視窗
        if self.master.winfo_exists():
            self.master.after(500, self._destroy_master)


    def on_closing(self):
        """處理點擊視窗關閉按鈕 (X) 的事件"""
        app.logger.debug("控制台視窗關閉按鈕被點擊。")
        if server_running: # 如果伺服器仍在運行
            parent = self.master if self.master.winfo_exists() else None
            confirm_exit = messagebox.askokcancel(
                "確認退出",
                "伺服器仍在運行中。\n\n確定要關閉伺服器並退出程式嗎？",
                parent=parent
            )
            if confirm_exit: # 如果用戶確認
                app.logger.info("使用者確認關閉。")
                self._request_server_shutdown() # 發送關閉請求
                if self.master.winfo_exists():
                    self.master.after(500, self._destroy_master) # 稍後關閉視窗
            else:
                 app.logger.info("使用者取消退出。") # 用戶取消則不執行任何操作
        else: # 如果伺服器已停止
            app.logger.info("關閉控制台視窗 (伺服器已停止)。")
            self._destroy_master() # 直接關閉視窗


    def _destroy_master(self):
        """安全地銷毀 Tkinter 根視窗"""
        global status_window_root # 使用全域視窗物件

        if status_window_root and status_window_root.winfo_exists():
            try:
                app.logger.info("正在銷毀 Tkinter 主視窗...")
                status_window_root.destroy() # 銷毀視窗
                app.logger.info("Tkinter 視窗已銷毀。")
            except Exception as e:
                app.logger.error(f"銷毀 Tkinter 視窗時發生錯誤: {e}")
            finally:
                status_window_root = None # 清空引用
        else:
            app.logger.info("Tkinter 視窗已銷毀或不存在。")


# --- 初始設定 GUI 函式 ---
def get_initial_config():
    """顯示初始設定視窗 (主要用於設定 Port)，並啟動 Flask 伺服器"""
    global server_port, server_thread, server_running # 使用全域變數

    config_window_cancelled = True # 標記視窗是否被取消
    config_root = tk.Tk() # 創建 Tkinter 根視窗
    config_root.title("伺服器啟動設定") # 設定標題
    config_root.geometry("350x180") # 設定大小
    config_root.resizable(False, False) # 禁止調整大小

    # --- 設定圖示 ---
    try:
        icon_path = resource_path("icon.ico")
        if os.path.exists(icon_path):
            config_root.iconbitmap(icon_path)
    except Exception:
        pass # 出錯則忽略

    # --- 設定樣式 ---
    try:
        ttk.Style(config_root).theme_use("clam")
    except tk.TclError:
        pass

    # --- 創建元件 ---
    ttk.Label(config_root, text="伺服器 Port:", font=("Segoe UI", 10)).grid(row=0, column=0, padx=(15, 5), pady=20, sticky="w")
    port_var = tk.StringVar(value=str(server_port)) # 綁定變數
    port_entry = ttk.Entry(config_root, textvariable=port_var, width=10, font=("Segoe UI", 10))
    port_entry.grid(row=0, column=1, padx=(0, 15), pady=20, sticky="w")
    port_entry.focus() # 預設焦點
    status_var = tk.StringVar(value="")
    status_label = ttk.Label(config_root, textvariable=status_var, foreground="red", wraplength=320, font=("Segoe UI", 9))
    status_label.grid(row=1, column=0, columnspan=2, padx=15, pady=(10, 5), sticky="ew")
    start_button = None # 初始化啟動按鈕變數

    # --- 啟動伺服器按鈕的回呼函式 ---
    def start_server_action():
        nonlocal start_button, config_window_cancelled, config_root # 使用外層變數
        global server_port, server_thread, server_running # 使用全域變數

        if start_button: # 禁用按鈕，顯示狀態
            start_button.state(["disabled"])
            status_var.set("正在檢查設定並啟動伺服器...")
            config_root.update_idletasks() # 更新介面

        try:
            port_input = int(port_var.get()) # 獲取輸入的 Port
            if not (1024 <= port_input <= 65535): # 驗證 Port 範圍
                raise ValueError("Port 需介於 1024 - 65535。")

            server_port = port_input # 更新全域 Port
            save_server_config() # 儲存設定
            status_var.set("") # 清空狀態訊息

            # 確保基礎資料夾存在
            for folder_path in [BASE_REPORT_FOLDER, BASE_DATA_FOLDER]:
                folder_path = resource_path(folder_path) # 使用 resource_path
                if not os.path.isdir(folder_path):
                    os.makedirs(folder_path, exist_ok=True)
                    app.logger.info(f"已建立資料夾: '{os.path.abspath(folder_path)}'")

            # --- 啟動 Flask 線程 ---
            app.logger.info("正在啟動 Flask 線程...")
            server_thread = threading.Thread(target=run_flask_app, daemon=True) # 設置為守護線程
            server_thread.start()
            time.sleep(1.5) # 等待伺服器啟動
            if not server_running: # 檢查伺服器是否成功啟動
                raise RuntimeError("伺服器啟動失敗。請檢查控制台記錄。")

            config_window_cancelled = False # 標記為未取消
            app.logger.info(f"Flask 伺服器已啟動於 Port {server_port}。")
            status_var.set(f"伺服器已啟動！") # 顯示成功訊息
            status_label.config(foreground="green")
            config_root.title("伺服器設定 (運行中)") # 更新標題
            config_root.after(1500, config_root.destroy) # 延遲關閉設定視窗

        except (ValueError, RuntimeError) as e:
            # 處理 Port 錯誤或啟動失敗
            error_message = f"啟動錯誤: {e}"
            status_var.set(error_message)
            status_label.config(foreground="red")
            app.logger.error(f"啟動失敗: {e}")
            # 重新啟用按鈕
            if config_root.winfo_exists() and start_button:
                try:
                    start_button.state(["!disabled"])
                except tk.TclError:
                    pass
        except Exception as ex:
            # 處理其他未預期錯誤
            error_message = f"未預期錯誤: {ex}"
            status_var.set(error_message)
            status_label.config(foreground="red")
            app.logger.error(f"未預期的啟動錯誤: {traceback.format_exc()}")
            # 重新啟用按鈕
            if config_root.winfo_exists() and start_button:
                try:
                    start_button.state(["!disabled"])
                except tk.TclError:
                    pass

    # 創建啟動按鈕
    start_button = ttk.Button(config_root, text="啟動伺服器", command=start_server_action, style="Accent.TButton")
    try: # 嘗試設定強調按鈕樣式
        ttk.Style().configure("Accent.TButton", font=("Segoe UI", 10, "bold"), foreground="white", background="#0078D4")
    except tk.TclError: # 若失敗則使用預設按鈕樣式
        ttk.Style().configure("TButton", font=("Segoe UI", 10))

    start_button.grid(row=2, column=0, columnspan=2, pady=15)
    # 綁定 Enter 鍵到啟動按鈕
    port_entry.bind("<Return>", lambda event=None: start_server_action())

    # --- 設定視窗關閉行為 ---
    def on_config_window_close():
        nonlocal config_window_cancelled, config_root
        if config_window_cancelled: # 如果在啟動前關閉
            app.logger.info("設定視窗在啟動前被關閉。")
            try:
                config_root.destroy() # 銷毀視窗
            except tk.TclError:
                pass

    config_root.protocol("WM_DELETE_WINDOW", on_config_window_close) # 綁定關閉按鈕事件

    # --- 將視窗置中並顯示 ---
    config_root.update_idletasks() # 更新視窗尺寸
    ws = config_root.winfo_screenwidth()
    hs = config_root.winfo_screenheight()
    w = config_root.winfo_width()
    h = config_root.winfo_height()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    config_root.geometry('%dx%d+%d+%d' % (w, h, int(x), int(y))) # 設定位置
    config_root.lift() # 移到最上層
    config_root.attributes("-topmost", True) # 保持在最上層
    config_root.after(100, lambda: config_root.attributes("-topmost", False)) # 短暫延遲後取消置頂
    config_root.mainloop() # 進入 Tkinter 事件循環
    # 返回是否成功啟動 (未被取消)
    return not config_window_cancelled


# --- 主執行區塊 ---
if __name__ == "__main__":
    # --- 設定信號處理函式 (用於 Ctrl+C 或終止信號) ---
    def signal_handler(sig, frame):
        print(f"\n接收到信號 {signal.Signals(sig).name} ({sig}). 正在關閉...")
        app.logger.warning(f"偵測到信號 {sig}。開始關閉程序...")
        global server_running, status_window_root, server_thread, server_port, selenium_driver_instance

        # --- 嘗試透過 HTTP 請求關閉 Flask 伺服器 ---
        if server_running:
            app.logger.info("信號處理: 嘗試 HTTP 關閉...")
            shutdown_url = f"http://127.0.0.1:{server_port}/shutdown"
            try:
                req = urllib.request.Request(shutdown_url, method="POST")
                urllib.request.urlopen(req, timeout=1.5) # 短暫超時
                time.sleep(0.5) # 給伺服器一點關閉時間
            except Exception as e:
                app.logger.error(f"信號處理: HTTP 關閉失敗: {e}")
                server_running = False # 標記伺服器已停止
        else:
            app.logger.info("信號處理: 伺服器未運行。")

        # --- 關閉 Selenium ---
        app.logger.info("信號處理: 正在檢查 Selenium...")
        with selenium_driver_lock:
            if selenium_driver_instance:
                app.logger.info("信號處理: 正在關閉 Selenium...")
                try:
                    selenium_driver_instance.quit()
                except Exception as e:
                    app.logger.error(f"信號處理: 關閉 Selenium 時發生錯誤: {e}")
                finally:
                    selenium_driver_instance = None # 清空引用
            else:
                app.logger.info("信號處理: 沒有 Selenium driver 實例。")

        # --- 關閉 Tkinter GUI ---
        if status_window_root and status_window_root.winfo_exists():
            app.logger.info("信號處理: 正在安排銷毀 Tkinter 視窗。")
            try:
                # 使用 after(0, ...) 在主線程中安全地銷毀視窗
                status_window_root.after(0, status_window_root.destroy)
            except Exception as e:
                app.logger.error(f"信號處理: 安排銷毀 Tkinter 時發生錯誤: {e}")

        # --- 等待 Flask 線程結束 ---
        if server_thread and server_thread.is_alive():
            app.logger.info("信號處理: 正在等待 Flask 線程...")
            server_thread.join(timeout=2.0) # 等待最多 2 秒
            if server_thread.is_alive():
                app.logger.warning("信號處理: Flask 線程未在超時內結束。")
            else:
                app.logger.info("信號處理: Flask 線程已結束。")
        else:
            app.logger.info("信號處理: Flask 線程未運行或已結束。")

        # --- 強制退出 ---
        app.logger.info("信號處理: 透過 os._exit 退出。")
        os._exit(1) # 使用 os._exit 強制退出，避免 Tkinter 在非主線程關閉時可能的問題

    # 註冊信號處理器
    signal.signal(signal.SIGINT, signal_handler) # Ctrl+C
    try:
        signal.signal(signal.SIGTERM, signal_handler) # 終止信號 (Linux/macOS)
    except AttributeError:
        # Windows 不支援 SIGTERM
        app.logger.warning("無法註冊 SIGTERM 處理器 (可能為 Windows 系統)。")

    # --- 應用程式啟動流程 ---
    app.logger.info("================ 應用程式啟動 ================")
    load_server_config() # 載入伺服器 Port 設定
    load_rules() # 載入全域排除規則
    load_common_notes() # <<< 新增：預先載入常用備註

    server_started_successfully = get_initial_config() # 顯示初始設定視窗並啟動伺服器
    if not server_started_successfully:
        # 如果啟動被取消或失敗
        app.logger.warning("啟動被取消或失敗。正在退出。")
        sys.exit(1)

    # --- 如果伺服器成功啟動，則啟動主 GUI ---
    app.logger.info(f"Flask 伺服器已啟動於 Port {server_port}。正在啟動 GUI...")
    try:
        status_window_root = tk.Tk() # 創建主 GUI 視窗
        gui = ServerControlGUI(status_window_root) # 實例化 GUI 類別
        # --- 將主 GUI 視窗置中 ---
        status_window_root.update_idletasks() # 更新視窗尺寸
        ws = status_window_root.winfo_screenwidth()
        hs = status_window_root.winfo_screenheight()
        w = status_window_root.winfo_width()
        h = status_window_root.winfo_height()
        x = (ws/2) - (w/2)
        y = (hs/2) - (h/2)
        status_window_root.geometry('%dx%d+%d+%d' % (w, h, int(x), int(y))) # 設定位置
        status_window_root.lift() # 移到最上層
        status_window_root.attributes("-topmost", True) # 保持在最上層
        status_window_root.after(100, lambda: status_window_root.attributes("-topmost", False)) # 取消置頂
        # --- 進入 Tkinter 主事件循環 ---
        status_window_root.mainloop()
        # --- 當主視窗關閉後 ---
        app.logger.info("Tkinter mainloop 已結束。")
    except Exception as e:
        # 處理 GUI 運行時的嚴重錯誤
        app.logger.error(f"嚴重的 Tkinter GUI 錯誤: {traceback.format_exc()}")
        if status_window_root and status_window_root.winfo_exists():
            try:
                status_window_root.destroy() # 嘗試銷毀視窗
            except:
                pass
        if server_running: # 如果伺服器仍在運行，嘗試緊急關閉
            app.logger.warning("嘗試緊急關閉伺服器...")
            signal_handler(signal.SIGTERM, None) # 觸發信號處理
        sys.exit(1) # 退出程式

    # --- GUI 關閉後的清理工作 ---
    app.logger.info("正在執行 GUI 後的清理工作...")
    # 關閉 Selenium
    with selenium_driver_lock:
        if selenium_driver_instance:
            app.logger.info("GUI 後清理: 正在關閉 Selenium...")
            try:
                selenium_driver_instance.quit()
            except Exception as e:
                app.logger.error(f"GUI 後清理: 關閉 Selenium 時發生錯誤: {e}")
            finally:
                selenium_driver_instance = None
    # 等待 Flask 線程結束
    if server_thread and server_thread.is_alive():
        app.logger.info("GUI 後清理: 正在等待 Flask 線程...")
        server_thread.join(timeout=2.0)
        if server_thread.is_alive():
            app.logger.warning("GUI 後清理: Flask 線程未在超時內結束。")
        else:
            app.logger.info("GUI 後清理: Flask 線程已結束。")
    # 最終確認伺服器狀態
    if server_running:
        app.logger.warning("清理後伺服器標記仍為運行中。")
        server_running = False

    app.logger.info("================ 應用程式正常退出 ================")
    sys.exit(0) # 正常退出
